#!/usr/bin/env python3
"""Rebuild the Risk Identification Template Pack with premium formatting."""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, Protection, NamedStyle
)
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PrintPageSetup
from copy import copy

# ── Brand Constants ──
NAVY = '1B2A4A'
GOLD = 'C5A55A'
TEAL = '2A7F7F'
GREY = '666666'
GREEN = '2D8B2D'
WHITE = 'FFFFFF'
OFFWHITE = 'F8F7F4'
INPUT_YELLOW = 'F5EDD8'
LIGHT_TEAL = 'D4ECEC'
LIGHT_GREY = 'E8E8E8'
TEXT = '333333'
TEXT_LIGHT = '5A5A5A'
RED_BG = 'F4CCCC'
AMBER_BG = 'FCE5CD'
GREEN_BG = 'D9EAD3'
RED_DARK = 'CC0000'
AMBER_DARK = 'E69138'
GREEN_DARK = '6AA84F'

# ── Reusable Styles ──
FONT_TITLE = Font(name='Calibri', size=14, bold=True, color=NAVY)
FONT_SUBTITLE = Font(name='Calibri', size=9, color=GREY, italic=True)
FONT_INSTRUCTION = Font(name='Calibri', size=9, color=GREY)
FONT_SECTION = Font(name='Calibri', size=12, bold=True, color=TEAL)
FONT_HEADER = Font(name='Calibri', size=11, bold=True, color=WHITE)
FONT_DATA = Font(name='Calibri', size=10, color=TEXT)
FONT_DATA_LIGHT = Font(name='Calibri', size=10, color=TEXT_LIGHT)
FONT_EXAMPLE = Font(name='Calibri', size=10, color=TEXT, italic=True)

FILL_NAVY = PatternFill('solid', fgColor=NAVY)
FILL_TEAL_LIGHT = PatternFill('solid', fgColor=LIGHT_TEAL)
FILL_INPUT = PatternFill('solid', fgColor=INPUT_YELLOW)
FILL_GREY = PatternFill('solid', fgColor=LIGHT_GREY)
FILL_WHITE = PatternFill('solid', fgColor=WHITE)
FILL_OFFWHITE = PatternFill('solid', fgColor=OFFWHITE)
FILL_RED = PatternFill('solid', fgColor=RED_BG)
FILL_AMBER = PatternFill('solid', fgColor=AMBER_BG)
FILL_GREEN = PatternFill('solid', fgColor=GREEN_BG)

ALIGN_WRAP = Alignment(wrap_text=True, vertical='top')
ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_HEADER = Alignment(horizontal='center', vertical='center', wrap_text=True)

THIN_BORDER = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0')
)


def set_col_widths(ws, widths):
    """Set column widths from a dict {col_letter: width}."""
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def write_header_row(ws, row, headers, start_col=1):
    """Write a navy header row with white text."""
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=start_col + i, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_NAVY
        cell.alignment = ALIGN_HEADER
        cell.border = THIN_BORDER


def write_data_row(ws, row, values, start_col=1, fill=None, font=None):
    """Write a data row with optional fill."""
    for i, v in enumerate(values):
        cell = ws.cell(row=row, column=start_col + i, value=v)
        cell.font = font or FONT_DATA
        cell.fill = fill or FILL_INPUT
        cell.alignment = ALIGN_WRAP
        cell.border = THIN_BORDER


def write_section_header(ws, row, col, text):
    """Write a teal section header."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = FONT_SECTION
    return cell


def setup_print(ws, landscape=True, fit_cols=1, fit_rows=0, print_title_rows=None):
    """Configure print setup for professional output."""
    ws.page_setup.orientation = 'landscape' if landscape else 'portrait'
    ws.page_setup.fitToWidth = fit_cols
    ws.page_setup.fitToHeight = fit_rows
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    if print_title_rows:
        ws.print_title_rows = print_title_rows


def add_autofilter(ws, header_row, num_cols, start_col=1):
    """Add auto-filter to header row."""
    start = f'{get_column_letter(start_col)}{header_row}'
    end = f'{get_column_letter(start_col + num_cols - 1)}{header_row}'
    ws.auto_filter.ref = f'{start}:{end}'


def add_dropdown(ws, cell_range, options):
    """Add data validation dropdown."""
    dv = DataValidation(type='list', formula1=f'"{options}"', allow_blank=True)
    dv.error = 'Please select from the list'
    dv.errorTitle = 'Invalid Entry'
    dv.showDropDown = False
    ws.add_data_validation(dv)
    dv.add(cell_range)


def add_score_conditional(ws, cell_range):
    """Add conditional formatting for 1-5 scores (green to red)."""
    ws.conditional_formatting.add(cell_range,
        CellIsRule(operator='equal', formula=['5'],
                   fill=PatternFill('solid', fgColor='EA4335'),
                   font=Font(color=WHITE, bold=True)))
    ws.conditional_formatting.add(cell_range,
        CellIsRule(operator='equal', formula=['4'],
                   fill=PatternFill('solid', fgColor='FBBC04'),
                   font=Font(color=TEXT, bold=True)))
    ws.conditional_formatting.add(cell_range,
        CellIsRule(operator='equal', formula=['3'],
                   fill=PatternFill('solid', fgColor='FDE293'),
                   font=Font(color=TEXT)))
    ws.conditional_formatting.add(cell_range,
        CellIsRule(operator='equal', formula=['2'],
                   fill=PatternFill('solid', fgColor='B7E1CD'),
                   font=Font(color=TEXT)))
    ws.conditional_formatting.add(cell_range,
        CellIsRule(operator='equal', formula=['1'],
                   fill=PatternFill('solid', fgColor='57BB8A'),
                   font=Font(color=WHITE)))


def add_rag_conditional(ws, cell_range):
    """Add RAG (Red/Amber/Green) conditional formatting for text values."""
    ws.conditional_formatting.add(cell_range,
        CellIsRule(operator='equal', formula=['"Red"'],
                   fill=PatternFill('solid', fgColor='EA4335'),
                   font=Font(color=WHITE, bold=True)))
    ws.conditional_formatting.add(cell_range,
        CellIsRule(operator='equal', formula=['"Amber"'],
                   fill=PatternFill('solid', fgColor='FBBC04'),
                   font=Font(color=TEXT, bold=True)))
    ws.conditional_formatting.add(cell_range,
        CellIsRule(operator='equal', formula=['"Green"'],
                   fill=PatternFill('solid', fgColor='57BB8A'),
                   font=Font(color=WHITE, bold=True)))


def add_status_conditional(ws, cell_range):
    """Add conditional formatting for status columns."""
    ws.conditional_formatting.add(cell_range,
        CellIsRule(operator='equal', formula=['"Open"'],
                   fill=PatternFill('solid', fgColor='EA4335'),
                   font=Font(color=WHITE)))
    ws.conditional_formatting.add(cell_range,
        CellIsRule(operator='equal', formula=['"In Progress"'],
                   fill=PatternFill('solid', fgColor='FBBC04'),
                   font=Font(color=TEXT)))
    ws.conditional_formatting.add(cell_range,
        CellIsRule(operator='equal', formula=['"Resolved"'],
                   fill=PatternFill('solid', fgColor='57BB8A'),
                   font=Font(color=WHITE)))
    ws.conditional_formatting.add(cell_range,
        CellIsRule(operator='equal', formula=['"Escalated"'],
                   fill=PatternFill('solid', fgColor='9900FF'),
                   font=Font(color=WHITE)))


# ══════════════════════════════════════════════════
# SHEET BUILDERS
# ══════════════════════════════════════════════════

def build_cover(wb):
    ws = wb.active
    ws.title = 'Cover'
    ws.sheet_properties.tabColor = NAVY

    set_col_widths(ws, {'A': 90})
    ws.sheet_view.showGridLines = False

    # Spacer
    ws.row_dimensions[1].height = 40

    # Title block
    ws.row_dimensions[2].height = 36
    c = ws.cell(row=2, column=1, value='BANK RISK IDENTIFICATION')
    c.font = Font(name='Calibri', size=28, bold=True, color=NAVY)

    ws.row_dimensions[3].height = 24
    c = ws.cell(row=3, column=1, value='The Complete Methodology \u2014 Template Pack')
    c.font = Font(name='Calibri', size=16, color=GOLD)

    ws.row_dimensions[4].height = 8

    c = ws.cell(row=5, column=1, value='EON Risk Services Ltd')
    c.font = Font(name='Calibri', size=12, color=TEAL)
    c = ws.cell(row=6, column=1, value='www.eonriskservices.com')
    c.font = Font(name='Calibri', size=10, color=TEAL, underline='single')

    ws.row_dimensions[7].height = 6

    # Version info
    c = ws.cell(row=8, column=1, value='Version 1.0  \u2022  February 2026')
    c.font = Font(name='Calibri', size=10, color=GREY)

    ws.row_dimensions[9].height = 20

    # Description
    desc = ('This workbook contains 29 structured templates covering every phase '
            'of the EON Risk Identification Methodology. Each tab corresponds to '
            'a key artifact described in the book "Bank Risk Identification: '
            'The Complete Methodology".')
    c = ws.cell(row=10, column=1, value=desc)
    c.font = Font(name='Calibri', size=10, color=TEXT)
    c.alignment = ALIGN_WRAP

    ws.row_dimensions[10].height = 48

    ws.row_dimensions[11].height = 12

    # Color legend
    c = ws.cell(row=12, column=1, value='TAB COLOUR GUIDE')
    c.font = Font(name='Calibri', size=11, bold=True, color=NAVY)

    legend = [
        (13, '\u25a0  Navy tabs (1\u20135)', 'Foundation \u2014 Phase 1: Taxonomy, regulatory mapping, context setting, risk criteria', NAVY),
        (14, '\u25a0  Gold tabs (6\u20139)', 'Identification \u2014 Phase 2: Workshops, SWIFT prompts, bottom-up templates, pre-assessment', GOLD),
        (15, '\u25a0  Teal tabs (10\u201314)', 'Reconciliation & Assessment \u2014 Phases 3\u20134: Gap analysis, scoring, interaction, bow-tie, concentration', TEAL),
        (16, '\u25a0  Grey tabs (15\u201320)', 'Documentation & Integration \u2014 Phases 5\u20136: Inventory, profiles, KRIs, board reporting, stress testing, capital', GREY),
        (17, '\u25a0  Green tabs (21\u201329)', 'Ongoing Cycle & Regulatory \u2014 Emerging risks, registers, triggers, KPIs, training, audit, traceability', GREEN),
    ]
    for row, label, desc, color in legend:
        c = ws.cell(row=row, column=1, value=f'{label}: {desc}')
        c.font = Font(name='Calibri', size=10, color=color)
        ws.row_dimensions[row].height = 20

    ws.row_dimensions[18].height = 12

    # Instructions
    c = ws.cell(row=19, column=1, value='HOW TO USE THIS WORKBOOK')
    c.font = Font(name='Calibri', size=11, bold=True, color=TEAL)

    instructions = [
        '1. Each template is pre-structured with the fields described in the book.',
        '2. Yellow-highlighted cells are for your input. Grey cells contain formulas or fixed structure.',
        '3. Dropdown lists are pre-configured \u2014 click any input cell to see available options.',
        '4. Example data (in italics) shows what a completed entry looks like. Replace with your own.',
        '5. Start with tabs 1\u20135 (Foundation), then work through Identification, Reconciliation, and so on.',
        '6. The Risk Inventory (tab 15) is the master register \u2014 all other tabs feed into it.',
        '7. Print setup is pre-configured for landscape A4. Headers repeat on every page.',
    ]
    for i, text in enumerate(instructions):
        r = 20 + i
        c = ws.cell(row=r, column=1, value=text)
        c.font = Font(name='Calibri', size=10, color=TEXT)
        ws.row_dimensions[r].height = 18

    ws.row_dimensions[27].height = 20

    # Footer
    c = ws.cell(row=28, column=1, value='\u00a9 2026 EON Risk Services Ltd. All rights reserved.')
    c.font = Font(name='Calibri', size=9, color=GREY)
    c = ws.cell(row=29, column=1, value='This template pack is provided under the terms of the EON Risk Services Toolkit licence.')
    c.font = Font(name='Calibri', size=9, color=GREY)

    setup_print(ws, landscape=True, fit_cols=1)
    return ws


def build_risk_taxonomy(wb):
    ws = wb.create_sheet('1. Risk Taxonomy')
    ws.sheet_properties.tabColor = NAVY

    set_col_widths(ws, {'A': 6, 'B': 22, 'C': 26, 'D': 30, 'E': 50, 'F': 20, 'G': 25})

    ws.cell(row=1, column=1, value='Risk Taxonomy (L1 / L2 / L3)').font = FONT_TITLE
    ws.cell(row=2, column=1, value='Book Reference: Chapter 4 \u2014 The Risk Taxonomy').font = FONT_SUBTITLE
    ws.cell(row=3, column=1, value="Define your institution's three-level risk classification. Populate L1 categories first, then decompose into L2 and L3.").font = FONT_INSTRUCTION

    headers = ['#', 'L1 Category', 'L2 Sub-Category', 'L3 Granular Risk', 'Definition', 'COSO Objective', 'Regulatory Mapping']
    write_header_row(ws, 5, headers)

    data = [
        [1, 'Credit Risk', 'Counterparty Credit Risk', 'Single-Name Concentration', 'Risk of loss from excessive exposure to a single counterparty or group of connected counterparties', 'Operations', 'CRR Art. 395\u2013403'],
        [2, 'Credit Risk', 'Counterparty Credit Risk', 'Wrong-Way Risk', 'Risk that exposure increases when counterparty credit quality deteriorates', 'Operations', 'CRR Art. 291'],
        [3, 'Credit Risk', 'Country/Transfer Risk', 'Sovereign Default', 'Risk of loss from sovereign inability or unwillingness to honour obligations', 'Strategic', 'EBA GL/2017/16'],
        [4, 'Credit Risk', 'Credit Concentration', 'Sectoral Concentration', 'Risk of loss from excessive exposure to a single economic sector', 'Operations', 'CRR Art. 395\u2013403'],
        [5, 'Market Risk', 'Interest Rate Risk', 'Yield Curve Risk', 'Risk of loss from non-parallel shifts in the yield curve', 'Operations', 'CRR Art. 325\u2013377'],
        [6, 'Market Risk', 'FX Risk', 'Translation Risk', 'Risk of loss from converting foreign currency positions to reporting currency', 'Reporting', 'IAS 21'],
        [7, 'Market Risk', 'Equity Risk', 'Concentrated Equity Position', 'Risk of loss from large single-stock or correlated equity holdings', 'Operations', 'CRR Art. 325\u2013377'],
        [8, 'Operational Risk', 'People Risk', 'Key Person Dependency', 'Risk of loss from over-reliance on individuals with critical knowledge or skills', 'Operations', 'BCBS Principles'],
        [9, 'Operational Risk', 'Technology Risk', 'Cyber Attack', 'Risk of loss from malicious intrusion into technology systems', 'Operations', 'DORA Art. 5\u201316'],
        [10, 'Operational Risk', 'Process Risk', 'Trade Settlement Failure', 'Risk of loss from failed or delayed settlement of financial transactions', 'Operations', 'CSDR'],
        [11, 'Liquidity Risk', 'Funding Risk', 'Wholesale Funding Concentration', 'Risk from over-reliance on short-term wholesale funding sources', 'Strategic', 'CRR Art. 411\u2013428'],
        [12, 'Liquidity Risk', 'Market Liquidity Risk', 'Asset Fire-Sale Risk', 'Risk of loss from forced liquidation of assets at below-market prices', 'Operations', 'CRR Art. 411\u2013428'],
        [13, 'Strategic Risk', 'Business Model Risk', 'Revenue Concentration', 'Risk from over-dependence on a single product, market, or client segment', 'Strategic', 'PRA SS31/15'],
        [14, 'Conduct Risk', 'Sales Practices', 'Mis-Selling', 'Risk of loss from sale of unsuitable products to customers', 'Compliance', 'FCA PRIN'],
        [15, 'Compliance Risk', 'AML/CFT', 'Sanctions Breach', 'Risk of processing transactions that violate sanctions regimes', 'Compliance', 'EU AMLD6'],
    ]
    for i, row_data in enumerate(data):
        write_data_row(ws, 6 + i, row_data, fill=FILL_INPUT, font=FONT_EXAMPLE)

    # Empty rows for user input
    for i in range(15, 50):
        ws.cell(row=6 + i, column=1, value=i + 1).font = FONT_DATA

    add_dropdown(ws, 'F6:F200', 'Strategic,Operations,Reporting,Compliance')
    add_autofilter(ws, 5, 7)
    ws.freeze_panes = 'A6'
    setup_print(ws, print_title_rows='5:5')
    return ws


def build_reg_mapping(wb):
    ws = wb.create_sheet('2. Reg Mapping')
    ws.sheet_properties.tabColor = NAVY

    set_col_widths(ws, {'A': 25, 'B': 25, 'C': 22, 'D': 22, 'E': 22, 'F': 22, 'G': 22, 'H': 25})

    ws.cell(row=1, column=1, value='Regulatory Mapping Table').font = FONT_TITLE
    ws.cell(row=2, column=1, value='Book Reference: Chapter 4, 15 \u2014 Maps taxonomy to each regulatory framework').font = FONT_SUBTITLE
    ws.cell(row=3, column=1, value='For each L1/L2 taxonomy node, record the corresponding category in each regulator\'s framework.').font = FONT_INSTRUCTION

    headers = ['L1 Category', 'L2 Sub-Category', 'PRA (UK)', 'Fed / OCC (US)', 'ECB / SSM (EU)', 'FINMA (CH)', 'APRA (AU)', 'Notes']
    write_header_row(ws, 5, headers)

    data = [
        ['Credit Risk', 'Counterparty Credit Risk', 'Credit Risk', 'Credit Risk', 'Credit Risk', 'Kreditrisiko', 'Credit Risk', 'CCP exposures may sit under Market Risk in some frameworks'],
        ['Credit Risk', 'Credit Concentration', 'Concentration Risk', 'Credit Concentration', 'Concentration Risk', 'Konzentrationsrisiko', 'Credit Risk', 'PRA treats as separate principal risk'],
        ['Market Risk', 'Interest Rate Risk', 'Market Risk', 'Market Risk', 'Market Risk', 'Marktrisiko', 'Market Risk', 'IRRBB separate under PRA/Fed'],
        ['Market Risk', 'FX Risk', 'Market Risk', 'Market Risk', 'Market Risk', 'Marktrisiko', 'Market Risk', ''],
        ['Operational Risk', 'Technology Risk', 'Operational Risk', 'Operational Risk', 'Operational Risk', 'Operationelles Risiko', 'Operational Risk', 'DORA applies for EU; PRA may split cyber out'],
        ['Operational Risk', 'People Risk', 'Operational Risk', 'Operational Risk', 'Operational Risk', 'Operationelles Risiko', 'Operational Risk', 'SM&CR adds individual accountability (UK)'],
        ['Liquidity Risk', 'Funding Risk', 'Liquidity Risk', 'Liquidity Risk', 'Liquidity Risk', 'Liquidit\u00e4tsrisiko', 'Liquidity Risk', 'Captured via LCR/NSFR across all frameworks'],
        ['Strategic Risk', 'Business Model Risk', 'Business Risk', 'Strategic Risk', 'Business Model Risk', 'Strategisches Risiko', 'Strategic Risk', 'PRA terms it "Business Risk" in SS31/15'],
        ['Conduct Risk', 'Sales Practices', 'Conduct Risk', 'Compliance Risk', 'Conduct Risk', 'Verhaltensrisiko', 'Conduct Risk', 'Fed bundles under Compliance; UK/AU separate'],
        ['Compliance Risk', 'AML/CFT', 'Financial Crime', 'BSA/AML', 'ML/TF Risk', 'Geldw\u00e4scherei', 'AML/CTF', 'Terminology varies significantly across jurisdictions'],
    ]
    for i, row_data in enumerate(data):
        write_data_row(ws, 6 + i, row_data, fill=FILL_INPUT, font=FONT_EXAMPLE)

    add_autofilter(ws, 5, 8)
    ws.freeze_panes = 'C6'
    setup_print(ws, print_title_rows='5:5')
    return ws


def build_pestle(wb):
    ws = wb.create_sheet('3. PESTLE Assessment')
    ws.sheet_properties.tabColor = NAVY

    set_col_widths(ws, {'A': 18, 'B': 45, 'C': 30, 'D': 18, 'E': 15, 'F': 30})

    ws.cell(row=1, column=1, value='PESTLE Assessment').font = FONT_TITLE
    ws.cell(row=2, column=1, value='Book Reference: Chapter 5 \u2014 Setting the Context').font = FONT_SUBTITLE
    ws.cell(row=3, column=1, value='Assess external environment factors. Map each finding to taxonomy categories and rate significance (1\u20135).').font = FONT_INSTRUCTION

    headers = ['PESTLE Dimension', 'Key Finding / Observation', 'Taxonomy Categories Affected', 'Significance (1\u20135)', 'Trend', 'Source / Evidence']
    write_header_row(ws, 5, headers)

    dimensions = {
        'POLITICAL': [
            ['Government stability and policy direction', 'Strategic Risk, Country Risk', 3, 'Stable', 'Cabinet Office policy tracker'],
            ['Trade policy / sanctions regime changes', 'Compliance Risk, Credit Risk', 4, 'Increasing', 'OFAC/OFSI updates Q4 2025'],
            ['Regulatory reform agenda (Basel 3.1 implementation)', 'Capital Risk, Compliance Risk', 4, 'Increasing', 'PRA PS9/24'],
        ],
        'ECONOMIC': [
            ['Interest rate environment and trajectory', 'Market Risk, Credit Risk', 5, 'Decreasing', 'BoE MPC minutes; forward curve'],
            ['Credit cycle position (late cycle indicators)', 'Credit Risk, Strategic Risk', 4, 'Increasing', 'BoE Financial Stability Report'],
            ['Inflation / deflation pressures', 'Market Risk, Strategic Risk', 3, 'Decreasing', 'ONS CPI data'],
            ['Commercial real estate repricing', 'Credit Risk, Capital Risk', 4, 'Increasing', 'MSCI UK Property Index'],
        ],
        'SOCIAL': [
            ['Demographic shifts affecting customer base', 'Strategic Risk', 2, 'Stable', 'ONS population projections'],
            ['Public trust in financial institutions', 'Reputational Risk, Conduct Risk', 3, 'Stable', 'Edelman Trust Barometer 2025'],
            ['Remote working impact on operational resilience', 'Operational Risk, Technology Risk', 3, 'Stable', 'Internal HR data'],
        ],
        'TECHNOLOGICAL': [
            ['AI/ML adoption in credit decisioning', 'Model Risk, Operational Risk', 4, 'Increasing', 'BoE AI discussion paper'],
            ['Cloud concentration (AWS/Azure/GCP)', 'Technology Risk, Third-Party Risk', 4, 'Increasing', 'FCA/PRA outsourcing review'],
            ['Quantum computing threat to cryptography', 'Technology Risk', 2, 'Increasing', 'NIST post-quantum standards'],
        ],
        'LEGAL': [
            ['GDPR enforcement intensification', 'Compliance Risk, Operational Risk', 3, 'Increasing', 'ICO enforcement tracker'],
            ['Senior Managers & Certification Regime expansion', 'Conduct Risk, Compliance Risk', 3, 'Stable', 'FCA SM&CR directory'],
        ],
        'ENVIRONMENTAL': [
            ['Climate transition risk (carbon-intensive portfolios)', 'Credit Risk, Strategic Risk', 4, 'Increasing', 'NGFS scenarios; PRA SS3/19'],
            ['Physical risk from extreme weather events', 'Credit Risk, Operational Risk', 3, 'Increasing', 'Met Office climate projections'],
            ['Biodiversity and nature-related financial risks', 'Credit Risk, Reputational Risk', 2, 'Increasing', 'TNFD framework v1.0'],
        ],
    }

    row = 6
    for dim, findings in dimensions.items():
        # Dimension header
        c = ws.cell(row=row, column=1, value=dim)
        c.font = Font(name='Calibri', size=10, bold=True, color=TEAL)
        c.fill = FILL_TEAL_LIGHT
        for col in range(2, 7):
            ws.cell(row=row, column=col).fill = FILL_TEAL_LIGHT
        row += 1
        for finding in findings:
            ws.cell(row=row, column=2, value=finding[0]).font = FONT_EXAMPLE
            ws.cell(row=row, column=2).fill = FILL_INPUT
            ws.cell(row=row, column=3, value=finding[1]).font = FONT_EXAMPLE
            ws.cell(row=row, column=3).fill = FILL_INPUT
            ws.cell(row=row, column=4, value=finding[2]).font = FONT_EXAMPLE
            ws.cell(row=row, column=4).fill = FILL_INPUT
            ws.cell(row=row, column=5, value=finding[3]).font = FONT_EXAMPLE
            ws.cell(row=row, column=5).fill = FILL_INPUT
            ws.cell(row=row, column=6, value=finding[4]).font = FONT_EXAMPLE
            ws.cell(row=row, column=6).fill = FILL_INPUT
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = THIN_BORDER
                ws.cell(row=row, column=col).alignment = ALIGN_WRAP
            row += 1
        row += 1  # Space between dimensions

    add_dropdown(ws, 'D6:D100', '1,2,3,4,5')
    add_dropdown(ws, 'E6:E100', 'Increasing,Stable,Decreasing,New')
    add_score_conditional(ws, 'D6:D100')

    ws.freeze_panes = 'A6'
    setup_print(ws, print_title_rows='5:5')
    return ws


def build_internal_env(wb):
    ws = wb.create_sheet('4. Internal Environment')
    ws.sheet_properties.tabColor = NAVY

    set_col_widths(ws, {'A': 30, 'B': 45, 'C': 40, 'D': 15, 'E': 35})

    ws.cell(row=1, column=1, value='Internal Environment Assessment (Seven COSO Elements)').font = FONT_TITLE
    ws.cell(row=2, column=1, value='Book Reference: Chapter 5 \u2014 Setting the Context').font = FONT_SUBTITLE
    ws.cell(row=3, column=1, value='Assess each of the seven COSO ERM internal environment elements. Rate as Strong / Adequate / Weak / Critical.').font = FONT_INSTRUCTION

    headers = ['COSO Element', 'Assessment Question', 'Finding / Evidence', 'Rating', 'Risks Implicated']
    write_header_row(ws, 5, headers)

    data = [
        ['1. Risk Management Philosophy', 'How does senior management communicate risk appetite and tolerance?', 'Board-approved RAS published annually; cascaded via BU risk committees', 'Strong', 'Strategic Risk'],
        ['1. Risk Management Philosophy', 'Is risk-taking rewarded or penalised in practice?', 'Bonus pool adjusted for risk metrics but front-office still primarily revenue-driven', 'Adequate', 'Conduct Risk, Strategic Risk'],
        ['2. Board Risk Oversight', 'Does the Board actively challenge risk assessments or rubber-stamp?', 'BRC minutes show substantive challenge; 3 independent risk experts on Board', 'Strong', 'All material risks'],
        ['2. Board Risk Oversight', 'Does the Board have independent risk expertise?', 'Two members with CRO experience at peer institutions', 'Strong', 'All material risks'],
        ['3. Integrity and Ethical Values', 'Have there been conduct failures in the past 3 years?', 'One sales practice remediation in retail (2024); no systemic issues', 'Adequate', 'Conduct Risk, Reputational Risk'],
        ['3. Integrity and Ethical Values', 'Is whistleblowing actively encouraged and protected?', 'Annual awareness campaign; 12 reports last year, all investigated', 'Adequate', 'Conduct Risk, Compliance Risk'],
        ['4. Commitment to Competence', 'Are risk management roles staffed with qualified professionals?', '85% of risk team hold FRM/PRM; vacancy rate 8% (above target)', 'Adequate', 'Operational Risk'],
        ['4. Commitment to Competence', 'Is there adequate training for risk identification?', 'No formal training programme for BU risk leads; ad-hoc only', 'Weak', 'All risks \u2014 identification quality'],
        ['5. Organisational Structure', 'Are risk functions independent of revenue-generating functions?', 'CRO reports to CEO with dotted line to Board; budget set by Board', 'Strong', 'All material risks'],
        ['5. Organisational Structure', 'Does the CRO have direct Board access?', 'Standing invitation to all Board meetings; private sessions quarterly', 'Strong', 'All material risks'],
        ['6. Assignment of Authority', 'Are risk owners clearly defined for all material risks?', 'Documented for top 15 principal risks; gaps in L2/L3 ownership', 'Adequate', 'Operational Risk'],
        ['6. Assignment of Authority', 'Do business unit heads accept accountability for risks?', 'Variable \u2014 some BUs treat risk ID as a compliance exercise', 'Weak', 'All BU-level risks'],
        ['7. Human Resource Standards', 'Do compensation structures incentivise excessive risk-taking?', 'Deferred comp and malus/clawback in place; short-term targets still dominant', 'Adequate', 'Conduct Risk, Strategic Risk'],
        ['7. Human Resource Standards', 'Is there key-person dependency in risk management?', 'CRO function has 2 single points of failure in quantitative risk', 'Weak', 'Operational Risk, Model Risk'],
    ]
    for i, row_data in enumerate(data):
        write_data_row(ws, 6 + i, row_data, fill=FILL_INPUT, font=FONT_EXAMPLE)

    add_dropdown(ws, 'D6:D100', 'Strong,Adequate,Weak,Critical')

    # Conditional formatting for ratings
    ws.conditional_formatting.add('D6:D100',
        CellIsRule(operator='equal', formula=['"Strong"'],
                   fill=PatternFill('solid', fgColor='57BB8A'), font=Font(color=WHITE)))
    ws.conditional_formatting.add('D6:D100',
        CellIsRule(operator='equal', formula=['"Adequate"'],
                   fill=PatternFill('solid', fgColor='B7E1CD'), font=Font(color=TEXT)))
    ws.conditional_formatting.add('D6:D100',
        CellIsRule(operator='equal', formula=['"Weak"'],
                   fill=PatternFill('solid', fgColor='FBBC04'), font=Font(color=TEXT)))
    ws.conditional_formatting.add('D6:D100',
        CellIsRule(operator='equal', formula=['"Critical"'],
                   fill=PatternFill('solid', fgColor='EA4335'), font=Font(color=WHITE)))

    add_autofilter(ws, 5, 5)
    ws.freeze_panes = 'A6'
    setup_print(ws, print_title_rows='5:5')
    return ws


def build_risk_criteria(wb):
    ws = wb.create_sheet('5. Risk Criteria')
    ws.sheet_properties.tabColor = NAVY

    set_col_widths(ws, {'A': 8, 'B': 18, 'C': 45, 'D': 40, 'E': 45, 'F': 40})

    ws.cell(row=1, column=1, value='Risk Criteria \u2014 Impact, Likelihood, Vulnerability, Speed of Onset').font = FONT_TITLE
    ws.cell(row=2, column=1, value='Book Reference: Chapter 5, 9 \u2014 Four-dimensional scoring framework').font = FONT_SUBTITLE

    # Financial Impact Scale
    row = 4
    write_section_header(ws, row, 1, 'FINANCIAL IMPACT SCALE')
    row = 5
    write_header_row(ws, row, ['Score', 'Rating', 'Description', 'CET1 Impact Anchor', 'Example'])
    scales = [
        [1, 'Incidental', 'Negligible financial impact', '<1 bps of CET1', 'Minor operational loss absorbed within budget'],
        [2, 'Minor', 'Small financial loss within normal volatility', '1\u201310 bps of CET1', 'Failed trade, small fraud, minor write-down'],
        [3, 'Moderate', 'Material financial loss requiring management attention', '10\u201350 bps of CET1', 'Significant credit default, trading loss, regulatory fine'],
        [4, 'Major', 'Severe financial loss threatening profitability', '50\u2013200 bps of CET1', 'Major counterparty failure, portfolio-level write-down'],
        [5, 'Extreme', 'Existential financial loss threatening solvency', '>200 bps of CET1', 'Systemic event, multiple simultaneous defaults'],
    ]
    for i, s in enumerate(scales):
        write_data_row(ws, 6 + i, s, fill=FILL_WHITE)
    add_score_conditional(ws, 'A6:A10')

    # Regulatory Impact Scale
    row = 12
    write_section_header(ws, row, 1, 'REGULATORY AND LEGAL IMPACT SCALE')
    row = 13
    write_header_row(ws, row, ['Score', 'Rating', 'Description', '', ''])
    reg_scales = [
        [1, 'Incidental', 'Informal supervisory feedback, no action required'],
        [2, 'Minor', 'Formal supervisory finding requiring remediation'],
        [3, 'Moderate', 'Enforcement action, material fine, s166 review'],
        [4, 'Major', 'Significant fine (>1% revenue), restrictions on business activities'],
        [5, 'Extreme', 'Licence revocation, forced restructuring, criminal proceedings'],
    ]
    for i, s in enumerate(reg_scales):
        write_data_row(ws, 14 + i, s, fill=FILL_WHITE)
    add_score_conditional(ws, 'A14:A18')

    # Reputational Impact Scale
    row = 20
    write_section_header(ws, row, 1, 'REPUTATIONAL IMPACT SCALE')
    row = 21
    write_header_row(ws, row, ['Score', 'Rating', 'Description', '', ''])
    rep_scales = [
        [1, 'Incidental', 'No external awareness; internal learning only'],
        [2, 'Minor', 'Limited trade press coverage; no customer impact'],
        [3, 'Moderate', 'National media coverage; some customer attrition'],
        [4, 'Major', 'Sustained negative media; material customer/investor loss'],
        [5, 'Extreme', 'Systemic loss of confidence; run risk; regulatory intervention'],
    ]
    for i, s in enumerate(rep_scales):
        write_data_row(ws, 22 + i, s, fill=FILL_WHITE)
    add_score_conditional(ws, 'A22:A26')

    # Likelihood Scale
    row = 28
    write_section_header(ws, row, 1, 'LIKELIHOOD SCALE')
    row = 29
    write_header_row(ws, row, ['Score', 'Rating', 'Description', 'Probability Anchor', 'Time Horizon'])
    lik_scales = [
        [1, 'Rare', 'Highly unlikely to occur in normal conditions', '<1% probability', '>25 years'],
        [2, 'Unlikely', 'Could occur but not expected', '1\u201310%', '10\u201325 years'],
        [3, 'Possible', 'Reasonable chance of occurring', '10\u201340%', '3\u201310 years'],
        [4, 'Likely', 'Expected to occur in planning horizon', '40\u201380%', '1\u20133 years'],
        [5, 'Almost Certain', 'Expected imminently or already occurring', '>80%', '<1 year'],
    ]
    for i, s in enumerate(lik_scales):
        write_data_row(ws, 30 + i, s, fill=FILL_WHITE)
    add_score_conditional(ws, 'A30:A34')

    # Vulnerability Scale
    row = 36
    write_section_header(ws, row, 1, 'VULNERABILITY SCALE')
    row = 37
    write_header_row(ws, row, ['Score', 'Rating', 'Description', 'Control Anchor', ''])
    vul_scales = [
        [1, 'Very Low', 'Multiple independent controls; tested and proven effective', 'Automated, redundant, independently validated', ''],
        [2, 'Low', 'Adequate controls in place and regularly tested', 'Documented, tested annually, minor gaps', ''],
        [3, 'Moderate', 'Controls exist but effectiveness uncertain or untested', 'In place but not independently tested', ''],
        [4, 'High', 'Material control weaknesses identified', 'Known gaps, manual workarounds, audit findings open', ''],
        [5, 'Very High', 'No effective controls or controls have failed', 'No controls, failed controls, or deliberately bypassed', ''],
    ]
    for i, s in enumerate(vul_scales):
        write_data_row(ws, 38 + i, s, fill=FILL_WHITE)
    add_score_conditional(ws, 'A38:A42')

    # Speed of Onset Scale
    row = 44
    write_section_header(ws, row, 1, 'SPEED OF ONSET SCALE')
    row = 45
    write_header_row(ws, row, ['Score', 'Rating', 'Description', 'Response Window', ''])
    speed_scales = [
        [1, 'Very Slow', 'Risk materialises over years; ample time to respond', '>2 years', ''],
        [2, 'Slow', 'Risk builds over quarters; time for strategic response', '6 months \u2013 2 years', ''],
        [3, 'Moderate', 'Risk materialises over weeks/months', '1\u20136 months', ''],
        [4, 'Fast', 'Risk materialises within days', '1\u201330 days', ''],
        [5, 'Instantaneous', 'Risk materialises with no warning; immediate impact', '<24 hours', ''],
    ]
    for i, s in enumerate(speed_scales):
        write_data_row(ws, 46 + i, s, fill=FILL_WHITE)
    add_score_conditional(ws, 'A46:A50')

    setup_print(ws, print_title_rows='1:2')
    return ws


def build_workshop_planner(wb):
    ws = wb.create_sheet('6. Workshop Planner')
    ws.sheet_properties.tabColor = GOLD

    set_col_widths(ws, {'A': 8, 'B': 28, 'C': 15, 'D': 55, 'E': 35})

    ws.cell(row=1, column=1, value='Top-Down SWIFT Workshop \u2014 Structure and Guide Words').font = FONT_TITLE
    ws.cell(row=2, column=1, value='Book Reference: Chapter 6 \u2014 Top-Down Identification').font = FONT_SUBTITLE

    write_section_header(ws, 4, 1, 'WORKSHOP STRUCTURE')
    write_header_row(ws, 5, ['Stage', 'Activity', 'Duration', 'Purpose', 'Output'])

    stages = [
        [1, 'Opening and Context', '30 min', 'Present PESTLE findings, internal environment summary, prior cycle results. Set scope and ground rules.', 'Shared understanding of current risk landscape'],
        [2, 'Independent Assessment Review', '30 min', 'Compare pre-workshop submissions (Tab 9). Identify common themes, outliers, and blind spots across participants.', 'Aggregated view of participant perspectives'],
        [3, 'SWIFT Systematic Walk-Through', '90 min', 'Apply guide words (below) across each L1 risk category using the prompt matrix (Tab 7). Challenge assumptions systematically.', 'Comprehensive risk identification across all categories'],
        [4, 'Emerging Risk Horizon Scan', '30 min', 'Identify risks not yet in taxonomy \u2014 1yr, 3yr, 5yr horizons. Use Delphi technique for divergent views.', 'Emerging risk candidates for follow-up'],
        [5, 'Prioritisation and Scoring', '45 min', 'Preliminary four-dimensional scoring of identified risks using criteria from Tab 5.', 'Ranked principal risk list with initial scores'],
        [6, 'Wrap-Up and Actions', '15 min', 'Assign risk owners, identify gaps requiring further analysis, note disagreements and assumptions.', 'Action log, assumption register entries, disagreement log entries'],
    ]
    for i, s in enumerate(stages):
        write_data_row(ws, 6 + i, s, fill=FILL_WHITE)

    # Guide words section
    write_section_header(ws, 14, 1, 'SWIFT GUIDE WORDS')
    write_header_row(ws, 15, ['#', 'Guide Word', 'Purpose', 'Example Prompt', ''])

    guides = [
        [1, 'What if...', 'Explore deviation from expected conditions', 'What if interest rates rise 300bps in 6 months?', ''],
        [2, 'What would happen if...', 'Explore consequences of specific events', 'What would happen if our largest counterparty defaulted?', ''],
        [3, 'Could someone...', 'Explore internal actor risks (fraud, error, conduct)', 'Could someone bypass trade confirmation controls?', ''],
        [4, 'Could something...', 'Explore external/systemic threats', 'Could something disrupt our primary data centre for >24hrs?', ''],
        [5, 'Has anyone considered...', 'Surface assumptions and blind spots', 'Has anyone considered our exposure to a single cloud provider?', ''],
        [6, 'What is the worst...', 'Explore tail risks and extreme scenarios', 'What is the worst that could happen if this control failed completely?', ''],
        [7, 'What changed since...', 'Identify newly emerged or evolved risks', 'What changed since last year that could affect our credit portfolio?', ''],
    ]
    for i, g in enumerate(guides):
        write_data_row(ws, 16 + i, g, fill=FILL_WHITE)

    ws.freeze_panes = 'A6'
    setup_print(ws, print_title_rows='5:5')
    return ws


def build_swift_matrix(wb):
    ws = wb.create_sheet('7. SWIFT Prompt Matrix')
    ws.sheet_properties.tabColor = GOLD

    set_col_widths(ws, {'A': 24, 'B': 32, 'C': 32, 'D': 32, 'E': 32, 'F': 32})

    ws.cell(row=1, column=1, value='SWIFT Prompt Matrix').font = FONT_TITLE
    ws.cell(row=2, column=1, value="Book Reference: Chapter 6 \u2014 Customise for your institution's risk profile").font = FONT_SUBTITLE
    ws.cell(row=3, column=1, value='Populate this matrix before the workshop. Each cell contains a tailored prompt for that risk/guide-word combination.').font = FONT_INSTRUCTION

    headers = ['L1 Risk Category \u2193 / Guide Word \u2192', 'What if...', 'Could someone...', 'Could something...', 'Has anyone considered...', 'What is the worst...']
    write_header_row(ws, 5, headers)

    categories = [
        'Credit Risk', 'Market Risk', 'Operational Risk', 'Liquidity Risk',
        'Strategic Risk', 'Conduct Risk', 'Compliance Risk',
        'Technology / Cyber Risk', 'Reputational Risk', 'Climate / ESG Risk'
    ]

    # Example prompts for first 3 rows
    examples = {
        'Credit Risk': [
            'What if our top 5 counterparties were all downgraded 2 notches simultaneously?',
            'Could someone approve a credit limit breach without proper escalation?',
            'Could something cause correlated defaults across our CRE portfolio?',
            'Has anyone considered the second-order credit impact of a major cyber event at a clearing house?',
            'What is the worst outcome if our largest single-name exposure defaulted with zero recovery?',
        ],
        'Market Risk': [
            'What if yields inverted 200bps more than current curve implies?',
            'Could someone mismark a position to hide losses?',
            'Could something trigger a liquidity gap in our bond portfolio?',
            'Has anyone considered basis risk between our hedges and underlying exposures?',
            'What is the worst impact of a simultaneous rates + FX + equity shock?',
        ],
        'Operational Risk': [
            'What if our core banking system was unavailable for 72 hours?',
            'Could someone exfiltrate customer data without detection?',
            'Could something cause a complete failure of our disaster recovery?',
            'Has anyone considered the operational risk of our dependency on 3 key vendors?',
            'What is the worst outcome of a ransomware attack during year-end processing?',
        ],
    }

    for i, cat in enumerate(categories):
        row = 6 + i
        ws.cell(row=row, column=1, value=cat).font = Font(name='Calibri', size=10, bold=True, color=NAVY)
        ws.cell(row=row, column=1).fill = FILL_TEAL_LIGHT
        ws.cell(row=row, column=1).border = THIN_BORDER
        if cat in examples:
            for j, prompt in enumerate(examples[cat]):
                c = ws.cell(row=row, column=2 + j, value=prompt)
                c.font = FONT_EXAMPLE
                c.fill = FILL_INPUT
                c.alignment = ALIGN_WRAP
                c.border = THIN_BORDER
        else:
            for j in range(5):
                c = ws.cell(row=row, column=2 + j)
                c.fill = FILL_INPUT
                c.border = THIN_BORDER

    ws.freeze_panes = 'B6'
    setup_print(ws, print_title_rows='5:5')
    return ws


def build_bottom_up(wb):
    ws = wb.create_sheet('8. Bottom-Up Template')
    ws.sheet_properties.tabColor = GOLD

    set_col_widths(ws, {
        'A': 12, 'B': 20, 'C': 35, 'D': 30, 'E': 30,
        'F': 25, 'G': 30, 'H': 18, 'I': 25, 'J': 12, 'K': 25
    })

    ws.cell(row=1, column=1, value='Standardised Bottom-Up Risk Assessment Template (11 Fields)').font = FONT_TITLE
    ws.cell(row=2, column=1, value='Book Reference: Chapter 7 \u2014 Bottom-Up Identification').font = FONT_SUBTITLE
    ws.cell(row=3, column=1, value='Each business unit completes one row per risk. This is the standardised submission format.').font = FONT_INSTRUCTION

    headers = ['Risk ID', 'Taxonomy (L1/L2)', 'Risk Definition', 'Direct Drivers', 'Indirect Drivers',
               'Quantitative Metrics', 'Current Controls', 'Risk Owner', 'Emerging Risk Indicators',
               'Data Quality', 'Qualitative Notes']
    write_header_row(ws, 5, headers)

    data = [
        ['CR-001', 'Credit / Counterparty', 'Risk of loss from counterparty default on OTC derivative obligations',
         'Counterparty credit deterioration, wrong-way risk, collateral disputes',
         'Market volatility increasing exposure at default, correlation breakdown in stressed markets',
         'Current exposure: $2.1B gross, $0.8B net of collateral',
         'Daily margining, ISDA CSAs, credit limits, CVA desk hedging',
         'Head of Credit Risk', 'Widening CDS spreads on top 10 counterparties',
         'High', 'Concentration in energy sector counterparties increasing'],
        ['MR-001', 'Market / Interest Rate', 'Risk of loss from adverse movements in interest rates across the banking book',
         'Central bank policy changes, inflation expectations, yield curve shifts',
         'Duration mismatch between assets and liabilities, basis risk between hedges and exposures',
         'EVE sensitivity: -$180M per +100bps parallel shift; NII at risk: $45M',
         'Duration limits, hedging programme, ALM committee oversight, monthly NII forecasts',
         'Head of Market Risk', 'Forward curve steepening beyond model assumptions',
         'High', 'Model assumptions for deposit behavioural maturity under review'],
        ['OR-001', 'Operational / Technology', 'Risk of loss from failure or compromise of critical technology infrastructure',
         'Ageing infrastructure, cyber threat landscape, third-party dependencies',
         'Regulatory expectations (DORA), reputational contagion from peer incidents',
         'System availability: 99.92% (target 99.95%); 3 P1 incidents in past 12 months',
         'Redundant data centres, 24/7 monitoring, annual DR testing, penetration testing',
         'CTO / Head of IT Risk', 'Increasing frequency of phishing attempts; vendor patch delays',
         'Medium', 'DR test in Q3 revealed 4-hour gap in trade capture failover'],
        ['LR-001', 'Liquidity / Funding', 'Risk of inability to meet payment obligations as they fall due without incurring unacceptable losses',
         'Wholesale funding maturity concentration, deposit outflow sensitivity',
         'Rating downgrade triggering collateral calls, market-wide liquidity stress',
         'LCR: 135% (min 100%); NSFR: 112%; survival horizon: 92 days',
         'Liquidity buffer (HQLA), contingency funding plan, diversified funding sources',
         'Group Treasurer', 'Wholesale funding renewal rates declining; deposit beta increasing',
         'High', 'Upcoming maturity wall: $3.2B in wholesale funding maturing Q2'],
        ['SR-001', 'Strategic / Business Model', 'Risk that revenue concentration in a single product line creates earnings volatility',
         'Mortgage lending represents 62% of net interest income; competitive pressure',
         'Fintech disruption, changing customer preferences, regulatory capital changes',
         'Mortgage NII: $420M of $680M total; HHI concentration index: 0.42',
         'Strategic planning process, product diversification targets, competitor monitoring',
         'CEO / Head of Strategy', 'Fintech mortgage market share growing 3% p.a.',
         'Medium', 'Board has approved diversification strategy but execution is 18 months behind plan'],
    ]
    for i, row_data in enumerate(data):
        write_data_row(ws, 6 + i, row_data, fill=FILL_INPUT, font=FONT_EXAMPLE)

    add_dropdown(ws, 'J6:J200', 'High,Medium,Low,Very Low')
    add_autofilter(ws, 5, 11)
    ws.freeze_panes = 'C6'
    setup_print(ws, print_title_rows='5:5')
    return ws


def build_pre_workshop(wb):
    ws = wb.create_sheet('9. Pre-Workshop Assessment')
    ws.sheet_properties.tabColor = GOLD

    set_col_widths(ws, {'A': 8, 'B': 40, 'C': 22, 'D': 45, 'E': 15})

    ws.cell(row=1, column=1, value='Pre-Workshop Independent Assessment').font = FONT_TITLE
    ws.cell(row=2, column=1, value='Book Reference: Chapter 6 \u2014 Each participant completes before the workshop').font = FONT_SUBTITLE
    ws.cell(row=3, column=1, value='List your top 10 risks independently. Do not discuss with other participants before the workshop.').font = FONT_INSTRUCTION

    ws.cell(row=4, column=1, value='Participant Name:').font = Font(name='Calibri', size=10, bold=True, color=NAVY)
    ws.cell(row=4, column=2).fill = FILL_INPUT
    ws.cell(row=4, column=2).border = THIN_BORDER

    headers = ['Rank', 'Risk Description', 'Taxonomy Category', 'Why This Risk Now?', 'Confidence']
    write_header_row(ws, 6, headers)

    examples = [
        [1, 'Commercial real estate credit losses accelerating', 'Credit / Concentration', 'Rising rates + remote work reducing occupancy; US regional bank failures showed contagion pattern', 'High'],
        [2, 'Cyber attack on payment infrastructure', 'Operational / Technology', 'Increasing sophistication of attacks; DORA compliance deadline approaching; peer incidents in 2025', 'High'],
        [3, 'Interest rate model risk (deposit behavioural assumptions)', 'Market / Model Risk', 'SVB demonstrated that deposit stability assumptions fail under stress; our models not yet recalibrated', 'Medium'],
        [4, 'Wholesale funding concentration', 'Liquidity / Funding', '$3.2B maturity wall in Q2; credit spreads widening; potential rating downgrade under review', 'High'],
    ]
    for i, row_data in enumerate(examples):
        write_data_row(ws, 7 + i, row_data, fill=FILL_INPUT, font=FONT_EXAMPLE)

    for i in range(5, 11):
        ws.cell(row=7 + i - 1, column=1, value=i).font = FONT_DATA
        for col in range(1, 6):
            ws.cell(row=7 + i - 1, column=col).fill = FILL_INPUT
            ws.cell(row=7 + i - 1, column=col).border = THIN_BORDER

    add_dropdown(ws, 'E7:E16', 'High,Medium,Low')
    ws.freeze_panes = 'A7'
    setup_print(ws, print_title_rows='6:6')
    return ws


def build_gap_analysis(wb):
    ws = wb.create_sheet('10. Gap Analysis')
    ws.sheet_properties.tabColor = TEAL

    set_col_widths(ws, {'A': 12, 'B': 35, 'C': 15, 'D': 18, 'E': 35, 'F': 20, 'G': 15})

    ws.cell(row=1, column=1, value='Reconciliation Gap Analysis').font = FONT_TITLE
    ws.cell(row=2, column=1, value='Book Reference: Chapter 8 \u2014 Reconciliation and Enterprise Portfolio View').font = FONT_SUBTITLE

    # Top-Down Gaps
    write_section_header(ws, 4, 1, 'TOP-DOWN GAPS (risks in workshop not in bottom-up submissions)')
    write_header_row(ws, 5, ['Risk ID', 'Risk Description', 'Source', 'Gap Type', 'Resolution', 'Assigned Owner', 'Status'])

    td_examples = [
        ['TD-001', 'Cross-portfolio correlation risk in stressed markets', 'Top-Down', 'Blind Spot', 'BUs assessed individual risks but not portfolio-level interaction; adding to BU templates for next cycle', 'Head of Credit Risk', 'In Progress'],
        ['TD-002', 'Regulatory change risk from Basel 3.1 implementation', 'Top-Down', 'Scope Gap', 'Compliance team did not include in BU submissions; assigned to Head of Regulatory Affairs', 'Head of Compliance', 'Open'],
        ['TD-003', 'Reputational risk from ESG criticism of fossil fuel lending', 'Top-Down', 'Blind Spot', 'Not captured by any BU; sustainability team to own going forward', 'Head of Sustainability', 'In Progress'],
    ]
    for i, row_data in enumerate(td_examples):
        write_data_row(ws, 6 + i, row_data, fill=FILL_INPUT, font=FONT_EXAMPLE)

    # Bottom-Up Gaps
    write_section_header(ws, 18, 1, 'BOTTOM-UP GAPS (risks in submissions not in workshop)')
    write_header_row(ws, 19, ['Risk ID', 'Risk Description', 'Business Unit', 'Gap Type', 'Escalation Required?', 'Resolution', 'Status'])

    bu_examples = [
        ['BU-001', 'Key person dependency in quantitative risk team', 'Risk Analytics', 'Escalation', 'Yes - exceeds materiality', 'Escalated to principal risk list; succession planning initiated', 'Resolved'],
        ['BU-002', 'Trade surveillance system gaps in fixed income', 'Trading', 'Legitimate BU Risk', 'No', 'Retained as BU-level risk; not material at group level', 'Resolved'],
    ]
    for i, row_data in enumerate(bu_examples):
        write_data_row(ws, 20 + i, row_data, fill=FILL_INPUT, font=FONT_EXAMPLE)

    add_dropdown(ws, 'G6:G16', 'Open,In Progress,Resolved,Escalated')
    add_dropdown(ws, 'G20:G30', 'Open,In Progress,Resolved,Escalated')
    add_dropdown(ws, 'E20:E30', 'Yes - exceeds materiality,Yes - appetite breach,No,Under review')
    add_status_conditional(ws, 'G6:G16')
    add_status_conditional(ws, 'G20:G30')

    add_autofilter(ws, 5, 7)
    ws.freeze_panes = 'A6'
    setup_print(ws, print_title_rows='5:5')
    return ws


def build_risk_scoring(wb):
    ws = wb.create_sheet('11. Risk Scoring')
    ws.sheet_properties.tabColor = TEAL

    set_col_widths(ws, {
        'A': 12, 'B': 28, 'C': 11, 'D': 11, 'E': 11, 'F': 11,
        'G': 12, 'H': 11, 'I': 11, 'J': 11, 'K': 11, 'L': 12, 'M': 12, 'N': 12
    })

    ws.cell(row=1, column=1, value='Four-Dimensional Risk Scoring Worksheet').font = FONT_TITLE
    ws.cell(row=2, column=1, value='Book Reference: Chapter 9 \u2014 Assessment').font = FONT_SUBTITLE
    ws.cell(row=3, column=1, value='Score each risk on four dimensions (1\u20135). Inherent = no controls. Residual = after controls. Grey cells are formulas.').font = FONT_INSTRUCTION

    headers = ['Risk ID', 'Risk Name', 'Inh. Impact', 'Inh. Likelihood', 'Inh. Vulnerability',
               'Inh. Speed', 'Inh. Score', 'Ctrl Effect.', 'Res. Impact', 'Res. Likelihood',
               'Res. Vulnerability', 'Res. Score', 'Data Quality', 'Material?']
    write_header_row(ws, 5, headers)

    examples = [
        ['CR-001', 'Counterparty Credit Risk', 4, 3, 3, 4, None, 3, 3, 2, 2, None, 'High', 'Yes'],
        ['CR-002', 'CRE Concentration', 4, 4, 4, 3, None, 2, 3, 3, 3, None, 'Medium', 'Yes'],
        ['MR-001', 'Interest Rate Risk (Banking Book)', 4, 4, 3, 5, None, 3, 3, 3, 2, None, 'High', 'Yes'],
        ['OR-001', 'Cyber / Technology Failure', 5, 3, 3, 5, None, 3, 4, 2, 2, None, 'Medium', 'Yes'],
        ['LR-001', 'Wholesale Funding Concentration', 4, 3, 4, 4, None, 2, 3, 2, 3, None, 'High', 'Yes'],
        ['SR-001', 'Revenue Concentration', 3, 3, 3, 2, None, 2, 2, 3, 2, None, 'Medium', 'No'],
        ['CO-001', 'Sanctions Breach Risk', 5, 2, 3, 5, None, 3, 4, 1, 2, None, 'High', 'Yes'],
    ]

    for i, row_data in enumerate(examples):
        r = 6 + i
        for j, v in enumerate(row_data):
            c = ws.cell(row=r, column=1 + j, value=v)
            c.font = FONT_EXAMPLE
            c.fill = FILL_INPUT
            c.alignment = ALIGN_WRAP if j < 2 else ALIGN_CENTER
            c.border = THIN_BORDER

        # Formulas for computed columns
        ws.cell(row=r, column=7, value=f'=AVERAGE(C{r}:F{r})').font = Font(name='Calibri', size=10, bold=True)
        ws.cell(row=r, column=7).fill = FILL_GREY
        ws.cell(row=r, column=7).alignment = ALIGN_CENTER
        ws.cell(row=r, column=7).border = THIN_BORDER
        ws.cell(row=r, column=7).number_format = '0.0'

        ws.cell(row=r, column=12, value=f'=AVERAGE(I{r}:K{r})').font = Font(name='Calibri', size=10, bold=True)
        ws.cell(row=r, column=12).fill = FILL_GREY
        ws.cell(row=r, column=12).alignment = ALIGN_CENTER
        ws.cell(row=r, column=12).border = THIN_BORDER
        ws.cell(row=r, column=12).number_format = '0.0'

    # Empty rows with formulas
    for i in range(len(examples), 50):
        r = 6 + i
        ws.cell(row=r, column=7, value=f'=IF(C{r}="","",AVERAGE(C{r}:F{r}))').fill = FILL_GREY
        ws.cell(row=r, column=7).border = THIN_BORDER
        ws.cell(row=r, column=7).alignment = ALIGN_CENTER
        ws.cell(row=r, column=7).number_format = '0.0'
        ws.cell(row=r, column=12, value=f'=IF(I{r}="","",AVERAGE(I{r}:K{r}))').fill = FILL_GREY
        ws.cell(row=r, column=12).border = THIN_BORDER
        ws.cell(row=r, column=12).alignment = ALIGN_CENTER
        ws.cell(row=r, column=12).number_format = '0.0'

    # Dropdowns and conditional formatting
    add_dropdown(ws, 'C6:F200', '1,2,3,4,5')
    add_dropdown(ws, 'H6:K200', '1,2,3,4,5')
    add_dropdown(ws, 'M6:M200', 'High,Medium,Low,Very Low')
    add_dropdown(ws, 'N6:N200', 'Yes,No')

    add_score_conditional(ws, 'C6:F55')
    add_score_conditional(ws, 'H6:K55')

    # Material? highlighting
    ws.conditional_formatting.add('N6:N55',
        CellIsRule(operator='equal', formula=['"Yes"'],
                   fill=PatternFill('solid', fgColor='EA4335'), font=Font(color=WHITE, bold=True)))

    add_autofilter(ws, 5, 14)
    ws.freeze_panes = 'C6'
    setup_print(ws, print_title_rows='5:5')
    return ws


def build_interaction_matrix(wb):
    ws = wb.create_sheet('12. Interaction Matrix')
    ws.sheet_properties.tabColor = TEAL

    cols = {'A': 25}
    risk_names = ['Credit Risk', 'Market Risk', 'Liquidity Risk', 'Operational Risk',
                  'Conduct Risk', 'Strategic Risk', 'Technology Risk', 'Reputational Risk',
                  'Climate Risk', 'Compliance Risk', 'Model Risk', 'Third-Party Risk',
                  'Country Risk', 'Capital Risk', 'Legal Risk']
    for i in range(len(risk_names)):
        cols[get_column_letter(2 + i)] = 14
    set_col_widths(ws, cols)

    ws.cell(row=1, column=1, value='Risk Interaction Matrix').font = FONT_TITLE
    ws.cell(row=2, column=1, value='Book Reference: Chapter 10 \u2014 Risk Interaction').font = FONT_SUBTITLE
    ws.cell(row=3, column=1, value='Mark relationships: T=Trigger, A=Amplifier, C=Correlation. Leave blank if no material interaction.').font = FONT_INSTRUCTION

    # Header row
    headers = ['Risk \u2193  affects  Risk \u2192'] + risk_names
    write_header_row(ws, 5, headers)

    # Example interactions
    interactions = {
        (0, 1): 'T', (0, 2): 'A', (0, 7): 'T', (0, 13): 'A',  # Credit -> Market, Liquidity, Reputational, Capital
        (1, 0): 'A', (1, 2): 'T', (1, 13): 'A',  # Market -> Credit, Liquidity, Capital
        (2, 0): 'A', (2, 1): 'A', (2, 7): 'T',  # Liquidity -> Credit, Market, Reputational
        (3, 5): 'A', (3, 7): 'T', (3, 6): 'C',  # Operational -> Strategic, Reputational, Technology
        (6, 3): 'T', (6, 7): 'T', (6, 9): 'A',  # Technology -> Operational, Reputational, Compliance
        (8, 0): 'A', (8, 5): 'T', (8, 7): 'A',  # Climate -> Credit, Strategic, Reputational
    }

    for i, risk in enumerate(risk_names):
        r = 6 + i
        ws.cell(row=r, column=1, value=risk).font = Font(name='Calibri', size=10, bold=True, color=NAVY)
        ws.cell(row=r, column=1).fill = FILL_TEAL_LIGHT
        ws.cell(row=r, column=1).border = THIN_BORDER

        for j in range(len(risk_names)):
            c = ws.cell(row=r, column=2 + j)
            c.alignment = ALIGN_CENTER
            c.border = THIN_BORDER
            if i == j:
                c.fill = FILL_GREY
                c.value = '\u2014'
                c.font = Font(color=GREY)
            elif (i, j) in interactions:
                c.value = interactions[(i, j)]
                c.font = FONT_EXAMPLE
                c.fill = FILL_INPUT
            else:
                c.fill = FILL_INPUT

    add_dropdown(ws, f'B6:{get_column_letter(1 + len(risk_names))}{5 + len(risk_names)}', 'T,A,C,')

    # Conditional formatting for interaction types
    rng = f'B6:{get_column_letter(1 + len(risk_names))}{5 + len(risk_names)}'
    ws.conditional_formatting.add(rng,
        CellIsRule(operator='equal', formula=['"T"'],
                   fill=PatternFill('solid', fgColor='EA4335'), font=Font(color=WHITE, bold=True)))
    ws.conditional_formatting.add(rng,
        CellIsRule(operator='equal', formula=['"A"'],
                   fill=PatternFill('solid', fgColor='FBBC04'), font=Font(color=TEXT, bold=True)))
    ws.conditional_formatting.add(rng,
        CellIsRule(operator='equal', formula=['"C"'],
                   fill=PatternFill('solid', fgColor='B7E1CD'), font=Font(color=TEXT, bold=True)))

    ws.freeze_panes = 'B6'
    setup_print(ws, print_title_rows='5:5')
    return ws


def build_bowtie(wb):
    ws = wb.create_sheet('13. Bow-Tie Template')
    ws.sheet_properties.tabColor = TEAL

    set_col_widths(ws, {'A': 6, 'B': 28, 'C': 28, 'D': 22, 'E': 28, 'F': 28, 'G': 28})

    ws.cell(row=1, column=1, value='Bow-Tie Analysis Template').font = FONT_TITLE
    ws.cell(row=2, column=1, value='Book Reference: Chapter 10 \u2014 Visual causal analysis for material risks').font = FONT_SUBTITLE

    ws.cell(row=3, column=1, value='Risk:').font = Font(name='Calibri', size=10, bold=True, color=NAVY)
    ws.cell(row=3, column=2, value='Cyber attack on payment infrastructure').font = FONT_EXAMPLE
    ws.cell(row=3, column=2).fill = FILL_INPUT
    ws.cell(row=3, column=4, value='Risk Owner:').font = Font(name='Calibri', size=10, bold=True, color=NAVY)
    ws.cell(row=3, column=5, value='CTO / Head of IT Risk').font = FONT_EXAMPLE
    ws.cell(row=3, column=5).fill = FILL_INPUT

    headers = ['#', 'CAUSES', 'PREVENTIVE BARRIERS', '\u2190 RISK EVENT \u2192', 'MITIGATING BARRIERS', 'CONSEQUENCES', 'RECOVERY CONTROLS']
    write_header_row(ws, 5, headers)

    # Worked example
    example = [
        [1, 'Phishing attack on employee credentials', 'MFA + security awareness training', 'CYBER ATTACK ON PAYMENT SYSTEMS', 'Transaction monitoring + anomaly detection', 'Fraudulent payments executed', 'Insurance claim + law enforcement'],
        [2, 'Unpatched vulnerability in payment gateway', 'Automated patch management + vuln scanning', '', 'Payment halt + manual processing fallback', 'System downtime (customer impact)', 'DR activation + customer comms'],
        [3, 'Insider threat (malicious or negligent)', 'Privileged access management + activity logging', '', 'Forensic investigation + containment', 'Data breach (customer PII)', 'Breach notification + credit monitoring'],
        [4, 'Supply chain compromise (vendor software)', 'Vendor security assessment + code signing', '', 'Network segmentation limiting blast radius', 'Regulatory investigation', 'Remediation plan + regulatory engagement'],
        [5, 'DDoS attack on public-facing services', 'DDoS mitigation service + CDN', '', 'Backup communication channels', 'Reputational damage', 'PR response + customer outreach'],
    ]
    for i, row_data in enumerate(example):
        write_data_row(ws, 6 + i, row_data, fill=FILL_INPUT, font=FONT_EXAMPLE)

    # Escalation Factors section
    write_section_header(ws, 18, 1, 'ESCALATION FACTORS')
    esc_headers = ['#', 'Barrier Ref', 'Escalation Factor', 'Escalation Control', 'Owner', 'Last Tested']
    write_header_row(ws, 19, esc_headers)

    esc_examples = [
        [1, 'Preventive #1 (MFA)', 'MFA bypass technique emerges (e.g. MFA fatigue attack)', 'Adaptive authentication + FIDO2 hardware keys for privileged users', 'CISO', '2025-09-15'],
        [2, 'Mitigating #1 (Transaction monitoring)', 'Attacker learns normal transaction patterns', 'ML-based anomaly detection + human review of flagged transactions', 'Head of Fraud', '2025-11-01'],
    ]
    for i, row_data in enumerate(esc_examples):
        write_data_row(ws, 20 + i, row_data, fill=FILL_INPUT, font=FONT_EXAMPLE)

    ws.freeze_panes = 'A6'
    setup_print(ws, print_title_rows='5:5')
    return ws


def build_concentration(wb):
    ws = wb.create_sheet('14. Concentration Analysis')
    ws.sheet_properties.tabColor = TEAL

    set_col_widths(ws, {'A': 22, 'B': 35, 'C': 25, 'D': 20, 'E': 25, 'F': 20})

    ws.cell(row=1, column=1, value='Concentration Analysis (Three Types)').font = FONT_TITLE
    ws.cell(row=2, column=1, value='Book Reference: Chapter 10 \u2014 Single-name, Structural, Systemic').font = FONT_SUBTITLE

    # Single-name
    write_section_header(ws, 4, 1, 'SINGLE-NAME CONCENTRATIONS')
    write_header_row(ws, 5, ['Concentration Type', 'Description', 'Exposure / Metric', 'Limit', 'Current vs Limit', 'Action Required'])
    sn = [
        ['Counterparty', 'Largest single counterparty exposure', '$1.8B gross ($0.6B net)', '10% of CET1 ($2.1B)', '86% of limit', 'Monitor; approaching threshold'],
        ['Sector', 'Commercial Real Estate lending', '$4.2B (18% of loan book)', '20% of loan book', '90% of limit', 'Origination review; stress test Q2'],
        ['Geography', 'UK domestic exposure', '67% of total assets', 'No hard limit', 'Board-flagged', 'Diversification strategy underway'],
        ['Asset Class', 'Sovereign bond portfolio (UK gilts)', '$6.1B', '25% of HQLA buffer', '78% of limit', 'Duration review'],
        ['Funding Source', 'Top 5 wholesale depositors', '28% of total wholesale funding', '30%', '93% of limit', 'Diversification programme'],
    ]
    for i, row_data in enumerate(sn):
        write_data_row(ws, 6 + i, row_data, fill=FILL_INPUT, font=FONT_EXAMPLE)

    # Structural
    write_section_header(ws, 13, 1, 'STRUCTURAL CONCENTRATIONS')
    write_header_row(ws, 14, ['Concentration Type', 'Description', 'Evidence', 'Risk Implications', 'Mitigants', ''])
    st = [
        ['Business Model Assumption', 'Revenue depends on stable net interest margin', 'NIM compression from 2.1% to 1.6% over 3 years', 'Earnings volatility if rates fall further', 'Fee income diversification; hedging programme', ''],
        ['Correlation Assumption', 'CRE and residential property assumed uncorrelated', 'Both fell >15% in 2008-09; model assumes r=0.3', 'Capital model understates tail risk', 'Stress test overlay; model conservatism adjustment', ''],
        ['Technology Dependency', 'Single core banking platform (Temenos)', 'No viable short-term alternative; vendor lock-in', 'Operational resilience risk; vendor concentration', 'Contractual protections; exit plan (3-year horizon)', ''],
    ]
    for i, row_data in enumerate(st):
        write_data_row(ws, 15 + i, row_data, fill=FILL_INPUT, font=FONT_EXAMPLE)

    # Systemic
    write_section_header(ws, 20, 1, 'SYSTEMIC CONCENTRATIONS')
    write_header_row(ws, 21, ['Concentration Type', 'Description', 'Evidence', 'Risk Implications', 'Mitigants', ''])
    sy = [
        ['Clearing House', 'LCH Ltd concentration for rates derivatives', '85% of cleared rates volume through LCH', 'CCP default or operational failure', 'Default fund contribution monitoring; bilateral backup', ''],
        ['Cloud Provider', 'AWS hosts 70% of cloud workloads', 'No multi-cloud strategy currently', 'DORA compliance gap; single point of failure', 'Multi-cloud assessment initiated; critical apps on-prem', ''],
    ]
    for i, row_data in enumerate(sy):
        write_data_row(ws, 22 + i, row_data, fill=FILL_INPUT, font=FONT_EXAMPLE)

    add_autofilter(ws, 5, 6)
    ws.freeze_panes = 'A6'
    setup_print(ws, print_title_rows='5:5')
    return ws


def build_risk_inventory(wb):
    ws = wb.create_sheet('15. Risk Inventory')
    ws.sheet_properties.tabColor = GREY

    set_col_widths(ws, {
        'A': 12, 'B': 18, 'C': 35, 'D': 15, 'E': 18, 'F': 10,
        'G': 28, 'H': 10, 'I': 10, 'J': 10, 'K': 12, 'L': 22, 'M': 12, 'N': 14
    })

    ws.cell(row=1, column=1, value='Fourteen-Field Risk Inventory').font = FONT_TITLE
    ws.cell(row=2, column=1, value='Book Reference: Chapter 11 \u2014 The master register. All other tabs feed into this.').font = FONT_SUBTITLE
    ws.cell(row=3, column=1, value='This is the central risk inventory. One row per risk. Material risks (residual score \u22653.0) get a full Risk Profile (Tab 16).').font = FONT_INSTRUCTION

    headers = ['Risk ID', 'Taxonomy (L1/L2/L3)', 'Risk Definition', 'COSO Category', 'Risk Owner',
               'Inh. Score', 'Key Controls', 'Ctrl Eff.', 'Res. Score', 'Material?',
               'Data Quality', 'Interactions', 'Trend', 'Next Review']
    write_header_row(ws, 5, headers)

    data = [
        ['CR-001', 'Credit / Counterparty / Single-Name', 'Risk of loss from excessive exposure to a single counterparty', 'Operations', 'Head of Credit Risk',
         3.5, 'Daily margining, credit limits, CVA hedging', 3, 2.3, 'Yes', 'High', '\u2192MR-001(T), \u2192LR-001(A)', 'Stable', '2026-06-30'],
        ['CR-002', 'Credit / Concentration / CRE', 'Risk of loss from concentrated exposure to commercial real estate', 'Operations', 'Head of Credit Risk',
         3.8, 'Sector limits, LTV caps, stress testing', 2, 3.0, 'Yes', 'Medium', '\u2192MR-001(C), \u2192SR-001(A)', 'Increasing', '2026-03-31'],
        ['MR-001', 'Market / Interest Rate / IRRBB', 'Risk of loss from adverse interest rate movements in the banking book', 'Operations', 'Head of Market Risk',
         4.0, 'Duration limits, hedging, ALM oversight', 3, 2.7, 'Yes', 'High', '\u2192LR-001(T), \u2192CR-001(A)', 'Decreasing', '2026-06-30'],
        ['OR-001', 'Operational / Technology / Cyber', 'Risk of loss from cyber attack on critical technology infrastructure', 'Operations', 'CTO',
         4.0, 'MFA, patch management, SOC, pen testing', 3, 2.7, 'Yes', 'Medium', '\u2192RE-001(T), \u2192CO-001(A)', 'Increasing', '2026-03-31'],
        ['LR-001', 'Liquidity / Funding / Wholesale', 'Risk of inability to meet obligations from wholesale funding concentration', 'Strategic', 'Group Treasurer',
         3.8, 'HQLA buffer, CFP, diversified sources', 2, 2.7, 'Yes', 'High', '\u2192CR-001(A), \u2192RE-001(T)', 'Stable', '2026-06-30'],
        ['SR-001', 'Strategic / Business Model / Revenue', 'Risk of earnings volatility from revenue concentration in mortgage lending', 'Strategic', 'Head of Strategy',
         2.8, 'Strategic plan, diversification targets', 2, 2.3, 'No', 'Medium', '\u2192CR-002(C)', 'Stable', '2026-09-30'],
        ['CO-001', 'Compliance / AML / Sanctions', 'Risk of processing transactions that violate sanctions regimes', 'Compliance', 'Head of Compliance',
         3.8, 'Screening systems, training, QA testing', 3, 2.3, 'Yes', 'High', '\u2192RE-001(T), \u2192LE-001(A)', 'Increasing', '2026-03-31'],
    ]
    for i, row_data in enumerate(data):
        write_data_row(ws, 6 + i, row_data, fill=FILL_INPUT, font=FONT_EXAMPLE)

    add_dropdown(ws, 'D6:D200', 'Strategic,Operations,Reporting,Compliance')
    add_dropdown(ws, 'J6:J200', 'Yes,No')
    add_dropdown(ws, 'K6:K200', 'High,Medium,Low,Very Low')
    add_dropdown(ws, 'M6:M200', 'Increasing,Stable,Decreasing')

    ws.conditional_formatting.add('J6:J200',
        CellIsRule(operator='equal', formula=['"Yes"'],
                   fill=PatternFill('solid', fgColor='EA4335'), font=Font(color=WHITE, bold=True)))

    add_autofilter(ws, 5, 14)
    ws.freeze_panes = 'C6'
    setup_print(ws, print_title_rows='5:5')
    return ws


def build_risk_profile(wb):
    ws = wb.create_sheet('16. Risk Profile')
    ws.sheet_properties.tabColor = GREY

    set_col_widths(ws, {'A': 32, 'B': 75})

    ws.cell(row=1, column=1, value='Risk Profile Template (14 Elements)').font = FONT_TITLE
    ws.cell(row=2, column=1, value='Book Reference: Chapter 11 \u2014 Complete one per material risk').font = FONT_SUBTITLE

    ws.cell(row=3, column=1, value='Risk ID:').font = Font(name='Calibri', size=10, bold=True, color=NAVY)
    ws.cell(row=3, column=2, value='OR-001').font = FONT_EXAMPLE
    ws.cell(row=3, column=2).fill = FILL_INPUT

    elements = [
        ('1. Risk Definition', 'Risk of loss from cyber attack on critical payment and trading infrastructure, including data exfiltration, system disruption, and fraudulent transaction execution.'),
        ('2. Taxonomy Classification', 'Operational Risk / Technology Risk / Cyber Attack (L1 / L2 / L3)'),
        ('3. COSO Objective Category', 'Operations'),
        ('4. Underlying Drivers', 'Direct: Increasing sophistication of threat actors, unpatched vulnerabilities, insider threats, supply chain compromise. Indirect: Remote working expanding attack surface, cloud migration, regulatory complexity (DORA).'),
        ('5. Current Exposure Metrics', '3 P1 incidents in past 12 months. 47 P2 incidents. Mean time to detect: 4.2 hours. Mean time to contain: 18 hours. Estimated worst-case loss: $85M (insurance coverage: $50M).'),
        ('6. Risk Appetite Position', 'Board risk appetite: zero tolerance for data breach affecting >10,000 customers. Current position: WITHIN appetite but trend deteriorating. Last breach: June 2025 (1,200 records, contained).'),
        ('7. Key Controls', 'Preventive: MFA (all users), automated patch management, endpoint detection & response, security awareness training (quarterly). Detective: SOC 24/7 monitoring, SIEM correlation, transaction anomaly detection. Control effectiveness rating: 3/5 (adequate but gaps in OT environment).'),
        ('8. Cost-Benefit Assessment', 'Annual cybersecurity spend: $12M. Estimated annual loss expectancy without controls: $35M. Current residual ALE: $8M. ROI on security investment: 2.3x. Proposed FIDO2 upgrade ($0.4M) would reduce credential-based attacks by estimated 80%.'),
        ('9. Interaction Effects', 'Triggers: Reputational Risk (any public breach). Amplifies: Compliance Risk (regulatory investigation follows breach). Correlated with: Third-Party Risk (vendor compromise pathway).'),
        ('10. Concentration Assessment', 'Technology concentration: single core banking platform. Cloud concentration: 70% on AWS. Vendor concentration: Crowdstrike for EDR across all endpoints.'),
        ('11. Trend and Outlook', 'Deteriorating. Threat landscape expanding (AI-powered attacks); attack surface growing (cloud migration). Improvement in controls partially offsetting. Net trend: Increasing over 12-month horizon.'),
        ('12. Emerging Dimensions', 'Quantum computing threat to current encryption (3\u20135 year horizon). AI-generated deepfake social engineering (already emerging). Supply chain attacks increasing in frequency.'),
        ('13. Forward-Looking Actions', '1. Deploy FIDO2 hardware keys for privileged access (Q1 2026). 2. Complete OT security assessment (Q2 2026). 3. Implement multi-cloud strategy for critical workloads (2026\u201327). 4. Quantum-resistant encryption pilot (2027).'),
        ('14. Data Quality and Limitations', 'Data quality: Medium. Incident data comprehensive but threat intelligence coverage is partial. Loss estimation relies on scenario analysis (limited historical loss data for cyber). Conservatism adjustment: +1 to vulnerability score.'),
    ]

    row = 5
    for label, example_text in elements:
        ws.cell(row=row, column=1, value=label).font = Font(name='Calibri', size=10, bold=True, color=NAVY)
        ws.cell(row=row, column=1).fill = FILL_TEAL_LIGHT
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=1).alignment = ALIGN_WRAP

        ws.cell(row=row + 1, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2).font = FONT_INSTRUCTION
        ws.cell(row=row + 1, column=2, value=example_text).font = FONT_EXAMPLE
        ws.cell(row=row + 1, column=2).fill = FILL_INPUT
        ws.cell(row=row + 1, column=2).alignment = ALIGN_WRAP
        ws.cell(row=row + 1, column=2).border = THIN_BORDER
        ws.row_dimensions[row + 1].height = 48
        row += 2

    setup_print(ws, landscape=True)
    return ws


def build_kri_dashboard(wb):
    ws = wb.create_sheet('17. KRI Dashboard')
    ws.sheet_properties.tabColor = GREY

    set_col_widths(ws, {
        'A': 12, 'B': 28, 'C': 35, 'D': 15, 'E': 12, 'F': 12, 'G': 12, 'H': 12, 'I': 15
    })

    ws.cell(row=1, column=1, value='Key Risk Indicator Dashboard').font = FONT_TITLE
    ws.cell(row=2, column=1, value='Book Reference: Chapter 11, 13 \u2014 Leading indicators with traffic-light thresholds').font = FONT_SUBTITLE

    headers = ['Risk ID', 'Risk Name', 'KRI Description', 'Current Value', 'Green \u2264', 'Amber \u2264', 'Red >', 'Status', 'Trend']
    write_header_row(ws, 4, headers)

    data = [
        ['CR-001', 'Counterparty Credit Risk', 'Avg CDS spread on top 10 counterparties (bps)', 145, 100, 200, 200, 'Amber', 'Increasing'],
        ['CR-001', 'Counterparty Credit Risk', 'Largest single-name exposure (% of CET1)', 8.6, 8, 10, 10, 'Amber', 'Stable'],
        ['CR-002', 'CRE Concentration', 'CRE portfolio as % of loan book', 18, 15, 20, 20, 'Amber', 'Increasing'],
        ['CR-002', 'CRE Concentration', 'Avg LTV on new CRE originations', 68, 65, 75, 75, 'Amber', 'Stable'],
        ['MR-001', 'Interest Rate Risk', 'EVE sensitivity per +100bps (\u0394$M)', -180, -150, -250, -250, 'Amber', 'Stable'],
        ['MR-001', 'Interest Rate Risk', 'NII at risk 12-month horizon ($M)', 45, 30, 60, 60, 'Amber', 'Decreasing'],
        ['OR-001', 'Cyber / Technology', 'P1 incidents (rolling 12 months)', 3, 2, 4, 4, 'Amber', 'Stable'],
        ['OR-001', 'Cyber / Technology', 'Mean time to detect (hours)', 4.2, 2, 6, 6, 'Amber', 'Decreasing'],
        ['OR-001', 'Cyber / Technology', 'Phishing click rate (%)', 2.8, 2, 5, 5, 'Amber', 'Decreasing'],
        ['LR-001', 'Funding Concentration', 'LCR (%)', 135, 120, 105, 105, 'Green', 'Stable'],
        ['LR-001', 'Funding Concentration', 'Top 5 depositors (% of wholesale)', 28, 25, 35, 35, 'Amber', 'Stable'],
        ['CO-001', 'Sanctions Breach Risk', 'Sanctions screening false positive rate (%)', 3.2, 5, 8, 8, 'Green', 'Decreasing'],
    ]
    for i, row_data in enumerate(data):
        write_data_row(ws, 5 + i, row_data, fill=FILL_INPUT, font=FONT_EXAMPLE)

    add_dropdown(ws, 'H5:H200', 'Green,Amber,Red')
    add_dropdown(ws, 'I5:I200', 'Increasing,Stable,Decreasing')
    add_rag_conditional(ws, 'H5:H200')

    add_autofilter(ws, 4, 9)
    ws.freeze_panes = 'C5'
    setup_print(ws, print_title_rows='4:4')
    return ws


def build_principal_risk_report(wb):
    ws = wb.create_sheet('18. Principal Risk Report')
    ws.sheet_properties.tabColor = GREY

    set_col_widths(ws, {
        'A': 12, 'B': 28, 'C': 12, 'D': 18, 'E': 12, 'F': 18, 'G': 12, 'H': 14, 'I': 25, 'J': 14
    })

    ws.cell(row=1, column=1, value='Principal Risk Report \u2014 Board Reporting Template (10 Items)').font = FONT_TITLE
    ws.cell(row=2, column=1, value='Book Reference: Chapter 12 \u2014 Integration').font = FONT_SUBTITLE
    ws.cell(row=3, column=1, value='This report is presented to the Board Risk Committee. It contains only material risks with residual scores above the materiality threshold.').font = FONT_INSTRUCTION

    headers = ['Risk ID', 'Risk Name', 'Res. Score', 'Dominant Dimension', 'Trend', 'Appetite Status', 'Data Quality', 'New/Changed?', 'Key KRI Alert', 'Action Required']
    write_header_row(ws, 5, headers)

    data = [
        ['CR-001', 'Counterparty Credit Risk', 2.3, 'Speed of Onset', '\u2192', 'Within Appetite', 'High', 'No', 'CDS spreads widening', 'Monitor'],
        ['CR-002', 'CRE Concentration', 3.0, 'Vulnerability', '\u2191', 'Approaching Limit', 'Medium', 'Changed', 'CRE % approaching limit', 'Stress test Q2'],
        ['MR-001', 'Interest Rate Risk', 2.7, 'Speed of Onset', '\u2193', 'Within Appetite', 'High', 'No', 'EVE sensitivity elevated', 'Monitor'],
        ['OR-001', 'Cyber / Technology', 2.7, 'Impact', '\u2191', 'Within Appetite', 'Medium', 'No', '3 P1 incidents (12m)', 'FIDO2 deployment'],
        ['LR-001', 'Funding Concentration', 2.7, 'Vulnerability', '\u2192', 'Within Appetite', 'High', 'No', 'Maturity wall Q2', 'Diversification'],
        ['CO-001', 'Sanctions Breach Risk', 2.3, 'Impact', '\u2191', 'Within Appetite', 'High', 'No', 'None', 'Monitor'],
    ]
    for i, row_data in enumerate(data):
        write_data_row(ws, 6 + i, row_data, fill=FILL_INPUT, font=FONT_EXAMPLE)

    add_dropdown(ws, 'D6:D200', 'Financial,Regulatory,Reputational,Customer/Operational,Impact,Likelihood,Vulnerability,Speed of Onset')
    add_dropdown(ws, 'F6:F200', 'Within Appetite,Approaching Limit,Exceeds Appetite')
    add_dropdown(ws, 'G6:G200', 'High,Medium,Low,Very Low')

    ws.conditional_formatting.add('F6:F55',
        CellIsRule(operator='equal', formula=['"Exceeds Appetite"'],
                   fill=PatternFill('solid', fgColor='EA4335'), font=Font(color=WHITE, bold=True)))
    ws.conditional_formatting.add('F6:F55',
        CellIsRule(operator='equal', formula=['"Approaching Limit"'],
                   fill=PatternFill('solid', fgColor='FBBC04'), font=Font(color=TEXT, bold=True)))
    ws.conditional_formatting.add('F6:F55',
        CellIsRule(operator='equal', formula=['"Within Appetite"'],
                   fill=PatternFill('solid', fgColor='57BB8A'), font=Font(color=WHITE)))

    add_autofilter(ws, 5, 10)
    ws.freeze_panes = 'C6'
    setup_print(ws, print_title_rows='5:5')
    return ws


def build_stress_scenario(wb):
    ws = wb.create_sheet('19. Stress Scenario Map')
    ws.sheet_properties.tabColor = GREY

    set_col_widths(ws, {'A': 12, 'B': 28, 'C': 30, 'D': 18, 'E': 22, 'F': 30})

    ws.cell(row=1, column=1, value='Material Risk to Stress Scenario Mapping').font = FONT_TITLE
    ws.cell(row=2, column=1, value='Book Reference: Chapter 12 \u2014 Links inventory to ICAAP/ILAAP/CCAR').font = FONT_SUBTITLE

    headers = ['Risk ID', 'Risk Name', 'Stress Scenario', 'Framework', 'Severity Calibration', 'Loss Estimate Basis']
    write_header_row(ws, 4, headers)

    data = [
        ['CR-001', 'Counterparty Credit Risk', 'Top 3 counterparties default simultaneously', 'ICAAP', '1-in-200 year', 'LGD model + exposure at default estimates'],
        ['CR-002', 'CRE Concentration', 'UK CRE values fall 35% over 2 years', 'ICAAP', '1-in-100 year', 'LTV distribution + forced sale haircuts'],
        ['MR-001', 'Interest Rate Risk', 'Parallel shift +400bps in 6 months', 'ICAAP', 'Severe but plausible', 'EVE and NII models'],
        ['OR-001', 'Cyber / Technology', 'Core banking system down 5 days', 'ICAAP', 'Severe operational', 'Scenario-based expert estimate'],
        ['LR-001', 'Funding Concentration', 'Wholesale funding market closure + rating downgrade', 'ILAAP', '1-in-200 year', 'Cash flow projection + collateral call estimates'],
        ['CR-002', 'CRE Concentration', 'GFC repeat: property -40%, unemployment +5%', 'Reverse Stress Test', 'Point of non-viability', 'Integrated credit + market model'],
        ['ALL', 'Combined scenario', 'Stagflation: rates +300bps, GDP -3%, unemployment +4%', 'CCAR', 'Severely adverse', 'Integrated balance sheet model'],
    ]
    for i, row_data in enumerate(data):
        write_data_row(ws, 5 + i, row_data, fill=FILL_INPUT, font=FONT_EXAMPLE)

    add_dropdown(ws, 'D5:D200', 'ICAAP,ILAAP,CCAR,Recovery Plan,Reverse Stress Test')

    add_autofilter(ws, 4, 6)
    ws.freeze_panes = 'C5'
    setup_print(ws, print_title_rows='4:4')
    return ws


def build_reg_econ_gap(wb):
    ws = wb.create_sheet('20. Reg vs Econ Gap')
    ws.sheet_properties.tabColor = GREY

    set_col_widths(ws, {'A': 12, 'B': 28, 'C': 30, 'D': 30, 'E': 22, 'F': 30})

    ws.cell(row=1, column=1, value='Regulatory vs Economic Risk Gap Analysis').font = FONT_TITLE
    ws.cell(row=2, column=1, value='Book Reference: Chapter 12, 15 \u2014 Four-step comparison').font = FONT_SUBTITLE

    headers = ['Risk ID', 'Risk Name', 'Regulatory Treatment (Pillar 1)', 'Economic Assessment (Pillar 2)', 'Gap Identified?', 'Pillar 2A Action']
    write_header_row(ws, 4, headers)

    data = [
        ['CR-002', 'CRE Concentration', 'Standard credit risk weights; no explicit concentration charge in Pillar 1', 'Concentration adds ~40bps to loss distribution tail; not captured by standard weights', 'Under-captured in Pillar 1', 'Add Pillar 2A capital buffer for CRE concentration (+$180M)'],
        ['MR-001', 'Interest Rate Risk (IRRBB)', 'Not in Pillar 1 (trading book only)', 'EVE sensitivity of $180M per +100bps; material to capital adequacy', 'Not captured in Pillar 1', 'Full IRRBB Pillar 2A charge ($220M)'],
        ['OR-001', 'Cyber / Technology', 'Standard operational risk charge (Basic Indicator approach)', 'Scenario-based estimate: $85M worst case, far exceeding standard charge', 'Under-captured in Pillar 1', 'Additional operational risk buffer ($45M)'],
        ['SR-001', 'Revenue Concentration', 'No Pillar 1 treatment for strategic risk', 'Earnings volatility scenario: $120M NII shortfall in adverse case', 'Not captured in Pillar 1', 'Capital planning buffer under Pillar 2B'],
        ['CR-001', 'Counterparty Credit Risk', 'SA-CCR / IMM provides adequate capture', 'Economic capital model broadly aligns with regulatory', 'No gap', 'No additional action required'],
        ['LR-001', 'Funding Concentration', 'LCR/NSFR address liquidity (not capital)', 'Liquidity contingency costs: $30M under stress; capital impact indirect', 'Not captured in Pillar 1', 'Captured in ILAAP; cross-reference in ICAAP narrative'],
    ]
    for i, row_data in enumerate(data):
        write_data_row(ws, 5 + i, row_data, fill=FILL_INPUT, font=FONT_EXAMPLE)

    add_dropdown(ws, 'E5:E200', 'No gap,Under-captured in Pillar 1,Not captured in Pillar 1,Over-captured in Pillar 1')

    ws.conditional_formatting.add('E5:E55',
        CellIsRule(operator='equal', formula=['"Not captured in Pillar 1"'],
                   fill=PatternFill('solid', fgColor='EA4335'), font=Font(color=WHITE)))
    ws.conditional_formatting.add('E5:E55',
        CellIsRule(operator='equal', formula=['"Under-captured in Pillar 1"'],
                   fill=PatternFill('solid', fgColor='FBBC04'), font=Font(color=TEXT)))
    ws.conditional_formatting.add('E5:E55',
        CellIsRule(operator='equal', formula=['"No gap"'],
                   fill=PatternFill('solid', fgColor='57BB8A'), font=Font(color=WHITE)))

    add_autofilter(ws, 4, 6)
    ws.freeze_panes = 'C5'
    setup_print(ws, print_title_rows='4:4')
    return ws


def build_emerging_risks(wb):
    ws = wb.create_sheet('21. Emerging Risks')
    ws.sheet_properties.tabColor = GREEN

    set_col_widths(ws, {'A': 10, 'B': 35, 'C': 20, 'D': 15, 'E': 15, 'F': 35, 'G': 14})

    ws.cell(row=1, column=1, value='Emerging Risk Register').font = FONT_TITLE
    ws.cell(row=2, column=1, value='Book Reference: Chapter 6, 13 \u2014 Risks not yet material but requiring monitoring').font = FONT_SUBTITLE

    headers = ['ER-ID', 'Risk Description', 'Source', 'Horizon', 'Potential Impact', 'Monitoring Actions', 'Review Date']
    write_header_row(ws, 4, headers)

    data = [
        ['ER-001', 'Quantum computing rendering current encryption obsolete', 'Horizon Scan', '3\u20135 years', '5 \u2014 Extreme', 'Track NIST post-quantum standards; pilot quantum-safe algorithms', '2026-06-30'],
        ['ER-002', 'AI-generated deepfake fraud (voice/video)', 'Industry Event', '<1 year', '4 \u2014 Major', 'Enhanced voice authentication; deepfake detection tools in evaluation', '2026-03-31'],
        ['ER-003', 'Biodiversity-related financial risk (TNFD)', 'Regulatory Alert', '1\u20133 years', '3 \u2014 Moderate', 'TNFD gap analysis; portfolio screening for nature-dependent sectors', '2026-06-30'],
        ['ER-004', 'Stablecoin/CBDC impact on deposit franchise', 'Workshop', '1\u20133 years', '3 \u2014 Moderate', 'Monitor BoE digital pound consultation; model deposit outflow scenarios', '2026-09-30'],
        ['ER-005', 'Geopolitical fragmentation disrupting cross-border payments', 'Delphi Panel', '1\u20133 years', '4 \u2014 Major', 'Correspondent banking review; sanctions scenario planning', '2026-06-30'],
    ]
    for i, row_data in enumerate(data):
        write_data_row(ws, 5 + i, row_data, fill=FILL_INPUT, font=FONT_EXAMPLE)

    add_dropdown(ws, 'C5:C200', 'Delphi Panel,Workshop,Horizon Scan,Regulatory Alert,Industry Event,Internal Intelligence')
    add_dropdown(ws, 'D5:D200', '<1 year,1-3 years,3-5 years,>5 years')

    add_autofilter(ws, 4, 7)
    ws.freeze_panes = 'B5'
    setup_print(ws, print_title_rows='4:4')
    return ws


def build_assumption_register(wb):
    ws = wb.create_sheet('22. Assumption Register')
    ws.sheet_properties.tabColor = GREEN

    set_col_widths(ws, {'A': 8, 'B': 40, 'C': 22, 'D': 25, 'E': 35, 'F': 14})

    ws.cell(row=1, column=1, value='Assumption Register').font = FONT_TITLE
    ws.cell(row=2, column=1, value='Book Reference: Chapter 6, 11, 13 \u2014 Tracks key assumptions underlying risk assessments').font = FONT_SUBTITLE

    headers = ['#', 'Assumption', 'Risks Affected', 'Evidence Basis', 'Challenge / Sensitivity', 'Review Date']
    write_header_row(ws, 4, headers)

    data = [
        [1, 'Retail deposit behavioural maturity is 3\u20135 years under normal conditions', 'MR-001, LR-001', 'Historical deposit flow analysis (10-year dataset)', 'SVB demonstrated deposits can flee in hours under stress; sensitivity: +200bps rates = 15% outflow in model vs 40% actual at SVB', '2026-03-31'],
        [2, 'CRE and residential property values are weakly correlated (r=0.3)', 'CR-002', 'Pre-2008 data shows low correlation; GFC showed r>0.7', 'If correlation rises to 0.7 in stress, tail losses increase ~$200M; model does not capture regime change', '2026-06-30'],
        [3, 'UK sovereign default probability is negligible', 'CR-001, MR-001', 'UK CDS < 20bps; historical precedent', 'Gilt crisis of 2022 showed market stress is possible; no sovereign default assumption may be optimistic for some scenarios', '2026-09-30'],
        [4, 'Cloud service providers (AWS) maintain 99.99% availability', 'OR-001', 'AWS SLA and historical uptime data', 'Extended outage (>24hrs) has occurred 3 times in 5 years for other providers; our DR assumes max 4hr RTO', '2026-03-31'],
    ]
    for i, row_data in enumerate(data):
        write_data_row(ws, 5 + i, row_data, fill=FILL_INPUT, font=FONT_EXAMPLE)

    add_autofilter(ws, 4, 6)
    ws.freeze_panes = 'B5'
    setup_print(ws, print_title_rows='4:4')
    return ws


def build_disagreement_log(wb):
    ws = wb.create_sheet('23. Disagreement Log')
    ws.sheet_properties.tabColor = GREEN

    set_col_widths(ws, {'A': 8, 'B': 22, 'C': 35, 'D': 35, 'E': 22, 'F': 15})

    ws.cell(row=1, column=1, value='Disagreement Log').font = FONT_TITLE
    ws.cell(row=2, column=1, value='Book Reference: Chapter 6, 9 \u2014 Preserves divergent views rather than forcing false consensus').font = FONT_SUBTITLE

    headers = ['#', 'Risk / Topic', 'Position A', 'Position B', 'Resolution', 'Status']
    write_header_row(ws, 4, headers)

    data = [
        [1, 'CRE Concentration (CR-002)', 'Head of Credit: "Residual score should be 2.5 \u2014 controls are adequate and LTV caps prevent tail losses"', 'CRO: "Residual should be 3.0+ \u2014 SVB/Signature showed CRE risk is systematically underestimated"', 'Resolved \u2014 Position B', 'Resolved - Position B'],
        [2, 'Cyber risk materiality (OR-001)', 'CTO: "Residual score of 2.3 \u2014 we have best-in-class controls"', 'Head of Ops Risk: "Should be 3.0 \u2014 3 P1 incidents in 12 months shows controls are not as effective as claimed"', 'Escalated to CRO; scored at 2.7 with action plan', 'Escalated to CRO'],
        [3, 'Deposit behavioural maturity assumption', 'Treasury: "3-5 year assumption is conservative enough"', 'Risk Analytics: "Post-SVB, we need to model a scenario where 30% of deposits leave in 48 hours"', 'Evidence Requested \u2014 Analytics to provide updated analysis by Q2', 'Evidence Requested'],
    ]
    for i, row_data in enumerate(data):
        write_data_row(ws, 5 + i, row_data, fill=FILL_INPUT, font=FONT_EXAMPLE)

    add_dropdown(ws, 'F5:F200', 'Open,Evidence Requested,Resolved - Position A,Resolved - Position B,Escalated to CRO')

    add_autofilter(ws, 4, 6)
    ws.freeze_panes = 'B5'
    setup_print(ws, print_title_rows='4:4')
    return ws


def build_event_triggers(wb):
    ws = wb.create_sheet('24. Event-Driven Triggers')
    ws.sheet_properties.tabColor = GREEN

    set_col_widths(ws, {'A': 28, 'B': 40, 'C': 35, 'D': 25, 'E': 15})

    ws.cell(row=1, column=1, value='Event-Driven Trigger Framework').font = FONT_TITLE
    ws.cell(row=2, column=1, value='Book Reference: Chapter 13, 16 \u2014 Six categories requiring immediate risk review').font = FONT_SUBTITLE

    headers = ['Trigger Category', 'Description', 'Examples', 'Response Required', 'Owner']
    write_header_row(ws, 4, headers)

    data = [
        ['1. Market Dislocation', 'Significant market events affecting risk profile', 'Interest rate shock, credit spread widening, FX crisis, equity crash >10%', 'Immediate risk inventory review within 48 hours', 'Head of Market Risk'],
        ['2. Regulatory Change', 'New regulation, enforcement action, or supervisory communication', 'New capital requirements, sanctions changes, conduct findings, DORA deadlines', 'Impact assessment within 1 week', 'Head of Compliance'],
        ['3. Peer Event', 'Material loss or failure at comparable institution', 'Peer bank failure, industry-wide fraud discovery, systemic event', 'Relevance assessment within 1 week', 'CRO'],
        ['4. Internal Incident', 'Material operational loss, near-miss, or control failure', 'System outage, fraud attempt, significant error, compliance breach', 'Root cause analysis + forward-looking assessment', 'Head of Ops Risk'],
        ['5. Strategic Change', 'M&A, new product/market entry, business model change', 'Acquisition, new geography, product launch, major outsourcing', 'Pre-decision risk assessment using methodology', 'Head of Strategy'],
        ['6. External Environment Shift', 'Geopolitical, macroeconomic, or climate event', 'War, pandemic, sovereign crisis, natural disaster', 'PESTLE refresh + workshop within 2 weeks', 'CRO'],
    ]
    for i, row_data in enumerate(data):
        write_data_row(ws, 5 + i, row_data, fill=FILL_WHITE)

    ws.freeze_panes = 'A5'
    setup_print(ws, print_title_rows='4:4')
    return ws


def build_process_kpis(wb):
    ws = wb.create_sheet('25. Process KPIs')
    ws.sheet_properties.tabColor = GREEN

    set_col_widths(ws, {'A': 42, 'B': 18, 'C': 14, 'D': 14, 'E': 14, 'F': 14})

    ws.cell(row=1, column=1, value='Process Performance Indicators').font = FONT_TITLE
    ws.cell(row=2, column=1, value='Book Reference: Chapter 3, 12, 13 \u2014 KPIs tracking the process itself').font = FONT_SUBTITLE

    headers = ['Indicator', 'Target', 'Q1 Actual', 'Q2 Actual', 'Q3 Actual', 'Q4 Actual']
    write_header_row(ws, 4, headers)

    data = [
        ['Bottom-up submission completion rate (%)', '100%', '92%', '', '', ''],
        ['Average days to complete bottom-up submissions', '\u2264 15 days', '18 days', '', '', ''],
        ['Workshop attendance rate (% of required participants)', '\u2265 90%', '94%', '', '', ''],
        ['Gap analysis completion within deadline', '100%', '100%', '', '', ''],
        ['Risk profiles updated within 30 days of assessment', '100%', '86%', '', '', ''],
        ['KRI reporting timeliness (% on time)', '\u2265 95%', '97%', '', '', ''],
        ['Data Quality: % of material risks rated High', '\u2265 60%', '57%', '', '', ''],
        ['Emerging risks identified per cycle', '\u2265 3', '5', '', '', ''],
        ['Scoring calibration variance (cross-BU)', '\u2264 0.5', '0.4', '', '', ''],
        ['Internal Audit findings outstanding (count)', '0', '2', '', '', ''],
        ['Days from event trigger to risk review completion', '\u2264 5 days', '3 days', '', '', ''],
        ['Board report delivered on schedule', '100%', '100%', '', '', ''],
    ]
    for i, row_data in enumerate(data):
        write_data_row(ws, 5 + i, row_data, fill=FILL_INPUT, font=FONT_EXAMPLE)

    ws.freeze_panes = 'B5'
    setup_print(ws, print_title_rows='4:4')
    return ws


def build_training(wb):
    ws = wb.create_sheet('26. Training Programme')
    ws.sheet_properties.tabColor = GREEN

    set_col_widths(ws, {'A': 30, 'B': 50, 'C': 20, 'D': 15, 'E': 15})

    ws.cell(row=1, column=1, value='Training Programme (Five-Level)').font = FONT_TITLE
    ws.cell(row=2, column=1, value='Book Reference: Chapter 13 \u2014 Structured training by audience level').font = FONT_SUBTITLE

    headers = ['Audience', 'Content', 'Frequency', 'Delivery', 'Status']
    write_header_row(ws, 4, headers)

    data = [
        ['Board / Board Risk Committee', 'Methodology overview, how to challenge risk assessments, appetite framework, interpreting the principal risk report', 'Annual + ad hoc', 'Presentation', 'Scheduled'],
        ['CRO Function / Risk ID Team', 'Full methodology, scoring calibration, facilitation skills, data quality assessment, integration requirements', 'Annual + quarterly', 'Workshop', 'Complete'],
        ['Business Unit Risk Leads', 'Bottom-up template completion, specialist assessments, data quality standards, escalation criteria', 'Annual', 'Workshop', 'Scheduled'],
        ['Front-Line Employees', 'Risk awareness, how to identify and report risks, escalation procedures, case studies', 'Annual', 'e-Learning', 'In Progress'],
        ['Internal Audit', 'Methodology in detail, audit criteria, seven-area programme, regulatory expectations for independent assurance', 'Annual', 'Workshop', 'Complete'],
    ]
    for i, row_data in enumerate(data):
        write_data_row(ws, 5 + i, row_data, fill=FILL_WHITE)

    ws.freeze_panes = 'A5'
    setup_print(ws, print_title_rows='4:4')
    return ws


def build_audit_programme(wb):
    ws = wb.create_sheet('27. Audit Programme')
    ws.sheet_properties.tabColor = GREEN

    set_col_widths(ws, {'A': 8, 'B': 25, 'C': 50, 'D': 18, 'E': 30, 'F': 14})

    ws.cell(row=1, column=1, value='Internal Audit Annual Programme (Seven Areas)').font = FONT_TITLE
    ws.cell(row=2, column=1, value='Book Reference: Chapter 13 \u2014 Independent assurance over risk identification').font = FONT_SUBTITLE

    headers = ['Area', 'Audit Domain', 'Key Questions', 'Rating', 'Findings', 'Action Due']
    write_header_row(ws, 4, headers)

    data = [
        [1, 'Completeness', 'Does the inventory cover all material risks? Are there gaps vs regulatory expectations? Were emerging risks considered?', 'Satisfactory', 'All principal risks covered; minor gap in third-party risk granularity', '2026-03-31'],
        [2, 'Quality', 'Are risk definitions clear? Are scores evidence-based? Is there scoring consistency across BUs?', 'Needs Improvement', 'Scoring variance across BUs (0.4 avg); 3 risk definitions too vague', '2026-03-31'],
        [3, 'Reconciliation', 'Was the top-down / bottom-up reconciliation documented? Were gaps resolved? Is the audit trail complete?', 'Satisfactory', 'Full documentation; 5 gaps identified and resolved; 1 escalated to principal risk list', ''],
        [4, 'Data Quality', 'Are data quality ratings accurate? Are conservatism adjustments applied where data is weak?', 'Needs Improvement', '2 risks rated "High" quality without sufficient evidence; conservatism adjustments not consistently applied', '2026-06-30'],
        [5, 'Documentation', 'Are risk profiles complete for all material risks? Is the audit trail maintained?', 'Satisfactory', 'All 6 material risks have profiles; version control in place', ''],
        [6, 'Integration', 'Does the inventory feed capital planning? Are stress scenarios linked to identified risks?', 'Satisfactory', 'ICAAP integration complete; ILAAP linkage adequate; CCAR mapping documented', ''],
        [7, 'Continual Improvement', 'Are lessons learned actioned? Are process KPIs tracked? Has the methodology evolved?', 'Needs Improvement', 'Lessons learned log exists but 3 actions overdue; KPIs tracked but not reported to Board', '2026-06-30'],
    ]
    for i, row_data in enumerate(data):
        write_data_row(ws, 5 + i, row_data, fill=FILL_INPUT, font=FONT_EXAMPLE)

    add_dropdown(ws, 'D5:D20', 'Satisfactory,Needs Improvement,Unsatisfactory')

    ws.conditional_formatting.add('D5:D20',
        CellIsRule(operator='equal', formula=['"Satisfactory"'],
                   fill=PatternFill('solid', fgColor='57BB8A'), font=Font(color=WHITE)))
    ws.conditional_formatting.add('D5:D20',
        CellIsRule(operator='equal', formula=['"Needs Improvement"'],
                   fill=PatternFill('solid', fgColor='FBBC04'), font=Font(color=TEXT)))
    ws.conditional_formatting.add('D5:D20',
        CellIsRule(operator='equal', formula=['"Unsatisfactory"'],
                   fill=PatternFill('solid', fgColor='EA4335'), font=Font(color=WHITE)))

    ws.freeze_panes = 'C5'
    setup_print(ws, print_title_rows='4:4')
    return ws


def build_reg_traceability(wb):
    ws = wb.create_sheet('28. Reg Traceability')
    ws.sheet_properties.tabColor = GREEN

    set_col_widths(ws, {'A': 32, 'B': 18, 'C': 18, 'D': 18, 'E': 18, 'F': 18, 'G': 18})

    ws.cell(row=1, column=1, value='Regulatory Traceability Matrix').font = FONT_TITLE
    ws.cell(row=2, column=1, value='Book Reference: Chapter 15 \u2014 Maps methodology phases to regulatory requirements').font = FONT_SUBTITLE

    headers = ['Regulatory Framework', 'Phase 1\nFoundation', 'Phase 2\nIdentification', 'Phase 3\nReconciliation',
               'Phase 4\nAssessment', 'Phase 5\nInteraction', 'Phase 6\nDocumentation']
    write_header_row(ws, 4, headers)

    frameworks = [
        ['ISO 31000:2018', '\u2713 Direct', '\u2713 Direct', '\u2713 Direct', '\u2713 Direct', '\u2713 Indirect', '\u2713 Direct'],
        ['ISO 31010:2019', '\u25d0 Partial', '\u2713 Direct', '\u25d0 Partial', '\u2713 Direct', '\u2713 Direct', '\u25d0 Partial'],
        ['COSO ERM (2017)', '\u2713 Direct', '\u2713 Direct', '\u2713 Direct', '\u2713 Direct', '\u2713 Direct', '\u2713 Direct'],
        ['BCBS Corporate Governance (2015)', '\u2713 Direct', '\u2713 Direct', '\u2713 Indirect', '\u2713 Direct', '\u25d0 Partial', '\u2713 Direct'],
        ['PRA SS31/15 (UK)', '\u2713 Direct', '\u2713 Direct', '\u2713 Direct', '\u2713 Direct', '\u2713 Direct', '\u2713 Direct'],
        ['Fed SR 15-18 (US)', '\u2713 Direct', '\u2713 Direct', '\u2713 Direct', '\u2713 Direct', '\u2713 Indirect', '\u2713 Direct'],
        ['ECB Guide to ICAAP (EU)', '\u2713 Direct', '\u2713 Direct', '\u2713 Direct', '\u2713 Direct', '\u2713 Direct', '\u2713 Direct'],
        ['EBA GL on ICAAP/ILAAP', '\u2713 Direct', '\u2713 Direct', '\u2713 Direct', '\u2713 Direct', '\u25d0 Partial', '\u2713 Direct'],
        ['CRD V / CRR II', '\u2713 Indirect', '\u25d0 Partial', '\u2014 N/A', '\u2713 Direct', '\u25d0 Partial', '\u2713 Direct'],
        ['DORA (EU)', '\u25d0 Partial', '\u2713 Indirect', '\u2014 N/A', '\u2713 Direct', '\u25d0 Partial', '\u2713 Direct'],
        ['EBA GL on ICT Risk', '\u25d0 Partial', '\u2713 Indirect', '\u2014 N/A', '\u2713 Direct', '\u25d0 Partial', '\u2713 Direct'],
        ['EBA Outsourcing Guidelines', '\u25d0 Partial', '\u2713 Indirect', '\u2014 N/A', '\u25d0 Partial', '\u25d0 Partial', '\u2713 Direct'],
        ['EU AMLD6', '\u25d0 Partial', '\u2713 Direct', '\u2014 N/A', '\u2713 Direct', '\u2014 N/A', '\u2713 Direct'],
        ['FCA SYSC (UK)', '\u2713 Direct', '\u2713 Direct', '\u2713 Indirect', '\u2713 Direct', '\u25d0 Partial', '\u2713 Direct'],
        ['SM&CR (UK)', '\u25d0 Partial', '\u2713 Indirect', '\u2014 N/A', '\u25d0 Partial', '\u2014 N/A', '\u2713 Direct'],
        ['APRA CPS 220 (AU)', '\u2713 Direct', '\u2713 Direct', '\u2713 Direct', '\u2713 Direct', '\u2713 Indirect', '\u2713 Direct'],
    ]
    for i, row_data in enumerate(frameworks):
        write_data_row(ws, 5 + i, row_data, fill=FILL_INPUT, font=FONT_EXAMPLE)

    # Conditional formatting for coverage types
    rng = 'B5:G20'
    ws.conditional_formatting.add(rng,
        CellIsRule(operator='beginsWith', formula=['"✓ Direct"'],
                   fill=PatternFill('solid', fgColor='57BB8A'), font=Font(color=WHITE)))
    ws.conditional_formatting.add(rng,
        CellIsRule(operator='beginsWith', formula=['"✓ Indirect"'],
                   fill=PatternFill('solid', fgColor='B7E1CD'), font=Font(color=TEXT)))
    ws.conditional_formatting.add(rng,
        CellIsRule(operator='beginsWith', formula=['"◐ Partial"'],
                   fill=PatternFill('solid', fgColor='FDE293'), font=Font(color=TEXT)))
    ws.conditional_formatting.add(rng,
        CellIsRule(operator='beginsWith', formula=['"— N/A"'],
                   fill=PatternFill('solid', fgColor=LIGHT_GREY), font=Font(color=GREY)))

    add_dropdown(ws, 'B5:G20', '\u2713 Direct,\u2713 Indirect,\u25d0 Partial,\u2014 N/A')

    ws.freeze_panes = 'B5'
    setup_print(ws, print_title_rows='4:4')
    return ws


def build_lessons_learned(wb):
    ws = wb.create_sheet('29. Lessons Learned')
    ws.sheet_properties.tabColor = GREEN

    set_col_widths(ws, {'A': 8, 'B': 30, 'C': 40, 'D': 40, 'E': 25, 'F': 14})

    ws.cell(row=1, column=1, value='Lessons-Learned Review').font = FONT_TITLE
    ws.cell(row=2, column=1, value='Book Reference: Chapter 3, 13 \u2014 Annual review of process effectiveness').font = FONT_SUBTITLE

    headers = ['#', 'Area', 'What Worked Well', 'What Needs Improvement', 'Action / Owner', 'Status']
    write_header_row(ws, 4, headers)

    data = [
        [1, 'Workshop Design & Facilitation', 'SWIFT guide words generated 40% more risks than previous unstructured approach; pre-workshop submissions reduced groupthink', 'Workshop ran 45 mins over time; PESTLE section too long', 'Shorten PESTLE to 15 min summary / Head of Risk ID', 'Open'],
        [2, 'Bottom-Up Template Quality', 'Standardised 11-field template improved consistency; data quality ratings helped identify gaps', '3 BUs submitted templates with vague definitions; "emerging risk indicators" field often left blank', 'Add worked examples to template; require field completion / BU Risk Leads', 'In Progress'],
        [3, 'Reconciliation Process', 'Gap analysis identified 5 top-down gaps and 2 escalation-worthy BU risks', 'Process took 3 weeks instead of target 2 weeks; manual spreadsheet reconciliation is error-prone', 'Evaluate GRC tool for automated reconciliation / CRO', 'Open'],
        [4, 'Scoring Consistency', 'Four-dimensional scoring reduced "everything is a 3" problem; calibration session improved cross-BU alignment', 'Vulnerability scores still show highest variance across BUs; speed-of-onset is hard to calibrate', 'Add vulnerability scoring examples to training / Head of Risk ID', 'Open'],
        [5, 'Data Quality', 'Data quality ratings added transparency; conservatism adjustments applied to 4 risks', 'DQ assessment criteria not well understood by BUs; "Medium" rating used as default', 'Revise DQ criteria with examples; training for BU leads / Risk Analytics', 'Open'],
        [6, 'Board Reporting', 'Principal risk report well-received; dominant dimension analysis was new and valued', 'Board requested more forward-looking content; trend arrows alone not sufficient', 'Add 12-month outlook narrative per risk / CRO', 'In Progress'],
        [7, 'Regulatory Alignment', 'Traceability matrix confirmed full coverage of PRA SS31/15 and Fed SR 15-18', 'DORA alignment gaps identified; EBA outsourcing guidelines not fully mapped', 'Complete DORA gap analysis / Compliance', 'In Progress'],
    ]
    for i, row_data in enumerate(data):
        write_data_row(ws, 5 + i, row_data, fill=FILL_INPUT, font=FONT_EXAMPLE)

    add_dropdown(ws, 'F5:F50', 'Open,In Progress,Complete')
    add_status_conditional(ws, 'F5:F50')

    # Extend status formatting for "Complete"
    ws.conditional_formatting.add('F5:F50',
        CellIsRule(operator='equal', formula=['"Complete"'],
                   fill=PatternFill('solid', fgColor='57BB8A'), font=Font(color=WHITE)))

    add_autofilter(ws, 4, 6)
    ws.freeze_panes = 'B5'
    setup_print(ws, print_title_rows='4:4')
    return ws


# ══════════════════════════════════════════════════
# MAIN BUILD
# ══════════════════════════════════════════════════

if __name__ == '__main__':
    import os
    print('Building Risk Identification Template Pack...')
    wb = openpyxl.Workbook()

    builders = [
        ('Cover', build_cover),
        ('1. Risk Taxonomy', build_risk_taxonomy),
        ('2. Reg Mapping', build_reg_mapping),
        ('3. PESTLE Assessment', build_pestle),
        ('4. Internal Environment', build_internal_env),
        ('5. Risk Criteria', build_risk_criteria),
        ('6. Workshop Planner', build_workshop_planner),
        ('7. SWIFT Prompt Matrix', build_swift_matrix),
        ('8. Bottom-Up Template', build_bottom_up),
        ('9. Pre-Workshop Assessment', build_pre_workshop),
        ('10. Gap Analysis', build_gap_analysis),
        ('11. Risk Scoring', build_risk_scoring),
        ('12. Interaction Matrix', build_interaction_matrix),
        ('13. Bow-Tie Template', build_bowtie),
        ('14. Concentration Analysis', build_concentration),
        ('15. Risk Inventory', build_risk_inventory),
        ('16. Risk Profile', build_risk_profile),
        ('17. KRI Dashboard', build_kri_dashboard),
        ('18. Principal Risk Report', build_principal_risk_report),
        ('19. Stress Scenario Map', build_stress_scenario),
        ('20. Reg vs Econ Gap', build_reg_econ_gap),
        ('21. Emerging Risks', build_emerging_risks),
        ('22. Assumption Register', build_assumption_register),
        ('23. Disagreement Log', build_disagreement_log),
        ('24. Event-Driven Triggers', build_event_triggers),
        ('25. Process KPIs', build_process_kpis),
        ('26. Training Programme', build_training),
        ('27. Audit Programme', build_audit_programme),
        ('28. Reg Traceability', build_reg_traceability),
        ('29. Lessons Learned', build_lessons_learned),
    ]

    for name, builder in builders:
        builder(wb)
        print(f'  {name} done')

    output = os.path.join(os.path.dirname(__file__), 'Risk_Identification_Template_Pack.xlsx')
    wb.save(output)
    size_kb = os.path.getsize(output) / 1024
    print(f'\nSaved: {output}')
    print(f'Size: {size_kb:.0f} KB')
    print(f'Sheets: {len(wb.sheetnames)}')
    print('Done.')
