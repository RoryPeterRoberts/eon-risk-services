#!/usr/bin/env python3
"""
Build the Risk Identification Template Pack — Excel workbook with all methodology artifacts.
Branded to EON Risk Services styling.

Usage: python build_template_pack.py
Output: book/output/Risk_Identification_Template_Pack.xlsx
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── EON Branding ──────────────────────────────────────────────────────────────
NAVY = "1B2A4A"
GOLD = "C5A55A"
TEAL = "2A7F7F"
OFFWHITE = "F8F7F4"
WHITE = "FFFFFF"
LIGHT_GOLD = "F5EDD8"
LIGHT_TEAL = "D4ECEC"
LIGHT_GREY = "E8E8E8"
RED_BG = "FDE8E8"
AMBER_BG = "FFF3D4"
GREEN_BG = "D4F5D4"

# Fonts
HEADER_FONT = Font(name="Arial", size=11, bold=True, color=WHITE)
SUBHEADER_FONT = Font(name="Arial", size=10, bold=True, color=NAVY)
BODY_FONT = Font(name="Arial", size=10, color="333333")
TITLE_FONT = Font(name="Arial", size=14, bold=True, color=NAVY)
SECTION_FONT = Font(name="Arial", size=12, bold=True, color=TEAL)
SMALL_FONT = Font(name="Arial", size=9, color="666666")
LINK_FONT = Font(name="Arial", size=10, color=TEAL, italic=True)

# Fills
NAVY_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
GOLD_FILL = PatternFill(start_color=GOLD, end_color=GOLD, fill_type="solid")
TEAL_FILL = PatternFill(start_color=TEAL, end_color=TEAL, fill_type="solid")
OFFWHITE_FILL = PatternFill(start_color=OFFWHITE, end_color=OFFWHITE, fill_type="solid")
LIGHT_GOLD_FILL = PatternFill(start_color=LIGHT_GOLD, end_color=LIGHT_GOLD, fill_type="solid")
LIGHT_TEAL_FILL = PatternFill(start_color=LIGHT_TEAL, end_color=LIGHT_TEAL, fill_type="solid")
LIGHT_GREY_FILL = PatternFill(start_color=LIGHT_GREY, end_color=LIGHT_GREY, fill_type="solid")
RED_FILL = PatternFill(start_color=RED_BG, end_color=RED_BG, fill_type="solid")
AMBER_FILL = PatternFill(start_color=AMBER_BG, end_color=AMBER_BG, fill_type="solid")
GREEN_FILL = PatternFill(start_color=GREEN_BG, end_color=GREEN_BG, fill_type="solid")

# Borders
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

# Alignment
WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")
CENTER = Alignment(horizontal="center", vertical="center")


def set_col_widths(ws, widths):
    """Set column widths from a dict {col_letter: width}."""
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def write_header_row(ws, row, headers, fill=NAVY_FILL, font=HEADER_FONT):
    """Write a formatted header row."""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = font
        cell.fill = fill
        cell.alignment = WRAP_CENTER
        cell.border = THIN_BORDER


def write_data_row(ws, row, values, font=BODY_FONT, fills=None):
    """Write a data row with optional per-cell fills."""
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = font
        cell.alignment = WRAP
        cell.border = THIN_BORDER
        if fills and col_idx <= len(fills) and fills[col_idx - 1]:
            cell.fill = fills[col_idx - 1]


def write_title(ws, row, title, subtitle=None):
    """Write a sheet title."""
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = TITLE_FONT
    if subtitle:
        cell = ws.cell(row=row + 1, column=1, value=subtitle)
        cell.font = SMALL_FONT


def write_section(ws, row, text):
    """Write a section header."""
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = SECTION_FONT


def write_instruction(ws, row, text, col=1):
    """Write instruction text."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = SMALL_FONT
    cell.alignment = WRAP


def parse_usd_to_millions(usd_str):
    """Convert '$85.0B' to 85000.0, '$691M' to 691.0, etc."""
    s = usd_str.strip().replace(",", "").replace("$", "")
    try:
        if s.endswith("B"):
            return round(float(s[:-1]) * 1000, 1)
        elif s.endswith("M"):
            return round(float(s[:-1]), 1)
        else:
            return float(s)
    except (ValueError, IndexError):
        return 0


# ══════════════════════════════════════════════════════════════════════════════
# SHEET BUILDERS
# ══════════════════════════════════════════════════════════════════════════════

def build_cover(wb):
    ws = wb.active
    ws.title = "Cover"
    ws.sheet_properties.tabColor = NAVY

    set_col_widths(ws, {"A": 80})

    ws.cell(row=2, column=1, value="Bank Risk Identification").font = Font(
        name="Arial", size=24, bold=True, color=NAVY
    )
    ws.cell(row=3, column=1, value="The Complete Methodology — Template Pack").font = Font(
        name="Arial", size=16, color=GOLD
    )
    ws.cell(row=5, column=1, value="EON Risk Services Ltd").font = Font(
        name="Arial", size=12, color=TEAL
    )
    ws.cell(row=6, column=1, value="www.eonriskservices.com").font = LINK_FONT

    ws.cell(row=8, column=1, value="This workbook contains structured templates for every phase of the risk identification methodology.").font = BODY_FONT
    ws.cell(row=9, column=1, value="Each tab corresponds to a key artifact described in the book. Tabs are colour-coded by phase:").font = BODY_FONT

    legend = [
        (11, "■ Navy tabs — Foundation (Phase 1)", NAVY),
        (12, "■ Gold tabs — Identification (Phase 2)", GOLD),
        (13, "■ Teal tabs — Reconciliation & Assessment (Phase 3–4)", TEAL),
        (14, "■ Grey tabs — Documentation & Integration (Phase 5–6)", "666666"),
        (15, "■ Green tabs — Ongoing Cycle & Regulatory", "2D8B2D"),
    ]
    for row, text, color in legend:
        ws.cell(row=row, column=1, value=text).font = Font(name="Arial", size=10, color=color)

    ws.cell(row=17, column=1, value="INSTRUCTIONS").font = SECTION_FONT
    instructions = [
        "1. Each template is pre-structured with the fields described in the book.",
        "2. Yellow-highlighted cells are for your input. Grey cells are reference/formula.",
        "3. Dropdown validations are provided where the book specifies fixed options.",
        "4. Adapt column widths and add rows as needed for your institution.",
        "5. The Risk Inventory (Tab 15) is the master sheet — other tabs feed into it.",
        "6. Tab 30 (Industry Loss Database) contains 179 historical bank failures — use it for taxonomy testing and calibration.",
        "7. See the book chapter references on each tab for methodology context.",
    ]
    for i, text in enumerate(instructions):
        ws.cell(row=19 + i, column=1, value=text).font = BODY_FONT


def build_taxonomy(wb):
    ws = wb.create_sheet("1. Risk Taxonomy")
    ws.sheet_properties.tabColor = NAVY

    set_col_widths(ws, {"A": 6, "B": 20, "C": 25, "D": 35, "E": 45, "F": 20, "G": 25})

    write_title(ws, 1, "Risk Taxonomy (L1 / L2 / L3)", "Book Reference: Chapter 4 — The Risk Taxonomy")
    write_instruction(ws, 3, "Define your institution's three-level risk classification. Populate L1 categories first, then L2 sub-categories, then L3 granular risks.")

    write_header_row(ws, 5, ["#", "L1 Category", "L2 Sub-Category", "L3 Granular Risk", "Definition", "COSO Objective", "Regulatory Mapping"])

    # Example rows
    examples = [
        [1, "Credit Risk", "Counterparty Credit Risk", "Single-Name Concentration", "Risk of loss from excessive exposure to a single counterparty", "Operations", "CRR Art. 395-403"],
        [2, "Credit Risk", "Counterparty Credit Risk", "Wrong-Way Risk", "Risk that exposure increases when counterparty credit quality deteriorates", "Operations", "CRR Art. 291"],
        [3, "Credit Risk", "Country/Transfer Risk", "Sovereign Default", "Risk of loss from sovereign inability or unwillingness to honour obligations", "Strategic", "EBA GL/2017/16"],
        [4, "Market Risk", "Interest Rate Risk", "Yield Curve Risk", "Risk of loss from non-parallel shifts in the yield curve", "Operations", "CRR Art. 325-377"],
        [5, "Market Risk", "FX Risk", "Translation Risk", "Risk of loss from converting foreign currency positions", "Reporting", "IAS 21"],
        [6, "Operational Risk", "People Risk", "Key Person Dependency", "Risk of loss from over-reliance on individuals with critical knowledge", "Operations", "BCBS Principles"],
        [7, "Operational Risk", "Technology Risk", "Cyber Attack", "Risk of loss from malicious intrusion into technology systems", "Operations", "DORA Art. 5-16"],
        [8, "Liquidity Risk", "Funding Risk", "Wholesale Funding Concentration", "Risk of loss from over-reliance on short-term wholesale funding", "Strategic", "CRR Art. 411-428"],
        ["", "", "", "", "", "", ""],
        ["", "", "", "", "", "", ""],
    ]
    for i, row_data in enumerate(examples):
        fills = [None, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL, None, None]
        write_data_row(ws, 6 + i, row_data, fills=fills)

    # COSO dropdown
    dv_coso = DataValidation(type="list", formula1='"Strategic,Operations,Reporting,Compliance"', allow_blank=True)
    dv_coso.error = "Select from: Strategic, Operations, Reporting, Compliance"
    ws.add_data_validation(dv_coso)
    dv_coso.add(f"F6:F200")


def build_regulatory_mapping(wb):
    ws = wb.create_sheet("2. Reg Mapping")
    ws.sheet_properties.tabColor = NAVY

    set_col_widths(ws, {"A": 25, "B": 25, "C": 22, "D": 22, "E": 22, "F": 22, "G": 22, "H": 22})

    write_title(ws, 1, "Regulatory Mapping Table", "Book Reference: Chapter 4, 15 — Maps taxonomy to each regulator's categories")
    write_instruction(ws, 3, "For each L1/L2 taxonomy node, record the corresponding category used by each applicable regulator.")

    write_header_row(ws, 5, ["L1 Category", "L2 Sub-Category", "PRA (UK)", "Fed / OCC (US)", "ECB / SSM (EU)", "FINMA (CH)", "APRA (AU)", "Notes"])

    examples = [
        ["Credit Risk", "Counterparty Credit Risk", "Credit Risk", "Credit Risk", "Credit Risk", "Kreditrisiko", "Credit Risk", ""],
        ["Market Risk", "Interest Rate Risk", "Market Risk", "Market Risk", "Market Risk", "Marktrisiko", "Market Risk", ""],
        ["Operational Risk", "Technology Risk", "Operational Risk", "Operational Risk", "Operational Risk", "Operationelles Risiko", "Operational Risk", "DORA applies for EU"],
        ["", "", "", "", "", "", "", ""],
    ]
    for i, row_data in enumerate(examples):
        write_data_row(ws, 6 + i, row_data, fills=[None, None, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL, None])


def build_pestle(wb):
    ws = wb.create_sheet("3. PESTLE Assessment")
    ws.sheet_properties.tabColor = NAVY

    set_col_widths(ws, {"A": 18, "B": 40, "C": 30, "D": 20, "E": 15, "F": 30})

    write_title(ws, 1, "PESTLE Assessment", "Book Reference: Chapter 5 — Setting the Context")
    write_instruction(ws, 3, "Assess external environment factors. Map each finding to taxonomy categories. Score significance (1-5). Update quarterly.")

    write_header_row(ws, 5, ["PESTLE Dimension", "Key Finding / Observation", "Taxonomy Categories Affected", "Significance (1-5)", "Trend", "Source / Evidence"])

    dimensions = [
        ("POLITICAL", [
            "Government stability and policy direction",
            "Trade policy / sanctions regime",
            "Regulatory reform agenda",
        ]),
        ("ECONOMIC", [
            "Interest rate environment and trajectory",
            "Credit cycle position",
            "Inflation / deflation pressures",
            "FX volatility and currency regime",
        ]),
        ("SOCIAL", [
            "Demographic shifts affecting customer base",
            "Public trust in financial institutions",
            "Workforce availability and skills",
        ]),
        ("TECHNOLOGICAL", [
            "Fintech disruption to business model",
            "Cyber threat landscape evolution",
            "AI/ML adoption in risk management",
        ]),
        ("LEGAL", [
            "Pending litigation exposure",
            "Regulatory enforcement trends",
            "Data protection / privacy requirements",
        ]),
        ("ENVIRONMENTAL", [
            "Physical climate risk to operations",
            "Transition risk to loan portfolio",
            "ESG regulatory requirements",
        ]),
    ]

    row = 6
    for dim, items in dimensions:
        ws.cell(row=row, column=1, value=dim).font = Font(name="Arial", size=10, bold=True, color=TEAL)
        ws.cell(row=row, column=1).fill = LIGHT_TEAL_FILL
        ws.cell(row=row, column=1).border = THIN_BORDER
        for col in range(2, 7):
            ws.cell(row=row, column=col).fill = LIGHT_TEAL_FILL
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1
        for item in items:
            write_data_row(ws, row, ["", item, "", "", "", ""],
                          fills=[None, None, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL])
            row += 1
        # Blank row for additional entries
        write_data_row(ws, row, ["", "", "", "", "", ""],
                      fills=[None, None, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL])
        row += 1

    dv_sig = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=True)
    ws.add_data_validation(dv_sig)
    dv_sig.add("D6:D100")

    dv_trend = DataValidation(type="list", formula1='"Increasing,Stable,Decreasing,New"', allow_blank=True)
    ws.add_data_validation(dv_trend)
    dv_trend.add("E6:E100")


def build_internal_environment(wb):
    ws = wb.create_sheet("4. Internal Environment")
    ws.sheet_properties.tabColor = NAVY

    set_col_widths(ws, {"A": 30, "B": 45, "C": 35, "D": 15, "E": 35})

    write_title(ws, 1, "Internal Environment Assessment (Seven COSO Elements)", "Book Reference: Chapter 5 — Setting the Context")
    write_instruction(ws, 3, "Assess each of the seven COSO ERM internal environment elements. This is a qualitative assessment informing risk identification.")

    write_header_row(ws, 5, ["COSO Element", "Assessment Question", "Finding / Evidence", "Rating", "Risks Implicated"])

    elements = [
        ("1. Risk Management Philosophy", "How does senior management communicate risk appetite and tolerance?"),
        ("1. Risk Management Philosophy", "Is risk-taking rewarded or penalised in practice?"),
        ("2. Board Risk Oversight", "Does the Board actively challenge risk assessments or rubber-stamp them?"),
        ("2. Board Risk Oversight", "Does the Board have independent risk expertise?"),
        ("3. Integrity and Ethical Values", "Have there been conduct failures in the past 3 years?"),
        ("3. Integrity and Ethical Values", "Is whistleblowing actively encouraged and protected?"),
        ("4. Commitment to Competence", "Are risk management roles staffed with qualified professionals?"),
        ("4. Commitment to Competence", "Is there adequate training for risk identification?"),
        ("5. Organisational Structure", "Are risk functions independent of revenue-generating functions?"),
        ("5. Organisational Structure", "Does the CRO have direct Board access?"),
        ("6. Assignment of Authority", "Are risk owners clearly defined for all material risks?"),
        ("6. Assignment of Authority", "Do business unit heads accept accountability for risks?"),
        ("7. Human Resource Standards", "Do compensation structures incentivise excessive risk-taking?"),
        ("7. Human Resource Standards", "Is there key-person dependency in risk management?"),
    ]

    for i, (element, question) in enumerate(elements):
        write_data_row(ws, 6 + i, [element, question, "", "", ""],
                      fills=[None, None, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL])

    dv_rating = DataValidation(type="list", formula1='"Strong,Adequate,Weak,Critical"', allow_blank=True)
    ws.add_data_validation(dv_rating)
    dv_rating.add("D6:D100")


def build_risk_criteria(wb):
    ws = wb.create_sheet("5. Risk Criteria")
    ws.sheet_properties.tabColor = NAVY

    set_col_widths(ws, {"A": 8, "B": 18, "C": 40, "D": 40, "E": 40, "F": 40})

    write_title(ws, 1, "Risk Criteria — Impact, Likelihood, Vulnerability, Speed of Onset", "Book Reference: Chapter 5, 9 — Four-dimensional scoring framework")

    # Financial Impact
    write_section(ws, 4, "FINANCIAL IMPACT SCALE")
    write_header_row(ws, 5, ["Score", "Rating", "Description", "CET1 Impact Anchor", "Example", ""])
    fin_rows = [
        [1, "Incidental", "Negligible financial impact", "<1 bps of CET1", "Minor operational loss absorbed within budget"],
        [2, "Minor", "Small financial loss within normal volatility", "1-10 bps of CET1", "Failed trade, small fraud, minor write-down"],
        [3, "Moderate", "Material financial loss requiring management attention", "10-50 bps of CET1", "Significant credit default, trading loss, regulatory fine"],
        [4, "Major", "Severe financial loss threatening profitability", "50-200 bps of CET1", "Major counterparty failure, portfolio-level write-down"],
        [5, "Extreme", "Existential financial loss threatening solvency", ">200 bps of CET1", "Systemic event, multiple simultaneous defaults"],
    ]
    for i, row_data in enumerate(fin_rows):
        write_data_row(ws, 6 + i, row_data)

    # Regulatory Impact
    write_section(ws, 12, "REGULATORY AND LEGAL IMPACT SCALE")
    write_header_row(ws, 13, ["Score", "Rating", "Description", "", "", ""])
    reg_rows = [
        [1, "Incidental", "Informal supervisory feedback, no action required", "", "", ""],
        [2, "Minor", "Formal supervisory finding requiring remediation", "", "", ""],
        [3, "Moderate", "Enforcement action, material fine, s166 review", "", "", ""],
        [4, "Major", "Significant fine (>1% revenue), restrictions on business", "", "", ""],
        [5, "Extreme", "Licence revocation, forced restructuring, criminal proceedings", "", "", ""],
    ]
    for i, row_data in enumerate(reg_rows):
        write_data_row(ws, 14 + i, row_data)

    # Reputational Impact
    write_section(ws, 20, "REPUTATIONAL IMPACT SCALE")
    write_header_row(ws, 21, ["Score", "Rating", "Description", "", "", ""])
    rep_rows = [
        [1, "Incidental", "No external awareness", "", "", ""],
        [2, "Minor", "Trade press coverage, limited stakeholder concern", "", "", ""],
        [3, "Moderate", "National media coverage, customer attrition begins", "", "", ""],
        [4, "Major", "Sustained negative coverage, significant customer/investor loss", "", "", ""],
        [5, "Extreme", "Total loss of market confidence, franchise value destroyed", "", "", ""],
    ]
    for i, row_data in enumerate(rep_rows):
        write_data_row(ws, 22 + i, row_data)

    # Customer/Operational Impact
    write_section(ws, 28, "CUSTOMER AND OPERATIONAL IMPACT SCALE")
    write_header_row(ws, 29, ["Score", "Rating", "Description", "", "", ""])
    cust_rows = [
        [1, "Incidental", "Negligible disruption, no customer impact", "", "", ""],
        [2, "Minor", "Brief disruption, small number of customers affected", "", "", ""],
        [3, "Moderate", "Significant disruption, material customer harm", "", "", ""],
        [4, "Major", "Extended outage, widespread customer harm, systemic process failure", "", "", ""],
        [5, "Extreme", "Total service failure, mass customer harm", "", "", ""],
    ]
    for i, row_data in enumerate(cust_rows):
        write_data_row(ws, 30 + i, row_data)

    # Likelihood Scale
    write_section(ws, 36, "LIKELIHOOD SCALE")
    write_header_row(ws, 37, ["Score", "Rating", "Annual Probability", "Description", "", ""])
    like_rows = [
        [1, "Rare", "<1%", "Conceivable but no precedent in institution or industry", "", ""],
        [2, "Unlikely", "1-10%", "Has occurred in industry but not at this institution recently", "", ""],
        [3, "Possible", "10-50%", "Has occurred at this institution or frequently in industry", "", ""],
        [4, "Likely", "50-90%", "Expected to occur within assessment horizon", "", ""],
        [5, "Frequent", ">90%", "Occurring regularly, near-certain within horizon", "", ""],
    ]
    for i, row_data in enumerate(like_rows):
        write_data_row(ws, 38 + i, row_data)

    # Vulnerability Scale
    write_section(ws, 44, "VULNERABILITY SCALE")
    write_header_row(ws, 45, ["Score", "Rating", "Description", "Control Environment", "", ""])
    vuln_rows = [
        [1, "Very Low", "Institution well-prepared, robust controls, tested", "Multiple independent preventive and detective controls", "", ""],
        [2, "Low", "Good preparedness, minor gaps", "Controls in place, some untested or manual", "", ""],
        [3, "Moderate", "Partial preparedness, known control gaps", "Key controls exist but effectiveness uncertain", "", ""],
        [4, "High", "Poor preparedness, significant control gaps", "Controls inadequate or not designed for this scenario", "", ""],
        [5, "Very High", "No preparedness, no relevant controls", "No controls, no experience, no capability", "", ""],
    ]
    for i, row_data in enumerate(vuln_rows):
        write_data_row(ws, 46 + i, row_data)

    # Speed of Onset
    write_section(ws, 52, "SPEED OF ONSET SCALE")
    write_header_row(ws, 53, ["Score", "Rating", "Timeframe", "Description", "", ""])
    speed_rows = [
        [1, "Very Slow", "> 2 years", "Gradual emergence, ample time to respond", "", ""],
        [2, "Slow", "1-2 years", "Developing over quarters, visible early indicators", "", ""],
        [3, "Moderate", "3-12 months", "Emerges within a year, limited response window", "", ""],
        [4, "Fast", "1 week - 3 months", "Rapid materialisation, reactive response only", "", ""],
        [5, "Immediate", "< 1 week", "Instantaneous or near-instantaneous crystallisation", "", ""],
    ]
    for i, row_data in enumerate(speed_rows):
        write_data_row(ws, 54 + i, row_data)

    # Data Quality Rating
    write_section(ws, 60, "DATA QUALITY RATING")
    write_header_row(ws, 61, ["Rating", "Description", "Typical Basis", "Conservatism Adjustment", "", ""])
    dq_rows = [
        ["High", "Robust quantitative data, validated models", "Internal loss data, market data, audited figures", "None"],
        ["Medium", "Reasonable data with some gaps", "Industry benchmarks, expert estimates with evidence", "Minor upward adjustment"],
        ["Low", "Limited data, significant expert judgement", "Analogous events, qualitative assessment", "Material upward adjustment"],
        ["Very Low", "No relevant data, pure expert judgement", "Hypothetical assessment, no precedent", "Maximum conservatism"],
    ]
    for i, row_data in enumerate(dq_rows):
        write_data_row(ws, 62 + i, row_data + ["", ""])


def build_workshop_planner(wb):
    ws = wb.create_sheet("6. Workshop Planner")
    ws.sheet_properties.tabColor = GOLD

    set_col_widths(ws, {"A": 8, "B": 25, "C": 15, "D": 50, "E": 30})

    write_title(ws, 1, "Top-Down SWIFT Workshop — Structure and Guide Words", "Book Reference: Chapter 6 — Top-Down Identification")

    # Workshop Structure
    write_section(ws, 4, "WORKSHOP STRUCTURE")
    write_header_row(ws, 5, ["Stage", "Activity", "Duration", "Purpose", "Output"])
    stages = [
        [1, "Opening and Context", "30 min", "Present PESTLE findings, internal environment summary, prior-year risks, industry events", "Shared understanding of current landscape"],
        [2, "Independent Assessment Review", "30 min", "Compare pre-workshop submissions, identify common themes and divergences", "Aggregated view of participant perspectives"],
        [3, "SWIFT Systematic Walk-Through", "90 min", "Apply guide words across each L1 risk category using prompt matrix", "Comprehensive risk identification across all categories"],
        [4, "Emerging Risk Horizon Scan", "30 min", "Identify risks not yet in taxonomy — 1yr, 3yr, 5yr horizons", "Emerging risk candidates for Delphi follow-up"],
        [5, "Prioritisation and Scoring", "45 min", "Preliminary four-dimensional scoring of identified risks", "Ranked principal risk list with initial scores"],
        [6, "Wrap-Up and Actions", "15 min", "Assign risk owners, identify gaps requiring further analysis", "Action log, assumption register entries, disagreement log entries"],
    ]
    for i, row_data in enumerate(stages):
        write_data_row(ws, 6 + i, row_data)

    # SWIFT Guide Words
    write_section(ws, 14, "SWIFT GUIDE WORDS")
    write_header_row(ws, 15, ["#", "Guide Word", "Purpose", "Example Prompt", ""])
    guide_words = [
        [1, "What if...", "Explore deviation from expected conditions", "What if interest rates rise 300bps in 6 months?"],
        [2, "What would happen if...", "Explore consequences of specific events", "What would happen if our largest counterparty defaulted?"],
        [3, "Could someone...", "Explore internal actor risks (fraud, error, conduct)", "Could someone bypass trade confirmation controls?"],
        [4, "Could something...", "Explore external/systemic threats", "Could something disrupt our primary data centre for >24hrs?"],
        [5, "Has anyone considered...", "Surface assumptions and blind spots", "Has anyone considered our exposure to a single cloud provider?"],
        [6, "What is the worst...", "Explore tail risks and extreme scenarios", "What is the worst outcome from our wholesale funding concentration?"],
        [7, "How would we know if...", "Test detection capabilities and early warning", "How would we know if a rogue trader was concealing losses?"],
        [8, "What changed since...", "Identify new or evolved risks", "What changed since last quarter that affects our credit portfolio?"],
    ]
    for i, row_data in enumerate(guide_words):
        write_data_row(ws, 16 + i, row_data + [""])


def build_swift_prompt_matrix(wb):
    ws = wb.create_sheet("7. SWIFT Prompt Matrix")
    ws.sheet_properties.tabColor = GOLD

    set_col_widths(ws, {"A": 22, "B": 30, "C": 30, "D": 30, "E": 30, "F": 30})

    write_title(ws, 1, "SWIFT Prompt Matrix", "Book Reference: Chapter 6 — Customise for your institution's context")
    write_instruction(ws, 3, "Populate this matrix before the workshop. Each cell contains a specific prompt combining the guide word (column) with the risk category (row).")

    write_header_row(ws, 5, ["L1 Risk Category ↓ / Guide Word →", "What if...", "Could someone...", "Could something...", "Has anyone considered...", "What is the worst..."])

    categories = [
        "Credit Risk",
        "Market Risk",
        "Operational Risk",
        "Liquidity Risk",
        "Strategic Risk",
        "Conduct Risk",
        "Compliance Risk",
        "Technology / Cyber Risk",
        "Reputational Risk",
        "Climate / ESG Risk",
    ]
    for i, cat in enumerate(categories):
        row_data = [cat, "", "", "", "", ""]
        fills = [None, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL]
        write_data_row(ws, 6 + i, row_data, fills=fills)


def build_bottom_up_template(wb):
    ws = wb.create_sheet("8. Bottom-Up Template")
    ws.sheet_properties.tabColor = GOLD

    set_col_widths(ws, {"A": 12, "B": 18, "C": 30, "D": 30, "E": 30, "F": 25, "G": 30, "H": 20, "I": 25, "J": 15, "K": 15})

    write_title(ws, 1, "Standardised Bottom-Up Risk Assessment Template (11 Fields)", "Book Reference: Chapter 7 — Bottom-Up Identification")
    write_instruction(ws, 3, "Each business unit completes one row per risk. This is the standard template; specialist sub-processes (RCSA, ICT, AML, etc.) have additional fields.")

    headers = [
        "Risk ID", "Taxonomy (L1/L2)", "Risk Definition", "Direct Drivers",
        "Indirect Drivers", "Quantitative Metrics", "Current Controls",
        "Risk Owner", "Emerging Risk Indicators", "Data Quality", "Qualitative Notes"
    ]
    write_header_row(ws, 5, headers)

    # Example row
    example = [
        "CR-001", "Credit / Counterparty",
        "Risk of loss from counterparty default on OTC derivative obligations",
        "Counterparty credit deterioration, wrong-way risk, collateral shortfall",
        "Market volatility increasing exposure at default, correlation between counterparties",
        "Current exposure: $2.1B gross, $0.8B net of collateral",
        "Daily margining, ISDA CSAs, credit limits, CVA desk hedging",
        "Head of Credit Risk",
        "Widening CDS spreads on top 10 counterparties",
        "High",
        "Concentration in energy sector counterparties increasing"
    ]
    write_data_row(ws, 6, example)

    # Blank rows for input
    for i in range(7, 30):
        for col in range(1, 12):
            cell = ws.cell(row=i, column=col)
            cell.fill = LIGHT_GOLD_FILL
            cell.border = THIN_BORDER

    dv_dq = DataValidation(type="list", formula1='"High,Medium,Low,Very Low"', allow_blank=True)
    ws.add_data_validation(dv_dq)
    dv_dq.add("J6:J200")


def build_pre_workshop(wb):
    ws = wb.create_sheet("9. Pre-Workshop Assessment")
    ws.sheet_properties.tabColor = GOLD

    set_col_widths(ws, {"A": 6, "B": 35, "C": 20, "D": 40, "E": 15})

    write_title(ws, 1, "Pre-Workshop Independent Assessment", "Book Reference: Chapter 6 — Each participant completes before the workshop")
    write_instruction(ws, 3, "List your top 10 risks independently. Do not discuss with other participants before submission.")

    ws.cell(row=4, column=1, value="Participant Name:").font = SUBHEADER_FONT
    ws.cell(row=4, column=2).fill = LIGHT_GOLD_FILL
    ws.cell(row=4, column=2).border = THIN_BORDER

    write_header_row(ws, 6, ["Rank", "Risk Description", "Taxonomy Category", "Why This Risk Now?", "Confidence"])

    for i in range(1, 11):
        write_data_row(ws, 6 + i, [i, "", "", "", ""],
                      fills=[None, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL])

    dv_conf = DataValidation(type="list", formula1='"High,Medium,Low"', allow_blank=True)
    ws.add_data_validation(dv_conf)
    dv_conf.add("E7:E16")


def build_gap_analysis(wb):
    ws = wb.create_sheet("10. Gap Analysis")
    ws.sheet_properties.tabColor = TEAL

    set_col_widths(ws, {"A": 12, "B": 30, "C": 15, "D": 15, "E": 30, "F": 20, "G": 15})

    write_title(ws, 1, "Reconciliation Gap Analysis", "Book Reference: Chapter 8 — Reconciliation and Enterprise Portfolio View")

    # Top-Down Gaps
    write_section(ws, 4, "TOP-DOWN GAPS (risks in workshop not in bottom-up submissions)")
    write_header_row(ws, 5, ["Risk ID", "Risk Description", "Source", "Gap Type", "Resolution", "Assigned Owner", "Status"])

    for i in range(6, 16):
        write_data_row(ws, i, ["", "", "Top-Down", "", "", "", ""],
                      fills=[LIGHT_GOLD_FILL, LIGHT_GOLD_FILL, None, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL])

    # Bottom-Up Gaps
    write_section(ws, 18, "BOTTOM-UP GAPS (risks in submissions not in workshop)")
    write_header_row(ws, 19, ["Risk ID", "Risk Description", "Business Unit", "Gap Type", "Escalation Required?", "Resolution", "Status"])

    for i in range(20, 30):
        write_data_row(ws, i, ["", "", "", "", "", "", ""],
                      fills=[LIGHT_GOLD_FILL, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL])

    dv_status = DataValidation(type="list", formula1='"Open,In Progress,Resolved,Escalated"', allow_blank=True)
    ws.add_data_validation(dv_status)
    dv_status.add("G6:G16")
    dv_status.add("G20:G30")

    dv_esc = DataValidation(type="list", formula1='"Yes - exceeds materiality,Yes - appetite breach,No,Under review"', allow_blank=True)
    ws.add_data_validation(dv_esc)
    dv_esc.add("E20:E30")


def build_scoring_worksheet(wb):
    ws = wb.create_sheet("11. Risk Scoring")
    ws.sheet_properties.tabColor = TEAL

    set_col_widths(ws, {
        "A": 12, "B": 25, "C": 10, "D": 10, "E": 10, "F": 10,
        "G": 12, "H": 10, "I": 10, "J": 10, "K": 10, "L": 12,
        "M": 12, "N": 12
    })

    write_title(ws, 1, "Four-Dimensional Risk Scoring Worksheet", "Book Reference: Chapter 9 — Assessment")
    write_instruction(ws, 3, "Score each risk on four dimensions (1-5). Inherent = no controls. Apply control effectiveness to derive residual. See Risk Criteria tab for scale definitions.")

    headers = [
        "Risk ID", "Risk Name",
        "Inh. Impact", "Inh. Likelihood", "Inh. Vulnerability", "Inh. Speed",
        "Inh. Score",
        "Ctrl Effect.", "Res. Impact", "Res. Likelihood", "Res. Vulnerability", "Res. Score",
        "Data Quality", "Material?"
    ]
    write_header_row(ws, 5, headers)

    # Example row
    ws.cell(row=6, column=1, value="CR-001").font = BODY_FONT
    ws.cell(row=6, column=2, value="Counterparty Credit Risk").font = BODY_FONT
    for col in range(1, 15):
        ws.cell(row=6, column=col).border = THIN_BORDER

    # Inherent scores
    for col, val in [(3, 4), (4, 3), (5, 3), (6, 4)]:
        ws.cell(row=6, column=col, value=val).font = BODY_FONT
        ws.cell(row=6, column=col).alignment = CENTER

    # Inherent composite (simple average for template)
    ws.cell(row=6, column=7).value = "=AVERAGE(C6:F6)"
    ws.cell(row=6, column=7).font = Font(name="Arial", size=10, bold=True)
    ws.cell(row=6, column=7).alignment = CENTER
    ws.cell(row=6, column=7).fill = LIGHT_GREY_FILL

    ws.cell(row=6, column=8, value=3).font = BODY_FONT
    ws.cell(row=6, column=8).alignment = CENTER

    for col, val in [(9, 3), (10, 2), (11, 2)]:
        ws.cell(row=6, column=col, value=val).font = BODY_FONT
        ws.cell(row=6, column=col).alignment = CENTER

    ws.cell(row=6, column=12).value = "=AVERAGE(I6:K6)"
    ws.cell(row=6, column=12).font = Font(name="Arial", size=10, bold=True)
    ws.cell(row=6, column=12).alignment = CENTER
    ws.cell(row=6, column=12).fill = LIGHT_GREY_FILL

    ws.cell(row=6, column=13, value="High").font = BODY_FONT
    ws.cell(row=6, column=14, value="Yes").font = BODY_FONT

    # Blank input rows with formulas
    for r in range(7, 50):
        for col in range(1, 15):
            cell = ws.cell(row=r, column=col)
            cell.border = THIN_BORDER
            if col in (3, 4, 5, 6, 8, 9, 10, 11):
                cell.fill = LIGHT_GOLD_FILL
            cell.alignment = CENTER
        # Formulas
        ws.cell(row=r, column=7).value = f"=IF(C{r}=\"\",\"\",AVERAGE(C{r}:F{r}))"
        ws.cell(row=r, column=7).fill = LIGHT_GREY_FILL
        ws.cell(row=r, column=12).value = f"=IF(I{r}=\"\",\"\",AVERAGE(I{r}:K{r}))"
        ws.cell(row=r, column=12).fill = LIGHT_GREY_FILL

    # Validations
    dv_score = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=True)
    ws.add_data_validation(dv_score)
    for col_letter in ["C", "D", "E", "F", "H", "I", "J", "K"]:
        dv_score.add(f"{col_letter}6:{col_letter}200")

    dv_dq = DataValidation(type="list", formula1='"High,Medium,Low,Very Low"', allow_blank=True)
    ws.add_data_validation(dv_dq)
    dv_dq.add("M6:M200")

    dv_mat = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(dv_mat)
    dv_mat.add("N6:N200")


def build_interaction_matrix(wb):
    ws = wb.create_sheet("12. Interaction Matrix")
    ws.sheet_properties.tabColor = TEAL

    set_col_widths(ws, {"A": 25})
    for i in range(2, 18):
        ws.column_dimensions[get_column_letter(i)].width = 14

    write_title(ws, 1, "Risk Interaction Matrix", "Book Reference: Chapter 10 — Risk Interaction")
    write_instruction(ws, 3, "Mark relationships: T=Trigger, A=Amplifier, C=Correlation, blank=no relationship. Read row→column: 'Row risk [relationship] Column risk'.")

    # Example risks
    risks = [
        "Credit Risk", "Market Risk", "Liquidity Risk", "Operational Risk",
        "Conduct Risk", "Strategic Risk", "Technology Risk", "Reputational Risk",
        "Climate Risk", "Compliance Risk", "Model Risk", "Third-Party Risk",
        "Country Risk", "Capital Risk", "Legal Risk"
    ]

    # Header row
    write_header_row(ws, 5, ["Risk ↓  affects  Risk →"] + risks)

    # Matrix rows
    for i, risk in enumerate(risks):
        row = 6 + i
        ws.cell(row=row, column=1, value=risk).font = SUBHEADER_FONT
        ws.cell(row=row, column=1).fill = LIGHT_TEAL_FILL
        ws.cell(row=row, column=1).border = THIN_BORDER
        for j in range(len(risks)):
            cell = ws.cell(row=row, column=2 + j)
            cell.border = THIN_BORDER
            cell.alignment = CENTER
            if i == j:
                cell.fill = LIGHT_GREY_FILL
                cell.value = "—"
            else:
                cell.fill = LIGHT_GOLD_FILL

    dv_rel = DataValidation(type="list", formula1='"T,A,C,"', allow_blank=True)
    ws.add_data_validation(dv_rel)
    dv_rel.add(f"B6:{get_column_letter(1 + len(risks))}{5 + len(risks)}")


def build_bowtie(wb):
    ws = wb.create_sheet("13. Bow-Tie Template")
    ws.sheet_properties.tabColor = TEAL

    set_col_widths(ws, {"A": 6, "B": 25, "C": 25, "D": 20, "E": 25, "F": 25, "G": 25})

    write_title(ws, 1, "Bow-Tie Analysis Template", "Book Reference: Chapter 10 — Visual causal analysis for material risks")

    ws.cell(row=3, column=1, value="Risk:").font = SUBHEADER_FONT
    ws.cell(row=3, column=2).fill = LIGHT_GOLD_FILL
    ws.cell(row=3, column=2).border = THIN_BORDER
    ws.cell(row=3, column=4, value="Risk Owner:").font = SUBHEADER_FONT
    ws.cell(row=3, column=5).fill = LIGHT_GOLD_FILL
    ws.cell(row=3, column=5).border = THIN_BORDER

    write_header_row(ws, 5, ["#", "CAUSES", "PREVENTIVE BARRIERS", "← RISK EVENT →", "MITIGATING BARRIERS", "CONSEQUENCES", "RECOVERY CONTROLS"])

    for i in range(1, 11):
        write_data_row(ws, 5 + i, [i, "", "", "", "", "", ""],
                      fills=[None, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL, LIGHT_TEAL_FILL, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL])

    write_section(ws, 18, "ESCALATION FACTORS")
    write_header_row(ws, 19, ["#", "Barrier Ref", "Escalation Factor", "Escalation Control", "Owner", "Last Tested", ""])
    for i in range(1, 6):
        write_data_row(ws, 19 + i, [i, "", "", "", "", "", ""],
                      fills=[None, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL, None])


def build_concentration(wb):
    ws = wb.create_sheet("14. Concentration Analysis")
    ws.sheet_properties.tabColor = TEAL

    set_col_widths(ws, {"A": 20, "B": 30, "C": 25, "D": 20, "E": 25, "F": 20})

    write_title(ws, 1, "Concentration Analysis (Three Types)", "Book Reference: Chapter 10 — Single-name, Structural, Systemic")

    write_section(ws, 4, "SINGLE-NAME CONCENTRATIONS")
    write_header_row(ws, 5, ["Concentration Type", "Description", "Exposure / Metric", "Limit", "Current vs Limit", "Action Required"])
    sn_rows = [
        ["Counterparty", "", "", "", "", ""],
        ["Sector", "", "", "", "", ""],
        ["Geography", "", "", "", "", ""],
        ["Asset Class", "", "", "", "", ""],
        ["Funding Source", "", "", "", "", ""],
    ]
    for i, row_data in enumerate(sn_rows):
        write_data_row(ws, 6 + i, row_data, fills=[None, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL])

    write_section(ws, 13, "STRUCTURAL CONCENTRATIONS")
    write_header_row(ws, 14, ["Concentration Type", "Description", "Evidence", "Risk Implications", "Mitigants", ""])
    st_rows = [
        ["Business Model Assumption", "", "", "", "", ""],
        ["Correlation Assumption", "", "", "", "", ""],
        ["Regulatory Interpretation", "", "", "", "", ""],
        ["Technology Dependency", "", "", "", "", ""],
    ]
    for i, row_data in enumerate(st_rows):
        write_data_row(ws, 15 + i, row_data, fills=[None, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL, None])

    write_section(ws, 21, "SYSTEMIC CONCENTRATIONS")
    write_header_row(ws, 22, ["Concentration Type", "Description", "Institution's Exposure", "System-Level Risk", "Diversification Possible?", ""])
    sy_rows = [
        ["Market Infrastructure", "", "", "", "", ""],
        ["Clearing/Settlement", "", "", "", "", ""],
        ["Cloud Provider", "", "", "", "", ""],
        ["Interbank Exposure", "", "", "", "", ""],
    ]
    for i, row_data in enumerate(sy_rows):
        write_data_row(ws, 23 + i, row_data, fills=[None, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL, None])


def build_risk_inventory(wb):
    ws = wb.create_sheet("15. Risk Inventory")
    ws.sheet_properties.tabColor = "666666"

    cols = {
        "A": 12, "B": 15, "C": 30, "D": 15, "E": 18,
        "F": 10, "G": 25, "H": 10, "I": 10,
        "J": 12, "K": 20, "L": 20, "M": 12, "N": 14
    }
    set_col_widths(ws, cols)

    write_title(ws, 1, "Fourteen-Field Risk Inventory", "Book Reference: Chapter 11 — The master register. All other tabs feed into this.")
    write_instruction(ws, 3, "This is the central risk inventory. One row per risk. Material risks (column J = Yes) require a full Risk Profile (Tab 16).")

    headers = [
        "Risk ID", "Taxonomy (L1/L2/L3)", "Risk Definition", "COSO Category",
        "Risk Owner", "Inherent Score", "Key Controls", "Control Effect.",
        "Residual Score", "Material?", "Data Quality", "Interactions",
        "Trend", "Next Review Date"
    ]
    write_header_row(ws, 5, headers)

    # Example row
    example = [
        "CR-001", "Credit / Counterparty / Single-Name",
        "Risk of loss from excessive exposure to a single counterparty",
        "Operations", "Head of Credit Risk",
        3.5, "Daily margining, credit limits, CVA hedging", 3, 2.3,
        "Yes", "High", "→MR-001(T), →LR-001(A)",
        "Stable", "2026-06-30"
    ]
    write_data_row(ws, 6, example)

    for r in range(7, 60):
        for col in range(1, 15):
            cell = ws.cell(row=r, column=col)
            cell.border = THIN_BORDER
            cell.fill = LIGHT_GOLD_FILL

    # Validations
    dv_coso = DataValidation(type="list", formula1='"Strategic,Operations,Reporting,Compliance"', allow_blank=True)
    ws.add_data_validation(dv_coso)
    dv_coso.add("D6:D200")

    dv_mat = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(dv_mat)
    dv_mat.add("J6:J200")

    dv_dq = DataValidation(type="list", formula1='"High,Medium,Low,Very Low"', allow_blank=True)
    ws.add_data_validation(dv_dq)
    dv_dq.add("K6:K200")

    dv_trend = DataValidation(type="list", formula1='"Increasing,Stable,Decreasing"', allow_blank=True)
    ws.add_data_validation(dv_trend)
    dv_trend.add("M6:M200")


def build_risk_profile(wb):
    ws = wb.create_sheet("16. Risk Profile")
    ws.sheet_properties.tabColor = "666666"

    set_col_widths(ws, {"A": 30, "B": 70})

    write_title(ws, 1, "Risk Profile Template (14 Elements)", "Book Reference: Chapter 11 — Complete one per material risk")

    ws.cell(row=3, column=1, value="Risk ID:").font = SUBHEADER_FONT
    ws.cell(row=3, column=2).fill = LIGHT_GOLD_FILL
    ws.cell(row=3, column=2).border = THIN_BORDER

    elements = [
        ("1. Risk Definition", "Plain-language description of what this risk is and its scope"),
        ("2. Taxonomy Classification", "L1 / L2 / L3 classification"),
        ("3. COSO Objective Category", "Strategic / Operations / Reporting / Compliance"),
        ("4. Underlying Drivers", "Direct and indirect drivers (use Ishikawa analysis)"),
        ("5. Current Exposure Metrics", "Dollar exposure, notional amounts, portfolio concentrations, customer counts"),
        ("6. Risk Appetite Position", "Board-approved appetite level and current position against it"),
        ("7. Key Controls", "Preventive and detective controls with type and effectiveness rating"),
        ("8. Cost-Benefit Assessment", "Analysis of control costs vs risk reduction benefit"),
        ("9. Key Risk Indicators", "Leading indicators with current values and G/A/R thresholds"),
        ("10. Scenario Linkage", "Which ICAAP/ILAAP/CCAR stress scenarios map to this risk"),
        ("11. Data Quality Assessment", "Rating (H/M/L/VL) with evidence basis and conservatism adjustment"),
        ("12. Risk Interactions", "Cross-references to correlated, triggering, or amplifying risks"),
        ("13. Trend and Outlook", "Direction of travel and forward-looking assessment"),
        ("14. Risk Owner", "Named individual accountable (not a committee)"),
    ]

    row = 5
    for element, description in elements:
        ws.cell(row=row, column=1, value=element).font = SUBHEADER_FONT
        ws.cell(row=row, column=1).fill = LIGHT_TEAL_FILL
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2, value=description).font = SMALL_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER
        row += 1
        # Input area
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2).fill = LIGHT_GOLD_FILL
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=2).alignment = WRAP
        ws.row_dimensions[row].height = 45
        row += 1


def build_kri_dashboard(wb):
    ws = wb.create_sheet("17. KRI Dashboard")
    ws.sheet_properties.tabColor = "666666"

    set_col_widths(ws, {"A": 12, "B": 25, "C": 30, "D": 15, "E": 12, "F": 12, "G": 12, "H": 12, "I": 15})

    write_title(ws, 1, "Key Risk Indicator Dashboard", "Book Reference: Chapter 11, 13 — Leading indicators with traffic-light thresholds")

    write_header_row(ws, 4, ["Risk ID", "Risk Name", "KRI Description", "Current Value", "Green ≤", "Amber ≤", "Red >", "Status", "Trend"])

    for r in range(5, 35):
        for col in range(1, 10):
            cell = ws.cell(row=r, column=col)
            cell.border = THIN_BORDER
            if col in (4, 5, 6, 7):
                cell.fill = LIGHT_GOLD_FILL
            cell.alignment = CENTER if col >= 4 else WRAP

    dv_status = DataValidation(type="list", formula1='"Green,Amber,Red"', allow_blank=True)
    ws.add_data_validation(dv_status)
    dv_status.add("H5:H200")

    dv_trend = DataValidation(type="list", formula1='"Increasing,Stable,Decreasing"', allow_blank=True)
    ws.add_data_validation(dv_trend)
    dv_trend.add("I5:I200")


def build_principal_risk_report(wb):
    ws = wb.create_sheet("18. Principal Risk Report")
    ws.sheet_properties.tabColor = "666666"

    set_col_widths(ws, {"A": 12, "B": 25, "C": 12, "D": 15, "E": 12, "F": 12, "G": 12, "H": 12, "I": 20, "J": 12})

    write_title(ws, 1, "Principal Risk Report — Board Reporting Template (10 Items)", "Book Reference: Chapter 12 — Integration")
    write_instruction(ws, 3, "This report is presented to the Board Risk Committee. It contains the ten required elements per the methodology.")

    write_header_row(ws, 5, [
        "Risk ID", "Risk Name", "Residual Score", "Dominant Dimension",
        "Trend", "Appetite Status", "Data Quality", "New/Changed?",
        "Key KRI Alert", "Action Required"
    ])

    for r in range(6, 30):
        for col in range(1, 11):
            cell = ws.cell(row=r, column=col)
            cell.border = THIN_BORDER
            cell.fill = LIGHT_GOLD_FILL
            cell.alignment = CENTER if col >= 3 else WRAP

    dv_dim = DataValidation(type="list", formula1='"Financial,Regulatory,Reputational,Customer/Operational"', allow_blank=True)
    ws.add_data_validation(dv_dim)
    dv_dim.add("D6:D200")

    dv_appetite = DataValidation(type="list", formula1='"Within Appetite,Approaching Limit,Exceeds Appetite"', allow_blank=True)
    ws.add_data_validation(dv_appetite)
    dv_appetite.add("F6:F200")


def build_stress_mapping(wb):
    ws = wb.create_sheet("19. Stress Scenario Map")
    ws.sheet_properties.tabColor = "666666"

    set_col_widths(ws, {"A": 12, "B": 25, "C": 25, "D": 15, "E": 20, "F": 25})

    write_title(ws, 1, "Material Risk to Stress Scenario Mapping", "Book Reference: Chapter 12 — Links inventory to ICAAP/ILAAP/CCAR")

    write_header_row(ws, 4, ["Risk ID", "Risk Name", "Stress Scenario", "Framework", "Severity Calibration", "Loss Estimate Basis"])

    for r in range(5, 40):
        for col in range(1, 7):
            cell = ws.cell(row=r, column=col)
            cell.border = THIN_BORDER
            cell.fill = LIGHT_GOLD_FILL

    dv_fw = DataValidation(type="list", formula1='"ICAAP,ILAAP,CCAR,Recovery Plan,Reverse Stress Test"', allow_blank=True)
    ws.add_data_validation(dv_fw)
    dv_fw.add("D5:D200")


def build_reg_gap(wb):
    ws = wb.create_sheet("20. Reg vs Econ Gap")
    ws.sheet_properties.tabColor = "666666"

    set_col_widths(ws, {"A": 12, "B": 25, "C": 25, "D": 25, "E": 20, "F": 25})

    write_title(ws, 1, "Regulatory vs Economic Risk Gap Analysis", "Book Reference: Chapter 12, 15 — Four-step comparison")

    write_header_row(ws, 4, ["Risk ID", "Risk Name", "Regulatory Treatment (Pillar 1)", "Economic Assessment (Pillar 2)", "Gap Identified?", "Pillar 2A Action"])

    for r in range(5, 35):
        for col in range(1, 7):
            cell = ws.cell(row=r, column=col)
            cell.border = THIN_BORDER
            cell.fill = LIGHT_GOLD_FILL

    dv_gap = DataValidation(type="list", formula1='"No gap,Under-captured in Pillar 1,Not captured in Pillar 1,Over-captured in Pillar 1"', allow_blank=True)
    ws.add_data_validation(dv_gap)
    dv_gap.add("E5:E200")


def build_emerging_risk_register(wb):
    ws = wb.create_sheet("21. Emerging Risks")
    ws.sheet_properties.tabColor = "2D8B2D"

    set_col_widths(ws, {"A": 12, "B": 30, "C": 15, "D": 15, "E": 15, "F": 30, "G": 15})

    write_title(ws, 1, "Emerging Risk Register", "Book Reference: Chapter 6, 13 — Risks not yet material but requiring monitoring")

    write_header_row(ws, 4, ["ER-ID", "Risk Description", "Source", "Horizon", "Potential Impact", "Monitoring Actions", "Review Date"])

    for r in range(5, 25):
        for col in range(1, 8):
            cell = ws.cell(row=r, column=col)
            cell.border = THIN_BORDER
            cell.fill = LIGHT_GOLD_FILL

    dv_source = DataValidation(type="list", formula1='"Delphi Panel,Workshop,Horizon Scan,Regulatory Alert,Industry Event,Internal Intelligence"', allow_blank=True)
    ws.add_data_validation(dv_source)
    dv_source.add("C5:C200")

    dv_horizon = DataValidation(type="list", formula1='"<1 year,1-3 years,3-5 years,>5 years"', allow_blank=True)
    ws.add_data_validation(dv_horizon)
    dv_horizon.add("D5:D200")


def build_assumption_register(wb):
    ws = wb.create_sheet("22. Assumption Register")
    ws.sheet_properties.tabColor = "2D8B2D"

    set_col_widths(ws, {"A": 8, "B": 35, "C": 20, "D": 20, "E": 30, "F": 15})

    write_title(ws, 1, "Assumption Register", "Book Reference: Chapter 6, 11, 13 — Tracks key assumptions underpinning risk assessments")

    write_header_row(ws, 4, ["#", "Assumption", "Risks Affected", "Evidence Basis", "Challenge / Sensitivity", "Review Date"])

    for r in range(5, 25):
        for col in range(1, 7):
            cell = ws.cell(row=r, column=col)
            cell.border = THIN_BORDER
            cell.fill = LIGHT_GOLD_FILL


def build_disagreement_log(wb):
    ws = wb.create_sheet("23. Disagreement Log")
    ws.sheet_properties.tabColor = "2D8B2D"

    set_col_widths(ws, {"A": 8, "B": 20, "C": 30, "D": 30, "E": 20, "F": 15})

    write_title(ws, 1, "Disagreement Log", "Book Reference: Chapter 6, 9 — Preserves divergent views rather than averaging them away")

    write_header_row(ws, 4, ["#", "Risk / Topic", "Position A", "Position B", "Resolution", "Status"])

    for r in range(5, 25):
        for col in range(1, 7):
            cell = ws.cell(row=r, column=col)
            cell.border = THIN_BORDER
            cell.fill = LIGHT_GOLD_FILL

    dv_status = DataValidation(type="list", formula1='"Open,Evidence Requested,Resolved - Position A,Resolved - Position B,Escalated to CRO"', allow_blank=True)
    ws.add_data_validation(dv_status)
    dv_status.add("F5:F200")


def build_event_triggers(wb):
    ws = wb.create_sheet("24. Event-Driven Triggers")
    ws.sheet_properties.tabColor = "2D8B2D"

    set_col_widths(ws, {"A": 25, "B": 35, "C": 30, "D": 20, "E": 15})

    write_title(ws, 1, "Event-Driven Trigger Framework", "Book Reference: Chapter 13, 16 — Six categories requiring immediate risk identification update")

    write_header_row(ws, 4, ["Trigger Category", "Description", "Examples", "Response Required", "Owner"])

    triggers = [
        ["1. Market Dislocation", "Significant market events affecting risk profile", "Interest rate shock, credit spread widening, FX crisis, equity crash", "Immediate risk inventory review within 48 hours", ""],
        ["2. Regulatory Change", "New regulation, enforcement action, or supervisory communication", "New capital requirements, sanctions changes, conduct findings", "Impact assessment within 1 week", ""],
        ["3. Peer Event", "Material loss or failure at comparable institution", "Peer bank failure, industry-wide fraud discovery, systemic event", "Relevance assessment within 1 week", ""],
        ["4. Internal Incident", "Material operational loss, near-miss, or control failure", "System outage, fraud attempt, significant error, compliance breach", "Root cause analysis + forward-looking assessment", ""],
        ["5. Strategic Change", "M&A, new product/market entry, business model change", "Acquisition, new geography, product launch, major outsourcing", "Pre-decision risk assessment using methodology", ""],
        ["6. External Environment Shift", "Geopolitical, macroeconomic, or climate event", "War, pandemic, sovereign crisis, natural disaster", "PESTLE refresh + workshop within 2 weeks", ""],
    ]
    for i, row_data in enumerate(triggers):
        write_data_row(ws, 5 + i, row_data, fills=[None, None, None, None, LIGHT_GOLD_FILL])


def build_process_performance(wb):
    ws = wb.create_sheet("25. Process KPIs")
    ws.sheet_properties.tabColor = "2D8B2D"

    set_col_widths(ws, {"A": 35, "B": 20, "C": 15, "D": 15, "E": 15, "F": 20})

    write_title(ws, 1, "Process Performance Indicators", "Book Reference: Chapter 3, 12, 13 — KPIs tracking the process itself")

    write_header_row(ws, 4, ["Indicator", "Target", "Q1 Actual", "Q2 Actual", "Q3 Actual", "Q4 Actual"])

    kpis = [
        ["Bottom-up submission completion rate (%)", "100%"],
        ["Average days to complete bottom-up submissions", "≤ 15 days"],
        ["Workshop attendance rate (% of required participants)", "≥ 90%"],
        ["Gap analysis completion within deadline", "100%"],
        ["Risk profiles updated within 30 days of assessment", "100%"],
        ["KRI reporting timeliness (% on time)", "≥ 95%"],
        ["Data Quality: % of material risks rated High", "≥ 60%"],
        ["Emerging risks identified per cycle", "≥ 3"],
        ["Scoring calibration variance (cross-BU)", "≤ 0.5"],
        ["Internal Audit findings outstanding (count)", "0"],
        ["Days from event trigger to risk review completion", "≤ 5 days"],
        ["Board report delivered on schedule", "100%"],
    ]
    for i, row_data in enumerate(kpis):
        write_data_row(ws, 5 + i, row_data + ["", "", "", ""],
                      fills=[None, None, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL])


def build_training_matrix(wb):
    ws = wb.create_sheet("26. Training Programme")
    ws.sheet_properties.tabColor = "2D8B2D"

    set_col_widths(ws, {"A": 25, "B": 40, "C": 15, "D": 15, "E": 15})

    write_title(ws, 1, "Training Programme (Five-Level)", "Book Reference: Chapter 13 — Structured training by audience")

    write_header_row(ws, 4, ["Audience", "Content", "Frequency", "Delivery", "Status"])

    training = [
        ["Board / Board Risk Committee", "Methodology overview, how to challenge risk assessments, appetite framework, principal risk report interpretation", "Annual + ad hoc", "Presentation", ""],
        ["CRO Function / Risk Identification Team", "Full methodology, scoring calibration, facilitation skills, reconciliation process, regulatory requirements", "Annual + quarterly refresher", "Workshop", ""],
        ["Business Unit Risk Leads", "Bottom-up template completion, specialist assessments, data quality standards, KRI design", "Annual", "Workshop", ""],
        ["Front-Line Employees", "Risk awareness, how to identify and report risks, escalation process, anonymous channels", "Annual", "e-Learning", ""],
        ["Internal Audit", "Methodology in detail, audit criteria, seven-area programme, independence requirements", "Annual", "Workshop", ""],
    ]
    for i, row_data in enumerate(training):
        write_data_row(ws, 5 + i, row_data, fills=[None, None, None, None, LIGHT_GOLD_FILL])


def build_audit_programme(wb):
    ws = wb.create_sheet("27. Audit Programme")
    ws.sheet_properties.tabColor = "2D8B2D"

    set_col_widths(ws, {"A": 8, "B": 22, "C": 40, "D": 15, "E": 25, "F": 15})

    write_title(ws, 1, "Internal Audit Annual Programme (Seven Areas)", "Book Reference: Chapter 13 — Independent assurance over risk identification")

    write_header_row(ws, 4, ["Area", "Audit Domain", "Key Questions", "Rating", "Findings", "Action Due"])

    areas = [
        [1, "Completeness", "Does the inventory cover all material risks? Are there gaps vs taxonomy, regulatory expectations, industry losses?"],
        [2, "Quality", "Are risk definitions clear? Are scores evidence-based? Is the methodology consistently applied?"],
        [3, "Reconciliation", "Was the top-down / bottom-up reconciliation documented? Were gaps resolved? Was CRO sign-off obtained?"],
        [4, "Data Quality", "Are data quality ratings accurate? Are conservatism adjustments applied where required?"],
        [5, "Documentation", "Are risk profiles complete for all material risks? Is the audit trail maintained? Are changes tracked?"],
        [6, "Integration", "Does the inventory feed capital planning? Are stress scenario linkages documented? Is Board reporting complete?"],
        [7, "Continual Improvement", "Are lessons learned actioned? Are process KPIs tracked? Has the taxonomy been updated?"],
    ]
    for i, row_data in enumerate(areas):
        write_data_row(ws, 5 + i, row_data + ["", "", ""],
                      fills=[None, None, None, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL])

    dv_rating = DataValidation(type="list", formula1='"Satisfactory,Needs Improvement,Unsatisfactory"', allow_blank=True)
    ws.add_data_validation(dv_rating)
    dv_rating.add("D5:D20")


def build_reg_traceability(wb):
    ws = wb.create_sheet("28. Reg Traceability")
    ws.sheet_properties.tabColor = "2D8B2D"

    set_col_widths(ws, {"A": 30, "B": 18, "C": 18, "D": 18, "E": 18, "F": 18, "G": 18})

    write_title(ws, 1, "Regulatory Traceability Matrix", "Book Reference: Chapter 15 — Maps methodology phases to regulatory requirements")

    write_header_row(ws, 4, [
        "Regulatory Framework", "Phase 1\nFoundation", "Phase 2\nIdentification",
        "Phase 3\nReconciliation", "Phase 4\nAssessment", "Phase 5\nInteraction",
        "Phase 6\nDocumentation"
    ])

    frameworks = [
        "ISO 31000:2018",
        "ISO 31010:2019",
        "COSO ERM (2017)",
        "BCBS Corporate Governance (2015)",
        "PRA SS31/15 (UK)",
        "Fed SR 15-18 (US)",
        "ECB Guide to ICAAP (EU)",
        "EBA GL on ICAAP/ILAAP",
        "CRD V / CRR II",
        "DORA (EU)",
        "EBA GL on ICT Risk",
        "EBA Outsourcing Guidelines",
        "EU AMLD6",
        "FCA SYSC (UK)",
        "SM&CR (UK)",
        "APRA CPS 220 (AU)",
    ]
    for i, fw in enumerate(frameworks):
        row_data = [fw, "", "", "", "", "", ""]
        write_data_row(ws, 5 + i, row_data,
                      fills=[None, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL])

    dv_comply = DataValidation(type="list", formula1='"✓ Direct,✓ Indirect,◐ Partial,— N/A"', allow_blank=True)
    ws.add_data_validation(dv_comply)
    dv_comply.add(f"B5:G{4 + len(frameworks)}")


def build_lessons_learned(wb):
    ws = wb.create_sheet("29. Lessons Learned")
    ws.sheet_properties.tabColor = "2D8B2D"

    set_col_widths(ws, {"A": 8, "B": 15, "C": 35, "D": 35, "E": 20, "F": 15})

    write_title(ws, 1, "Lessons-Learned Review", "Book Reference: Chapter 3, 13 — Annual review of process effectiveness")

    write_header_row(ws, 4, ["#", "Area", "What Worked Well", "What Needs Improvement", "Action / Owner", "Status"])

    areas = [
        "Workshop Design & Facilitation",
        "Bottom-Up Template Quality",
        "Reconciliation Process",
        "Scoring Consistency",
        "Data Quality",
        "Board Reporting",
        "Regulatory Alignment",
        "Technology Support",
        "Stakeholder Engagement",
        "Timeline / Timeliness",
    ]
    for i, area in enumerate(areas):
        write_data_row(ws, 5 + i, [i + 1, area, "", "", "", ""],
                      fills=[None, None, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL, LIGHT_GOLD_FILL])


def build_industry_loss_database(wb):
    """Build the Industry Loss Database tab from the sanitized appendix."""
    ws = wb.create_sheet("30. Industry Loss DB")
    ws.sheet_properties.tabColor = NAVY

    set_col_widths(ws, {
        "A": 6, "B": 38, "C": 22, "D": 8,
        "E": 50, "F": 18, "G": 14, "H": 24, "I": 14
    })

    write_title(ws, 1, "Industry Loss Database (179 Events)",
                "Book Reference: Chapter 16, Appendix A — Empirical foundation for the methodology")
    write_instruction(ws, 3,
        "179 bank and financial institution loss events spanning 35 countries and six decades. "
        "Use Excel auto-filter to sort and filter by any column. USD ($M) column enables numeric sorting by loss magnitude.")

    headers = ["#", "Institution", "Country", "Year", "Risk Event",
               "L1 Category", "COSO", "Est. Loss (Original)", "USD ($M)"]
    write_header_row(ws, 5, headers)

    # Parse events from appendix
    appendix_path = os.path.join(os.path.dirname(__file__), "sanitized", "appendix_a.md")
    if not os.path.exists(appendix_path):
        appendix_path = os.path.join(os.path.dirname(__file__), "appendix_a.md")

    events = []
    with open(appendix_path, "r", encoding="utf-8") as f:
        in_event_table = False
        for line in f:
            line = line.strip()
            if line.startswith("| # |"):
                in_event_table = True
                continue
            if in_event_table and line.startswith("|--"):
                continue
            if in_event_table and line.startswith("|"):
                parts = [p.strip() for p in line.split("|")[1:-1]]
                if len(parts) >= 9 and parts[0].isdigit():
                    events.append(parts)
            elif in_event_table and not line.startswith("|"):
                break

    # Write data rows with alternating shading
    alt_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    for i, parts in enumerate(events):
        row = 6 + i
        num = int(parts[0])
        year = int(parts[3])
        usd_millions = parse_usd_to_millions(parts[8])

        values = [num, parts[1], parts[2], year, parts[4],
                  parts[5], parts[6], parts[7], usd_millions]

        fill = alt_fill if i % 2 == 1 else None
        fills = [fill] * 9 if fill else None
        write_data_row(ws, row, values, fills=fills)

        # Center-align #, Year, USD columns
        ws.cell(row=row, column=1).alignment = CENTER
        ws.cell(row=row, column=4).alignment = CENTER
        ws.cell(row=row, column=9).alignment = CENTER
        ws.cell(row=row, column=9).number_format = '#,##0'

    # Auto-filter for sorting
    last_row = 5 + len(events)
    ws.auto_filter.ref = f"A5:I{last_row}"

    # Freeze header row
    ws.freeze_panes = "A6"

    return len(events)


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    wb = Workbook()

    print("Building template pack...")

    build_cover(wb)
    print("  ✓ Cover")

    build_taxonomy(wb)
    print("  ✓ 1. Risk Taxonomy")

    build_regulatory_mapping(wb)
    print("  ✓ 2. Regulatory Mapping")

    build_pestle(wb)
    print("  ✓ 3. PESTLE Assessment")

    build_internal_environment(wb)
    print("  ✓ 4. Internal Environment")

    build_risk_criteria(wb)
    print("  ✓ 5. Risk Criteria")

    build_workshop_planner(wb)
    print("  ✓ 6. Workshop Planner")

    build_swift_prompt_matrix(wb)
    print("  ✓ 7. SWIFT Prompt Matrix")

    build_bottom_up_template(wb)
    print("  ✓ 8. Bottom-Up Template")

    build_pre_workshop(wb)
    print("  ✓ 9. Pre-Workshop Assessment")

    build_gap_analysis(wb)
    print("  ✓ 10. Gap Analysis")

    build_scoring_worksheet(wb)
    print("  ✓ 11. Risk Scoring")

    build_interaction_matrix(wb)
    print("  ✓ 12. Interaction Matrix")

    build_bowtie(wb)
    print("  ✓ 13. Bow-Tie Template")

    build_concentration(wb)
    print("  ✓ 14. Concentration Analysis")

    build_risk_inventory(wb)
    print("  ✓ 15. Risk Inventory (Master)")

    build_risk_profile(wb)
    print("  ✓ 16. Risk Profile")

    build_kri_dashboard(wb)
    print("  ✓ 17. KRI Dashboard")

    build_principal_risk_report(wb)
    print("  ✓ 18. Principal Risk Report")

    build_stress_mapping(wb)
    print("  ✓ 19. Stress Scenario Mapping")

    build_reg_gap(wb)
    print("  ✓ 20. Regulatory vs Economic Gap")

    build_emerging_risk_register(wb)
    print("  ✓ 21. Emerging Risk Register")

    build_assumption_register(wb)
    print("  ✓ 22. Assumption Register")

    build_disagreement_log(wb)
    print("  ✓ 23. Disagreement Log")

    build_event_triggers(wb)
    print("  ✓ 24. Event-Driven Triggers")

    build_process_performance(wb)
    print("  ✓ 25. Process Performance KPIs")

    build_training_matrix(wb)
    print("  ✓ 26. Training Programme")

    build_audit_programme(wb)
    print("  ✓ 27. Audit Programme")

    build_reg_traceability(wb)
    print("  ✓ 28. Regulatory Traceability")

    build_lessons_learned(wb)
    print("  ✓ 29. Lessons Learned")

    event_count = build_industry_loss_database(wb)
    print(f"  ✓ 30. Industry Loss Database ({event_count} events)")

    # Save
    output_dir = os.path.join(os.path.dirname(__file__), "output")
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, "Risk_Identification_Template_Pack.xlsx")
    wb.save(output_path)

    size_kb = os.path.getsize(output_path) / 1024
    print(f"\n  Template pack saved: {output_path} ({size_kb:.0f} KB)")
    print(f"  Sheets: {len(wb.sheetnames)}")
    print(f"  Tabs: {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    main()
