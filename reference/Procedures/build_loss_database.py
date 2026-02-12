#!/usr/bin/env python3
"""
Build Industry Loss Database
EON Risk Services Ltd

Parses Historical Risk Identification Failures PDF, enriches with taxonomy,
narratives, and standardised fields, then outputs Excel, CSV, and Markdown.
"""

import re
import subprocess
import os
import sys
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ═══════════════════════════════════════════════════════════════════════════════
# CONFIGURATION
# ═══════════════════════════════════════════════════════════════════════════════

SOURCE_PDF = os.path.expanduser(
    "~/EONRiskServices/reference/Methods of Risk Identification/"
    "Historical Risk Identification Failures.pdf"
)
OUTPUT_DIR = os.path.expanduser(
    "~/EONRiskServices/reference/Procedures/"
)

NAVY = "1B2A4A"
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
BODY_FONT = Font(name="Calibri", size=10)
BODY_ALIGN = Alignment(vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

# ═══════════════════════════════════════════════════════════════════════════════
# STEP 1: PARSE PDF
# ═══════════════════════════════════════════════════════════════════════════════

def extract_text(pdf_path):
    """Extract text from PDF using pdftotext with layout preservation."""
    result = subprocess.run(
        ["pdftotext", "-layout", pdf_path, "-"],
        capture_output=True, text=True, check=True
    )
    return result.stdout


def parse_entries(raw_text):
    """Parse raw text into list of dicts with institution, year, risk_event, loss_str, include."""
    entries = []
    lines = raw_text.strip().split("\n")

    for line in lines:
        line = line.strip()
        if not line:
            continue
        # Skip header lines
        if line.startswith("HISTORICAL") or line.startswith("Financial Institution"):
            continue

        # Pattern: Institution  Year  Risk Event  Loss  Yes/No
        # Use regex to find the year (4-digit number) as anchor
        m = re.match(
            r'^(.+?)\s{2,}(\d{4})\s+(.+?)\s{2,}(.+?)(?:\s{2,}(Yes|No))?\s*$',
            line
        )
        if m:
            institution = m.group(1).strip()
            year = int(m.group(2))
            risk_event = m.group(3).strip()
            loss_str = m.group(4).strip()
            include = m.group(5).strip() if m.group(5) else ""
            entries.append({
                'institution': institution,
                'year': year,
                'risk_event': risk_event,
                'loss_str': loss_str,
                'include': include,
            })
        else:
            print(f"  WARNING: Could not parse line: {line[:80]}...")

    return entries


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 2: COUNTRY LOOKUP
# ═══════════════════════════════════════════════════════════════════════════════

# Parenthetical patterns in institution names
COUNTRY_FROM_NAME = {
    "Malaysia": "Malaysia",
    "US": "United States",
    "USA": "United States",
    "Ireland": "Ireland",
    "Greece": "Greece",
    "Australia": "Australia",
    "Italy": "Italy",
    "Andorra": "Andorra",
    "Spain": "Spain",
    "Portugal": "Portugal",
    "Dominican Republic": "Dominican Republic",
    "Costa Rica": "Costa Rica",
    "Russia": "Russia",
    "Germany": "Germany",
    "UK": "United Kingdom",
    "Belgium": "Belgium",
    "Belgium/Netherlands": "Belgium",
    "France": "France",
    "Switzerland": "Switzerland",
    "Denmark": "Denmark",
    "Singapore": "Singapore",
    "India": "India",
    "Iceland": "Iceland",
    "China": "China",
    "Japan": "Japan",
    "Pakistan": "Pakistan",
    "Netherlands": "Netherlands",
    "Sweden": "Sweden",
    "Canada": "Canada",
    "Nigeria": "Nigeria",
    "Cyprus": "Cyprus",
    "Ukraine": "Ukraine",
    "South Africa": "South Africa",
    "Malta": "Malta",
    "Austria": "Austria",
    "UAE": "United Arab Emirates",
    "London Whale": "United States",  # JPMorgan
    "3CIF": "France",  # Crédit Immobilier de France
}

# Hardcoded lookup for institutions without parenthetical country
INSTITUTION_COUNTRY = {
    "Adelphia Communications": "United States",
    "Allied Irish Bank": "Ireland",
    "AMCORE Financial": "United States",
    "Anglo Irish Bank": "Ireland",
    "Arthur Andersen": "United States",
    "Bank of America": "United States",
    "Bank of New England": "United States",
    "BankUnited FSB": "United States",
    "Barings Bank": "United Kingdom",
    "Bear Stearns": "United States",
    "Bernard L. Madoff Investment Securities LLC": "United States",
    "CIT Group": "United States",
    "Citigroup": "United States",
    "Colonial BancGroup": "United States",
    "Countrywide Financial": "United States",
    "Enron": "United States",
    "Fannie Mae": "United States",
    "Fannie Mae and Freddie Mac": "United States",
    "Freddie Mac": "United States",
    "Fremont General Corporation": "United States",
    "Global Crossing": "United States",
    "HealthSouth Corporation": "United States",
    "HSBC": "United Kingdom",
    "Hwang's Archegos Capital Management": "United States",
    "IndyMac Bancorp": "United States",
    "Irwin Financial Corporation": "United States",
    "Lehman Brothers": "United States",
    "Lincoln Savings and Loan": "United States",
    "Long-Term Capital Management": "United States",
    "Merrill Lynch": "United States",
    "Metropolitan Life Insurance Company": "United States",
    "MF Global": "United States",
    "New Century Financial Corporation": "United States",
    "Parmalat": "Italy",
    "Refco": "United States",
    "Stanford Financial Group": "United States",
    "Tyco International": "United States",
    "Vivendi": "France",
    "Washington Mutual": "United States",
    "WorldCom": "United States",
    "American International Group (AIG)": "United States",
    "American International Group": "United States",
    "American Savings & Loan": "United States",
    "Bank of Cyprus": "Cyprus",
    "Banesto": "Spain",  # Banesto (Spain) in text but no parens
    "Banco do Brasil": "Brazil",
    "Carillion": "United Kingdom",  # text says Carillion (UK) but parsing may strip
    "Icelandic Banks (Kaupthing, Landsbanki, Glitnir)": "Iceland",
    "Caixa Economica Federal": "Brazil",  # text has (Brazil) but just in case
    "Petrobras": "Brazil",  # text has (Brazil)
    "Banco Intercontinental": "Dominican Republic",
    "Banco Nacional": "Costa Rica",
    "Luckin Coffee": "China",
    "Maybank": "Malaysia",
    "Olympus Corporation": "Japan",
    "Toshiba": "Japan",
    "Banco de Valencia": "Spain",
    "Banco Popular": "Spain",
    "Bankia": "Spain",
    "Caja Madrid": "Spain",
    "Dexia": "Belgium",
    "Postbank": "Ireland",  # Irish Postbank
    "Rabobank": "Netherlands",
}

REGION_MAP = {
    "United States": "North America",
    "Canada": "North America",
    "Costa Rica": "Latin America",
    "Brazil": "Latin America",
    "Dominican Republic": "Latin America",
    "United Kingdom": "Europe",
    "Ireland": "Europe",
    "France": "Europe",
    "Germany": "Europe",
    "Switzerland": "Europe",
    "Italy": "Europe",
    "Spain": "Europe",
    "Portugal": "Europe",
    "Greece": "Europe",
    "Belgium": "Europe",
    "Netherlands": "Europe",
    "Austria": "Europe",
    "Denmark": "Europe",
    "Sweden": "Europe",
    "Iceland": "Europe",
    "Cyprus": "Europe",
    "Malta": "Europe",
    "Andorra": "Europe",
    "Russia": "Europe",
    "Ukraine": "Europe",
    "Australia": "Asia-Pacific",
    "Singapore": "Asia-Pacific",
    "Japan": "Asia-Pacific",
    "China": "Asia-Pacific",
    "India": "Asia-Pacific",
    "Malaysia": "Asia-Pacific",
    "Pakistan": "Asia-Pacific",
    "South Korea": "Asia-Pacific",
    "South Africa": "Africa",
    "Nigeria": "Africa",
    "United Arab Emirates": "Middle East",
}


def extract_country(institution):
    """Extract country from institution name parenthetical or lookup table."""
    # Try parenthetical extraction
    paren = re.search(r'\(([^)]+)\)\s*$', institution)
    if paren:
        key = paren.group(1).strip()
        if key in COUNTRY_FROM_NAME:
            return COUNTRY_FROM_NAME[key]

    # Try hardcoded lookup (exact match first)
    if institution in INSTITUTION_COUNTRY:
        return INSTITUTION_COUNTRY[institution]

    # Try partial match on institution name
    for inst_key, country in INSTITUTION_COUNTRY.items():
        if inst_key in institution or institution in inst_key:
            return country

    # Try parenthetical content as direct country name
    if paren:
        key = paren.group(1).strip()
        # Check if it looks like a country name
        for region_country in REGION_MAP:
            if key.lower() in region_country.lower() or region_country.lower() in key.lower():
                return region_country

    return "Unknown"


def get_region(country):
    """Map country to region."""
    return REGION_MAP.get(country, "Other")


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 3: CURRENCY PARSING AND FX CONVERSION
# ═══════════════════════════════════════════════════════════════════════════════

# Approximate FX rates to USD (indicative, for comparability only)
FX_TO_USD = {
    "USD": 1.0,
    "EUR": 1.18,
    "GBP": 1.35,
    "AUD": 0.72,
    "CHF": 1.10,
    "DKK": 0.16,
    "SGD": 0.74,
    "CNY": 0.15,
    "MYR": 0.24,
    "SEK": 0.11,
    "RUB": 0.013,
    "ZAR": 0.06,    # South African Rand (R)
    "CAD": 0.78,
}

CURRENCY_PATTERNS = [
    (r'^\$', "USD"),
    (r'^€', "EUR"),
    (r'^£', "GBP"),
    (r'^AUD\s', "AUD"),
    (r'^CHF\s', "CHF"),
    (r'^DKK\s', "DKK"),
    (r'^SGD\s', "SGD"),
    (r'^CNY\s', "CNY"),
    (r'^MYR\s', "MYR"),
    (r'^SEK\s', "SEK"),
    (r'^RUB\s', "RUB"),
    (r'^R(?:\d)', "ZAR"),
    (r'CAD\s*$', "CAD"),
    (r'AUD\s*$', "AUD"),
]


def parse_loss(loss_str):
    """Parse loss string into (original_amount_str, currency, usd_amount)."""
    if not loss_str or loss_str.strip() == "":
        return loss_str, "USD", 0.0

    s = loss_str.strip()

    # Detect currency
    currency = "USD"
    for pattern, curr in CURRENCY_PATTERNS:
        if re.search(pattern, s):
            currency = curr
            break

    # Extract numeric value
    # Remove currency symbols and prefixes
    cleaned = re.sub(r'^[\$€£]', '', s)
    cleaned = re.sub(r'^(AUD|CHF|DKK|SGD|CNY|MYR|SEK|RUB|CAD|R)\s*', '', cleaned)

    # Handle special cases like "$30 Billion (600$ Liabilities)"
    cleaned = re.sub(r'\(.*?\)', '', cleaned).strip()
    # Remove trailing currency markers
    cleaned = re.sub(r'\s*(AUD|CAD|USD)\s*$', '', cleaned).strip()

    # Extract number and multiplier
    num_match = re.search(r'([\d,.]+)\s*(billion|million|mil|trillion)?', cleaned, re.IGNORECASE)
    if num_match:
        num_str = num_match.group(1).replace(',', '')
        try:
            amount = float(num_str)
        except ValueError:
            return loss_str, currency, 0.0

        multiplier_str = (num_match.group(2) or "").lower()
        if multiplier_str in ("billion",):
            amount *= 1e9
        elif multiplier_str in ("million", "mil"):
            amount *= 1e6
        elif multiplier_str in ("trillion",):
            amount *= 1e12

        usd_amount = amount * FX_TO_USD.get(currency, 1.0)
        return loss_str, currency, usd_amount

    return loss_str, currency, 0.0


def loss_bucket(usd_amount):
    """Categorise USD loss into magnitude bucket."""
    if usd_amount <= 0:
        return "Unknown"
    b = usd_amount / 1e9
    if b < 1:
        return "< $1B"
    elif b < 10:
        return "$1-10B"
    elif b < 50:
        return "$10-50B"
    else:
        return "$50B+"


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 4: L1 RISK TAXONOMY (Basel Categories)
# ═══════════════════════════════════════════════════════════════════════════════

L1_KEYWORDS = {
    "Credit Risk": [
        "overexposure", "subprime", "mortgage", "loan", "debt",
        "npls", "non-performing", "credit", "default swap",
        "collateralized", "real estate", "property", "sovereign",
        "infrastructure loan", "bad loan",
    ],
    "Market Risk": [
        "trading", "leverage", "market risk", "proprietary",
        "concentration risk", "securities investment",
    ],
    "Operational Risk": [
        "fraud", "rogue trader", "unauthorized", "ponzi",
        "accounting fraud", "accounting irregularit", "mismanagement",
        "money laundering", "sanctions", "looting", "misappropriation",
        "data breach", "cyber", "it systems", "software glitch",
        "system failure", "corruption", "misrepresentation",
        "false accounting", "hidden loan", "hidden loss",
        "spying", "governance failure", "regulatory capture",
    ],
    "Conduct Risk": [
        "mis-selling", "fake account", "fees for no service",
        "interest rate rigging", "libor", "euribor", "tax evasion",
        "ppi", "scandal", "ethical",
    ],
    "Liquidity Risk": [
        "liquidity", "wholesale funding", "insolvency",
        "funding market",
    ],
    "Strategic Risk": [
        "acquisition", "overexpansion", "expansion",
        "pension deficit", "valuation of liabilities",
        "oil and gas sector", "insurance dispute",
        "debt concealment",
    ],
}


def classify_risk_l1(event_text):
    """Classify risk event into Basel L1 category using keyword matching."""
    text_lower = event_text.lower()

    # Score each category
    scores = {}
    for category, keywords in L1_KEYWORDS.items():
        score = sum(1 for kw in keywords if kw in text_lower)
        if score > 0:
            scores[category] = score

    if not scores:
        return "Operational Risk"  # default

    # Prioritisation rules for ambiguous cases
    # "Fraud" entries should be Operational unless clearly market/credit
    if "fraud" in text_lower or "accounting" in text_lower:
        if "Operational Risk" in scores:
            return "Operational Risk"

    # "Overexposure to X" is primarily Credit Risk
    if "overexposure" in text_lower and "Credit Risk" in scores:
        if "trading" not in text_lower and "leverage" not in text_lower:
            return "Credit Risk"

    # "Unauthorized trading" is Market Risk
    if "unauthorized" in text_lower and "trading" in text_lower:
        return "Market Risk"

    # "Liquidity" or "funding" with insolvency
    if "liquidity" in text_lower or ("funding" in text_lower and "insolvency" in text_lower):
        return "Liquidity Risk"

    # Mis-selling and conduct issues
    if "mis-selling" in text_lower or "fake account" in text_lower or "rigging" in text_lower:
        return "Conduct Risk"

    # Return highest scoring
    return max(scores, key=scores.get)


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 5: NARRATIVE ENRICHMENT DATA
# Each entry: (risk_l2, coso_category, root_cause, risk_id_failure, outcome)
# Indexed by position in the parsed entry list (PDF alphabetical order)
# ═══════════════════════════════════════════════════════════════════════════════

# Import enrichment from companion data module
import importlib.util
_spec = importlib.util.spec_from_file_location(
    "enrichment_data",
    os.path.join(os.path.dirname(os.path.abspath(__file__)), "enrichment_data.py")
)
_mod = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(_mod)
ENRICHMENT = _mod.ENRICHMENT

# ── Post-parse loss amount corrections ──────────────────────────────────────
# The PDF source text contains loss figures that have been corrected through
# fact-checking against primary sources. These overrides align the database
# with the verified book text without requiring PDF edits.
# Format: index → (corrected_original_str, currency, usd_amount)
POST_PARSE_LOSS_CORRECTIONS = {
    5: ("~€4 billion", "EUR", 4_720_000_000.0),        # Alpha Bank: was €10.9B (misattributed — Alpha was one of four Greek banks)
    66: ("$630 million", "USD", 630_000_000.0),          # Deutsche Bank: was $150M (combined UK/US mirror-trading fines)
    160: ("$2.3 billion", "USD", 2_300_000_000.0),       # UBS/Adoboli: was $2B (understated by $300M)
}


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 6: BUILD ENRICHED DATAFRAME
# ═══════════════════════════════════════════════════════════════════════════════

def build_database(entries):
    """Build the full enriched DataFrame from parsed entries."""
    rows = []
    for i, e in enumerate(entries):
        institution = e['institution']
        year = e['year']
        risk_event = e['risk_event']
        loss_str = e['loss_str']
        include = e['include']

        country = extract_country(institution)
        region = get_region(country)
        decade = (year // 10) * 10
        risk_l1 = classify_risk_l1(risk_event)

        orig_loss, currency, usd_amount = parse_loss(loss_str)

        # Apply post-parse corrections where book fact-checking found errors
        if i in POST_PARSE_LOSS_CORRECTIONS:
            orig_loss, currency, usd_amount = POST_PARSE_LOSS_CORRECTIONS[i]

        bucket = loss_bucket(usd_amount)

        # Narrative enrichment
        enrich = ENRICHMENT.get(i, {})
        risk_l2 = enrich.get('risk_l2', '')
        coso_cat = enrich.get('coso_category', '')
        root_cause = enrich.get('root_cause', '')
        risk_id_fail = enrich.get('risk_id_failure', '')
        outcome = enrich.get('outcome', '')

        rows.append({
            'Financial Institution': institution,
            'Country': country,
            'Region': region,
            'Year': year,
            'Decade': f"{decade}s",
            'Risk Event': risk_event,
            'Risk Taxonomy L1': risk_l1,
            'Risk Taxonomy L2': risk_l2,
            'COSO Objective Category': coso_cat,
            'Root Cause Summary': root_cause,
            'Risk Identification Failure': risk_id_fail,
            'Outcome': outcome,
            'Estimated Loss (Original)': orig_loss,
            'Original Currency': currency,
            'Standardized Loss (USD)': usd_amount,
            'Loss Magnitude Bucket': bucket,
            'Include in Inventory': include if include else "No",
        })

    df = pd.DataFrame(rows)
    return df


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 7: BUILD EXCEL WORKBOOK
# ═══════════════════════════════════════════════════════════════════════════════

def auto_width(ws, min_width=10, max_width=50):
    """Set column widths based on content."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        lengths = []
        for cell in col_cells:
            if cell.value:
                lengths.append(len(str(cell.value)))
        if lengths:
            best = min(max(max(lengths), min_width), max_width)
            ws.column_dimensions[col_letter].width = best + 2


def style_header(ws):
    """Apply navy header styling to first row."""
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def style_body(ws):
    """Apply body styling."""
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = BODY_ALIGN
            cell.border = THIN_BORDER


def build_excel(df, output_path):
    """Build formatted Excel workbook with 3 sheets."""
    wb = Workbook()

    # ── Sheet 1: Loss Database ──
    ws1 = wb.active
    ws1.title = "Loss Database"

    # Write headers
    headers = list(df.columns)
    ws1.append(headers)

    # Write data
    for _, row in df.iterrows():
        values = []
        for col in headers:
            val = row[col]
            if col == 'Standardized Loss (USD)' and isinstance(val, (int, float)):
                values.append(round(val, 0))
            else:
                values.append(val)
        ws1.append(values)

    style_header(ws1)
    style_body(ws1)

    # Format USD column as number
    usd_col_idx = headers.index('Standardized Loss (USD)') + 1
    for row_idx in range(2, ws1.max_row + 1):
        cell = ws1.cell(row=row_idx, column=usd_col_idx)
        cell.number_format = '#,##0'

    # Add Excel Table for filtering
    table_ref = f"A1:{get_column_letter(len(headers))}{len(df) + 1}"
    tab = Table(displayName="LossDatabase", ref=table_ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False
    )
    ws1.add_table(tab)

    # Freeze panes (below header, after first column)
    ws1.freeze_panes = "B2"

    # Set reasonable column widths
    col_widths = {
        'Financial Institution': 40,
        'Country': 18,
        'Region': 16,
        'Year': 8,
        'Decade': 8,
        'Risk Event': 55,
        'Risk Taxonomy L1': 18,
        'Risk Taxonomy L2': 25,
        'COSO Objective Category': 18,
        'Root Cause Summary': 55,
        'Risk Identification Failure': 60,
        'Outcome': 40,
        'Estimated Loss (Original)': 22,
        'Original Currency': 12,
        'Standardized Loss (USD)': 20,
        'Loss Magnitude Bucket': 16,
        'Include in Inventory': 14,
    }
    for i, col_name in enumerate(headers, 1):
        ws1.column_dimensions[get_column_letter(i)].width = col_widths.get(col_name, 15)

    # ── Sheet 2: Summary Statistics ──
    ws2 = wb.create_sheet("Summary Statistics")
    _write_summary_sheet(ws2, df)

    # ── Sheet 3: Priority Examples ──
    ws3 = wb.create_sheet("Priority Examples")
    priority = df[df['Include in Inventory'].str.upper() == 'YES'].copy()

    ws3.append(headers)
    for _, row in priority.iterrows():
        values = []
        for col in headers:
            val = row[col]
            if col == 'Standardized Loss (USD)' and isinstance(val, (int, float)):
                values.append(round(val, 0))
            else:
                values.append(val)
        ws3.append(values)

    style_header(ws3)
    style_body(ws3)
    for i, col_name in enumerate(headers, 1):
        ws3.column_dimensions[get_column_letter(i)].width = col_widths.get(col_name, 15)

    # Save
    wb.save(output_path)
    print(f"  Excel: {output_path}")


def _write_summary_sheet(ws, df):
    """Write summary statistics to worksheet."""
    row_num = 1

    def write_section(title, data_dict, start_row):
        ws.cell(row=start_row, column=1, value=title).font = Font(
            name="Calibri", bold=True, size=12, color=NAVY
        )
        r = start_row + 1
        ws.cell(row=r, column=1, value="Category").font = Font(bold=True)
        ws.cell(row=r, column=2, value="Count").font = Font(bold=True)
        ws.cell(row=r, column=3, value="Total USD Loss").font = Font(bold=True)
        r += 1
        for key, (count, total) in data_dict.items():
            ws.cell(row=r, column=1, value=str(key))
            ws.cell(row=r, column=2, value=count)
            c = ws.cell(row=r, column=3, value=round(total, 0))
            c.number_format = '#,##0'
            r += 1
        return r + 1

    # By L1 Risk Type
    l1_stats = {}
    for cat in sorted(df['Risk Taxonomy L1'].unique()):
        subset = df[df['Risk Taxonomy L1'] == cat]
        l1_stats[cat] = (len(subset), subset['Standardized Loss (USD)'].sum())
    row_num = write_section("By Risk Taxonomy (L1)", l1_stats, row_num)

    # By Decade
    decade_stats = {}
    for dec in sorted(df['Decade'].unique()):
        subset = df[df['Decade'] == dec]
        decade_stats[dec] = (len(subset), subset['Standardized Loss (USD)'].sum())
    row_num = write_section("By Decade", decade_stats, row_num)

    # By Region
    region_stats = {}
    for reg in sorted(df['Region'].unique()):
        subset = df[df['Region'] == reg]
        region_stats[reg] = (len(subset), subset['Standardized Loss (USD)'].sum())
    row_num = write_section("By Region", region_stats, row_num)

    # By Loss Magnitude
    bucket_order = ["< $1B", "$1-10B", "$10-50B", "$50B+", "Unknown"]
    bucket_stats = {}
    for b in bucket_order:
        subset = df[df['Loss Magnitude Bucket'] == b]
        if len(subset) > 0:
            bucket_stats[b] = (len(subset), subset['Standardized Loss (USD)'].sum())
    row_num = write_section("By Loss Magnitude", bucket_stats, row_num)

    # By COSO Category
    coso_stats = {}
    for cat in sorted(df['COSO Objective Category'].unique()):
        if not cat:
            continue
        subset = df[df['COSO Objective Category'] == cat]
        coso_stats[cat] = (len(subset), subset['Standardized Loss (USD)'].sum())
    row_num = write_section("By COSO Objective Category", coso_stats, row_num)

    # Top 20 Losses
    ws.cell(row=row_num, column=1, value="Top 20 Losses by USD Amount").font = Font(
        name="Calibri", bold=True, size=12, color=NAVY
    )
    row_num += 1
    top_headers = ["Rank", "Institution", "Year", "Risk Event", "USD Loss", "L1 Category"]
    for j, h in enumerate(top_headers, 1):
        ws.cell(row=row_num, column=j, value=h).font = Font(bold=True)
    row_num += 1
    top20 = df.nlargest(20, 'Standardized Loss (USD)')
    for rank, (_, r) in enumerate(top20.iterrows(), 1):
        ws.cell(row=row_num, column=1, value=rank)
        ws.cell(row=row_num, column=2, value=r['Financial Institution'])
        ws.cell(row=row_num, column=3, value=r['Year'])
        ws.cell(row=row_num, column=4, value=r['Risk Event'])
        c = ws.cell(row=row_num, column=5, value=round(r['Standardized Loss (USD)'], 0))
        c.number_format = '#,##0'
        ws.cell(row=row_num, column=6, value=r['Risk Taxonomy L1'])
        row_num += 1

    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 20


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 8: CSV EXPORT
# ═══════════════════════════════════════════════════════════════════════════════

def export_csv(df, output_path):
    """Export DataFrame to UTF-8 CSV with BOM for Excel compatibility."""
    df.to_csv(output_path, index=False, encoding='utf-8-sig')
    print(f"  CSV:   {output_path}")


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 9: MARKDOWN SUMMARY
# ═══════════════════════════════════════════════════════════════════════════════

def generate_markdown(df, output_path):
    """Generate Markdown summary report."""
    lines = []
    lines.append("# Industry Loss Database — Summary Report")
    lines.append(f"\n*Generated: {datetime.now().strftime('%Y-%m-%d')} | EON Risk Services Ltd*\n")
    lines.append(f"**Total entries:** {len(df)}\n")

    total_usd = df['Standardized Loss (USD)'].sum()
    lines.append(f"**Aggregate estimated losses (USD, indicative):** ${total_usd/1e12:.1f} trillion\n")

    # By Risk Taxonomy L1
    lines.append("## Distribution by Risk Taxonomy (L1)\n")
    lines.append("| Category | Count | % | Aggregate USD Loss |")
    lines.append("|---|---|---|---|")
    for cat in sorted(df['Risk Taxonomy L1'].unique()):
        subset = df[df['Risk Taxonomy L1'] == cat]
        pct = len(subset) / len(df) * 100
        total = subset['Standardized Loss (USD)'].sum()
        lines.append(f"| {cat} | {len(subset)} | {pct:.1f}% | ${total/1e9:.1f}B |")

    # By Decade
    lines.append("\n## Distribution by Decade\n")
    lines.append("| Decade | Count | % | Aggregate USD Loss |")
    lines.append("|---|---|---|---|")
    for dec in sorted(df['Decade'].unique()):
        subset = df[df['Decade'] == dec]
        pct = len(subset) / len(df) * 100
        total = subset['Standardized Loss (USD)'].sum()
        lines.append(f"| {dec} | {len(subset)} | {pct:.1f}% | ${total/1e9:.1f}B |")

    # By Region
    lines.append("\n## Distribution by Region\n")
    lines.append("| Region | Count | % | Aggregate USD Loss |")
    lines.append("|---|---|---|---|")
    for reg in sorted(df['Region'].unique()):
        subset = df[df['Region'] == reg]
        pct = len(subset) / len(df) * 100
        total = subset['Standardized Loss (USD)'].sum()
        lines.append(f"| {reg} | {len(subset)} | {pct:.1f}% | ${total/1e9:.1f}B |")

    # By COSO Category
    lines.append("\n## Distribution by COSO Objective Category\n")
    lines.append("| Category | Count | % |")
    lines.append("|---|---|---|")
    for cat in ["Strategic", "Operations", "Reporting", "Compliance"]:
        subset = df[df['COSO Objective Category'] == cat]
        if len(subset) > 0:
            pct = len(subset) / len(df) * 100
            lines.append(f"| {cat} | {len(subset)} | {pct:.1f}% |")

    # Timeline Patterns
    lines.append("\n## Key Timeline Patterns\n")

    # S&L Crisis
    sl = df[(df['Year'] >= 1985) & (df['Year'] <= 1995)]
    lines.append(f"**S&L Crisis Era (1985-1995):** {len(sl)} entries\n")

    # GFC
    gfc = df[(df['Year'] >= 2007) & (df['Year'] <= 2010)]
    lines.append(f"**Global Financial Crisis (2007-2010):** {len(gfc)} entries — "
                 f"${gfc['Standardized Loss (USD)'].sum()/1e12:.1f} trillion aggregate loss\n")

    # Conduct era
    conduct = df[(df['Year'] >= 2012) & (df['Year'] <= 2020) &
                 (df['Risk Taxonomy L1'].isin(['Conduct Risk', 'Operational Risk']))]
    lines.append(f"**Post-GFC Conduct/Operational Era (2012-2020):** {len(conduct)} entries\n")

    # Top 10 losses
    lines.append("\n## Top 10 Losses\n")
    lines.append("| Rank | Institution | Year | Event | USD Loss |")
    lines.append("|---|---|---|---|---|")
    top10 = df.nlargest(10, 'Standardized Loss (USD)')
    for rank, (_, r) in enumerate(top10.iterrows(), 1):
        usd = r['Standardized Loss (USD)']
        lines.append(f"| {rank} | {r['Financial Institution']} | {r['Year']} | "
                     f"{r['Risk Event'][:50]} | ${usd/1e9:.1f}B |")

    # Priority Examples with full narratives
    lines.append("\n## Priority Examples (Inventory Selections)\n")
    priority = df[df['Include in Inventory'].str.upper() == 'YES']
    lines.append(f"*{len(priority)} entries selected for the example inventory.*\n")

    for _, r in priority.iterrows():
        lines.append(f"### {r['Financial Institution']} ({r['Year']})\n")
        lines.append(f"- **Risk Event:** {r['Risk Event']}")
        lines.append(f"- **L1 / L2:** {r['Risk Taxonomy L1']} / {r['Risk Taxonomy L2']}")
        lines.append(f"- **COSO Category:** {r['COSO Objective Category']}")
        lines.append(f"- **Estimated Loss:** {r['Estimated Loss (Original)']} "
                     f"(~${r['Standardized Loss (USD)']/1e9:.1f}B USD)")
        lines.append(f"- **Root Cause:** {r['Root Cause Summary']}")
        lines.append(f"- **Risk Identification Failure:** {r['Risk Identification Failure']}")
        lines.append(f"- **Outcome:** {r['Outcome']}")
        lines.append("")

    # Common Risk Identification Failure Modes
    lines.append("\n## Common Risk Identification Failure Modes\n")
    lines.append("Analysis of the database reveals recurring patterns in how risk "
                 "identification processes failed:\n")
    failure_modes = [
        ("Concentration Blindness", "Failure to identify portfolio or sector concentration "
         "as a systemic risk, often masked by diversification narratives."),
        ("Model Overreliance", "Excessive trust in quantitative models (VaR, credit ratings) "
         "that failed to capture tail risks or correlation breakdowns."),
        ("Governance Bypass", "Risk identification frameworks existed on paper but were "
         "circumvented by dominant individuals or cultural pressure."),
        ("Silo Thinking", "Risk identification conducted within business lines without "
         "aggregation, missing cross-portfolio and enterprise-level exposures."),
        ("Cultural Suppression", "Institutional culture that discouraged challenge, dissent, "
         "or escalation of risk concerns."),
        ("Emerging Risk Blindness", "Failure to identify new risk types (cyber, conduct, "
         "structured products) that fell outside traditional taxonomies."),
        ("Control Environment Failure", "Weak internal controls allowed fraud, unauthorized "
         "trading, or accounting manipulation to go undetected."),
        ("Information Asymmetry", "Key risk information held by front office or management "
         "but not shared with risk functions or the board."),
        ("Regulatory Arbitrage Masking", "Complex structures designed to optimize capital "
         "that simultaneously obscured the true risk profile."),
        ("Complacency", "Extended periods of low losses bred overconfidence and "
         "reduced vigilance in risk identification processes."),
    ]
    for mode, desc in failure_modes:
        lines.append(f"- **{mode}:** {desc}")

    lines.append("\n---\n*This database is maintained by EON Risk Services Ltd as a "
                 "foundational input to the Bank Risk Identification Process. "
                 "Loss figures are indicative and drawn from public sources. "
                 "USD standardisation uses approximate exchange rates for comparability only.*\n")

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))
    print(f"  MD:    {output_path}")


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    print("=" * 70)
    print("Industry Loss Database Builder — EON Risk Services")
    print("=" * 70)

    # Parse
    print("\n[1/6] Extracting text from PDF...")
    raw_text = extract_text(SOURCE_PDF)

    print("[2/6] Parsing entries...")
    entries = parse_entries(raw_text)
    print(f"  Parsed {len(entries)} entries")

    # Build
    print("[3/6] Building enriched database...")
    df = build_database(entries)

    # Validate
    print(f"\n  Validation:")
    print(f"    Rows: {len(df)}")
    for col in ['Financial Institution', 'Year', 'Risk Event', 'Country', 'Region',
                'Risk Taxonomy L1', 'Risk Taxonomy L2', 'COSO Objective Category']:
        nulls = df[col].isna().sum() + (df[col] == '').sum()
        status = "OK" if nulls == 0 else f"WARN: {nulls} empty"
        print(f"    {col}: {status}")

    priority_count = len(df[df['Include in Inventory'].str.upper() == 'YES'])
    print(f"    Priority examples: {priority_count}")

    unknown_countries = df[df['Country'] == 'Unknown']
    if len(unknown_countries) > 0:
        print(f"    WARNING: {len(unknown_countries)} entries with unknown country:")
        for _, r in unknown_countries.iterrows():
            print(f"      - {r['Financial Institution']}")

    # Output
    print("\n[4/6] Building Excel workbook...")
    build_excel(df, os.path.join(OUTPUT_DIR, "Industry_Loss_Database.xlsx"))

    print("[5/6] Exporting CSV...")
    export_csv(df, os.path.join(OUTPUT_DIR, "Industry_Loss_Database.csv"))

    print("[6/6] Generating Markdown summary...")
    generate_markdown(df, os.path.join(OUTPUT_DIR, "Industry_Loss_Database_Summary.md"))

    print("\n" + "=" * 70)
    print("COMPLETE")
    print("=" * 70)


if __name__ == "__main__":
    main()
